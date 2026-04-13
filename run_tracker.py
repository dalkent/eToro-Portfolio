#!/usr/bin/env python3
"""
run_tracker.py
──────────────
Daily portfolio briefing script.  Runs at 05:45 via Windows Task Scheduler.

Steps:
  1. Read eToro_Master.xlsx for live portfolio state
  2. Fetch current prices via yfinance
  3. Write Obsidian daily metrics note
  4. Create a Gmail draft to ndaley1313@gmail.com with the summary
     (requires one-time setup: run setup_gmail_auth.py first)

Usage:
    python run_tracker.py           # full run
    python run_tracker.py --demo    # demo data (no Excel / no Gmail)
"""

import os
import sys
import base64
import json
import logging
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
DATA_DIR    = BASE_DIR / "data"
MASTER      = DATA_DIR / "eToro_Master.xlsx"
LOG_DIR     = BASE_DIR / "logs"
TOKEN_FILE  = BASE_DIR / "gmail_token.json"
CREDS_FILE  = BASE_DIR / "gmail_credentials.json"

OBSIDIAN_DIR = Path(r"C:\Users\Neil\OneDrive\Obsidian\Daley's Brain\Projects\eToro & Investing\Reference")
RECIPIENT    = "ndaley1313@gmail.com"

DEMO = "--demo" in sys.argv

# ── Logging ────────────────────────────────────────────────────────────────────
LOG_DIR.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "tracker_run.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ── Excel loader ───────────────────────────────────────────────────────────────

def load_excel():
    import openpyxl
    log.info("Reading %s ...", MASTER)
    wb = openpyxl.load_workbook(str(MASTER), data_only=True)

    ws_a = wb["Assumptions"]
    gbpusd = 1.34
    for row in ws_a.iter_rows(min_row=3, max_row=6, values_only=True):
        if row[0] == "GBP/USD" and row[1]:
            try:
                gbpusd = float(row[1])
            except Exception:
                pass
            break

    ws_p = wb["Portfolio"]
    holdings = []
    cash = 0.0
    for row in ws_p.iter_rows(min_row=3, max_row=300, values_only=True):
        company = str(row[1] or "").strip()
        if not company:
            continue
        if company == "CASH":
            try:
                cash = float(row[10] or 0)
            except Exception:
                pass
            continue
        if "GRAND TOTAL" in company.upper():
            continue
        ticker   = str(row[3] or "").strip()
        yahoo    = str(row[4] or ticker).strip()
        sector   = str(row[2] or "Other").strip()
        currency = str(row[5] or "USD").strip()
        try:
            units    = float(row[8]  or 0)
            invested = float(row[10] or 0)
        except Exception:
            continue
        holdings.append({
            "company": company, "ticker": ticker, "yahoo": yahoo,
            "sector": sector, "currency": currency,
            "units": units, "invested": invested,
        })

    return holdings, cash, gbpusd


# ── Price fetch ────────────────────────────────────────────────────────────────

def fetch_prices(holdings):
    try:
        import yfinance as yf
    except ImportError:
        log.warning("yfinance not installed — skipping live prices")
        return {}

    YF_OVERRIDES = {"BTC": "BTC-USD", "Roku": "ROKU"}
    yahoo_set = list({YF_OVERRIDES.get(h["yahoo"], h["yahoo"]) for h in holdings if h["yahoo"]})
    if not yahoo_set:
        return {}

    log.info("Fetching %d prices from Yahoo Finance ...", len(yahoo_set))
    prices = {}
    for orig_yahoo in {h["yahoo"] for h in holdings if h["yahoo"]}:
        yf_t = YF_OVERRIDES.get(orig_yahoo, orig_yahoo)
        try:
            hist = yf.Ticker(yf_t).history(period="2d")
            if not hist.empty:
                prices[orig_yahoo] = float(hist["Close"].iloc[-1])
        except Exception:
            pass
    return prices


# ── Compute metrics ────────────────────────────────────────────────────────────

def enrich(holdings, cash, gbpusd, prices):
    total_value = cash
    total_invested = cash

    for h in holdings:
        raw = prices.get(h["yahoo"])
        if raw is not None:
            if h["currency"] == "GBp":
                price_usd = (raw / 100) * gbpusd
            else:
                price_usd = raw
            h["current_value"] = h["units"] * price_usd
        else:
            h["current_value"] = h["invested"]
        h["pnl"] = h["current_value"] - h["invested"]
        h["roi"] = (h["pnl"] / h["invested"] * 100) if h["invested"] else 0.0
        total_value    += h["current_value"]
        total_invested += h["invested"]

    total_pnl = total_value - total_invested
    total_roi = (total_pnl / total_invested * 100) if total_invested else 0.0
    return total_value, total_invested, total_pnl, total_roi


# ── Demo data ──────────────────────────────────────────────────────────────────

def demo_data():
    holdings = [
        {"company": "Invesco QQQ Trust", "ticker": "QQQ", "yahoo": "QQQ",
         "sector": "ETF", "currency": "USD", "units": 8.0, "invested": 3440.0,
         "current_value": 3560.0, "pnl": 120.0, "roi": 3.49},
        {"company": "Bitcoin", "ticker": "BTC", "yahoo": "BTC-USD",
         "sector": "Crypto", "currency": "USD", "units": 0.05, "invested": 3000.0,
         "current_value": 3400.0, "pnl": 400.0, "roi": 13.33},
        {"company": "Apple Inc.", "ticker": "AAPL", "yahoo": "AAPL",
         "sector": "Stocks", "currency": "USD", "units": 10.0, "invested": 1700.0,
         "current_value": 1855.0, "pnl": 155.0, "roi": 9.12},
        {"company": "Tesla Inc.", "ticker": "TSLA", "yahoo": "TSLA",
         "sector": "Stocks", "currency": "USD", "units": 5.0, "invested": 1250.0,
         "current_value": 1150.0, "pnl": -100.0, "roi": -8.00},
    ]
    cash = 4000.0
    total_value = sum(h["current_value"] for h in holdings) + cash
    total_invested = sum(h["invested"] for h in holdings) + cash
    total_pnl = total_value - total_invested
    total_roi = (total_pnl / total_invested * 100) if total_invested else 0.0
    return holdings, cash, total_value, total_invested, total_pnl, total_roi


# ── Build report strings ───────────────────────────────────────────────────────

def build_markdown(date_str, holdings, cash, total_value, total_invested, total_pnl, total_roi, is_demo):
    demo_tag = " *[demo]*" if is_demo else ""
    lines = [
        f"# Daily Portfolio Metrics — {date_str}{demo_tag}",
        "",
        f"*Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC*",
        "",
        "## Account Overview",
        "",
        "| Metric | Value |",
        "|--------|-------|",
        f"| Portfolio Value | **${total_value:,.2f}** |",
        f"| Cash Balance | ${cash:,.2f} |",
        f"| Total Invested | ${total_invested:,.2f} |",
        f"| Unrealised P&L | ${ '+' if total_pnl >= 0 else ''}{total_pnl:,.2f} |",
        f"| Total Return | **{'+' if total_roi >= 0 else ''}{total_roi:.2f}%** |",
        "",
        "## Open Positions",
        "",
        "| Ticker | Company | Sector | Units | Value | P&L | P&L% |",
        "|--------|---------|--------|------:|------:|----:|-----:|",
    ]
    for h in sorted(holdings, key=lambda x: -x["current_value"]):
        pnl_sign = "+" if h["pnl"] >= 0 else ""
        roi_sign = "+" if h["roi"] >= 0 else ""
        lines.append(
            f"| {h['ticker']} | {h['company']} | {h['sector']} "
            f"| {h['units']:.4f} | ${h['current_value']:,.2f} "
            f"| {pnl_sign}${h['pnl']:,.2f} | {roi_sign}{h['roi']:.2f}% |"
        )

    # Allocation by sector
    by_sector = {}
    for h in holdings:
        by_sector.setdefault(h["sector"], 0.0)
        by_sector[h["sector"]] += h["current_value"]
    portfolio_value = total_value - cash
    lines += [
        "",
        "## Allocation by Sector",
        "",
        "| Sector | Value | % |",
        "|--------|------:|--:|",
    ]
    for sector, val in sorted(by_sector.items(), key=lambda x: -x[1]):
        pct = (val / portfolio_value * 100) if portfolio_value else 0
        lines.append(f"| {sector} | ${val:,.2f} | {pct:.1f}% |")

    lines += [
        "",
        "---",
        "*Source: eToro Portfolio Tracker — [[eToro Portfolio Tracker – Valuation Methodology & Assumptions Guide]]*",
    ]
    return "\n".join(lines)


def build_email_html(date_str, holdings, cash, total_value, total_invested, total_pnl, total_roi, is_demo):
    demo_tag = " [demo]" if is_demo else ""
    pnl_color = "#10b981" if total_pnl >= 0 else "#ef4444"
    pnl_sign  = "+" if total_pnl >= 0 else ""
    roi_sign  = "+" if total_roi >= 0 else ""

    rows = ""
    for h in sorted(holdings, key=lambda x: -x["current_value"]):
        color = "#10b981" if h["pnl"] >= 0 else "#ef4444"
        ps    = "+" if h["pnl"] >= 0 else ""
        rs    = "+" if h["roi"] >= 0 else ""
        rows += (
            f"<tr><td>{h['ticker']}</td><td>{h['company']}</td>"
            f"<td style='text-align:right'>${h['current_value']:,.2f}</td>"
            f"<td style='text-align:right;color:{color}'>{ps}${h['pnl']:,.2f}</td>"
            f"<td style='text-align:right;color:{color}'>{rs}{h['roi']:.2f}%</td></tr>"
        )

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"/></head>
<body style="font-family:Arial,sans-serif;font-size:14px;color:#1f2937;max-width:640px;margin:0 auto">
<h2 style="color:#111827">eToro Daily Brief — {date_str}{demo_tag}</h2>
<table style="border-collapse:collapse;width:100%;margin-bottom:16px">
  <tr><td style="padding:4px 8px;font-weight:bold">Portfolio Value</td>
      <td style="padding:4px 8px;font-weight:bold">${total_value:,.2f}</td></tr>
  <tr style="background:#f9fafb"><td style="padding:4px 8px">Cash</td>
      <td style="padding:4px 8px">${cash:,.2f}</td></tr>
  <tr><td style="padding:4px 8px">Total Invested</td>
      <td style="padding:4px 8px">${total_invested:,.2f}</td></tr>
  <tr style="background:#f9fafb"><td style="padding:4px 8px">Unrealised P&amp;L</td>
      <td style="padding:4px 8px;color:{pnl_color}">{pnl_sign}${total_pnl:,.2f}</td></tr>
  <tr><td style="padding:4px 8px;font-weight:bold">Total Return</td>
      <td style="padding:4px 8px;font-weight:bold;color:{pnl_color}">{roi_sign}{total_roi:.2f}%</td></tr>
</table>

<h3 style="color:#374151">Open Positions</h3>
<table style="border-collapse:collapse;width:100%;font-size:13px">
  <thead><tr style="background:#e5e7eb">
    <th style="padding:4px 8px;text-align:left">Ticker</th>
    <th style="padding:4px 8px;text-align:left">Company</th>
    <th style="padding:4px 8px;text-align:right">Value</th>
    <th style="padding:4px 8px;text-align:right">P&amp;L</th>
    <th style="padding:4px 8px;text-align:right">P&amp;L%</th>
  </tr></thead>
  <tbody>{rows}</tbody>
</table>

<p style="font-size:11px;color:#9ca3af;margin-top:24px">
  Generated by eToro Portfolio Tracker · {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC
</p>
</body></html>"""


# ── Obsidian writer ────────────────────────────────────────────────────────────

def write_obsidian(date_str, md_content):
    OBSIDIAN_DIR.mkdir(parents=True, exist_ok=True)
    # Single static file — overwrites previous so only one copy exists in the vault
    path = OBSIDIAN_DIR / "Daily Portfolio Metrics.md"
    path.write_text(md_content, encoding="utf-8")
    log.info("Obsidian note written: %s", path)
    return path


# ── Gmail draft ────────────────────────────────────────────────────────────────

def create_gmail_draft(subject, html_body):
    """Create a Gmail draft using stored OAuth token. Returns True on success."""
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
    except ImportError:
        log.warning("google-api-python-client not installed — skipping Gmail draft")
        return False

    if not TOKEN_FILE.exists():
        log.warning(
            "Gmail token not found (%s). Run setup_gmail_auth.py first to enable drafts.",
            TOKEN_FILE,
        )
        return False

    try:
        creds = Credentials.from_authorized_user_file(
            str(TOKEN_FILE),
            scopes=["https://www.googleapis.com/auth/gmail.compose"],
        )
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
            TOKEN_FILE.write_text(creds.to_json(), encoding="utf-8")

        service = build("gmail", "v1", credentials=creds)

        msg = MIMEMultipart("alternative")
        msg["To"]      = RECIPIENT
        msg["Subject"] = subject
        msg.attach(MIMEText(html_body, "html"))

        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        service.users().drafts().create(
            userId="me",
            body={"message": {"raw": raw}},
        ).execute()
        log.info("Gmail draft created for %s", RECIPIENT)
        return True
    except Exception as exc:
        log.error("Gmail draft failed: %s", exc)
        return False


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    today = datetime.now()
    date_str = today.strftime("%Y-%m-%d")
    log.info("=== run_tracker.py starting (%s) ===", date_str)

    if DEMO:
        log.info("Running in demo mode")
        holdings, cash, total_value, total_invested, total_pnl, total_roi = demo_data()
    else:
        try:
            raw_holdings, cash, gbpusd = load_excel()
            prices = fetch_prices(raw_holdings)
            total_value, total_invested, total_pnl, total_roi = enrich(
                raw_holdings, cash, gbpusd, prices
            )
            holdings = raw_holdings
        except Exception as exc:
            log.error("Excel load failed: %s — falling back to demo data", exc)
            holdings, cash, total_value, total_invested, total_pnl, total_roi = demo_data()
            DEMO_FLAG = True

    md  = build_markdown(date_str, holdings, cash, total_value, total_invested, total_pnl, total_roi, DEMO)
    html = build_email_html(date_str, holdings, cash, total_value, total_invested, total_pnl, total_roi, DEMO)

    write_obsidian(date_str, md)

    subject = f"eToro Daily Brief — {date_str}"
    ok = create_gmail_draft(subject, html)
    if not ok:
        # Save draft locally as fallback
        fallback = LOG_DIR / f"gmail_draft_{date_str}.html"
        fallback.write_text(html, encoding="utf-8")
        log.info("Draft saved locally (Gmail not configured): %s", fallback)

    log.info("=== run_tracker.py complete ===")


if __name__ == "__main__":
    main()
