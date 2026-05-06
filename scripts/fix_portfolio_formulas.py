"""Repair Portfolio sheet rows where the live-price and related formulas use
the broken bare-ticker form ('=...stockhistory("HOOD",...)') instead of the
indirection form that goes through Tickers!M (which produces XNAS:HOOD,
XLON:KGF, etc.).

Background
----------
sync_portfolio.py's _add_to_portfolio() writes formulas with a bare Yahoo
ticker hardcoded into the STOCKHISTORY call. Excel STOCKHISTORY only accepts
that format for symbols with a recognised exchange suffix (.L for LSE). For
NASDAQ/NYSE tickers like HOOD, CRSP, PLTR it returns #VALUE! and the IFERROR
falls through to "N/A" or zero.

All older rows in the Portfolio sheet use a corrected indirection form:
  =IFERROR(INDEX(_xlfn.STOCKHISTORY(
        INDEX(Tickers!$M:$M, MATCH(D{r}, Tickers!$D:$D, 0)),
        TODAY()-7, TODAY(), 0, 0
      ), 1, 2),
    IFERROR(INDEX(Tickers!$N:$N, MATCH(D{r}, Tickers!$D:$D, 0)), ""))

Tickers!$M is a per-row formula that produces XNAS:TICKER for NASDAQ,
XNYS:TICKER for NYSE, XLON:TICKER for FTSE, etc.

This script also rebuilds the columns that newer rows had off-by-one or
missing entirely (col 29 'Live Price GBP/USD', col 35 'Range Position',
col 36 'Yesterday Close', etc.) so they match older rows.

Usage
-----
    python scripts/fix_portfolio_formulas.py            # dry-run
    python scripts/fix_portfolio_formulas.py --apply
    python scripts/fix_portfolio_formulas.py --apply --backup
"""
from __future__ import annotations
import argparse
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import _envloader  # noqa: F401

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
MASTER = BASE_DIR / "data" / "eToro_Master.xlsx"

# Portfolio col layout (1-indexed):
# A=#  B=Company  C=Sector  D=eToro Ticker  E=Yahoo Ticker  F=Currency
# G=eToro ID  H=Trades  I=Units Held  J=Avg Buy  K=Invested(USD)
# L=Live Price (Local)  M=Price (GBP/USD)  N=Current Value (USD)  O=Weight
# P=P&L (USD)  Q=ROI %  R=Div 2023  S=Div 2024  T=Div 2025  U=Div 2026E
# V=Total Divs  W=Div Return %  X=ROI no Divs  Y=ROI with Divs
# (col 26 reserved)
# AA=Daily Chg  AB=Target (GBP/USD)  AC=Live Price (GBP/USD)
# AD=Value Ratio  AE=Signal  AF=Date Checked
# AG=52W High  AH=52W Low  AI=Range Position  AJ=Yesterday Close


# Pattern that identifies a broken row: bare ticker hardcoded in stockhistory()
BARE_STOCKHISTORY = re.compile(r'stockhistory\("([A-Z][A-Z0-9.\-]*)"\s*,', re.IGNORECASE)


def looks_broken(formula):
    """Return True if the formula uses the bare-ticker stockhistory pattern."""
    if not isinstance(formula, str):
        return False
    if BARE_STOCKHISTORY.search(formula):
        return True
    return False


def correct_formulas(r):
    """Return the dict {col_idx: formula_string} for the indirection forms
    we want every row to have. r is the 1-indexed row number."""
    # Indirection lookups
    M_LOOKUP = f'INDEX(Tickers!$M:$M,MATCH(D{r},Tickers!$D:$D,0))'
    N_LOOKUP = f'INDEX(Tickers!$N:$N,MATCH(D{r},Tickers!$D:$D,0))'

    return {
        # L: Live Price (Local)
        12: (
            f'=IFERROR(INDEX(_xlfn.STOCKHISTORY({M_LOOKUP},TODAY()-7,TODAY(),0,0),1,2),'
            f'IFERROR({N_LOOKUP},""))'
        ),
        # AC (29): Live Price (GBP/USD) - mirrors M
        29: f'=M{r}',
        # AD (30): Value Ratio
        30: f'=IFERROR(IF(AB{r}="N/A","N/A",AB{r}/M{r}),"N/A")',
        # AE (31): Signal
        31: (
            f'=IF(OR(AD{r}="N/A",AD{r}=""),"-",'
            f'IF(AD{r}>=1.25,"Strong Buy",IF(AD{r}>=1.1,"Buy",'
            f'IF(AD{r}>=0.9,"Fair Value",IF(AD{r}>=0.75,"Sell","Strong Sell")))))'
        ),
        # AG (33): 52W High
        33: (
            f'=IFERROR(MAX(INDEX(_xludf.stockhistory({M_LOOKUP},'
            f'DATE(YEAR(TODAY())-1,MONTH(TODAY()),DAY(TODAY())),TODAY(),1,0),0,2)),"N/A")'
        ),
        # AH (34): 52W Low
        34: (
            f'=IFERROR(MIN(INDEX(_xludf.stockhistory({M_LOOKUP},'
            f'DATE(YEAR(TODAY())-1,MONTH(TODAY()),DAY(TODAY())),TODAY(),1,0),0,2)),"N/A")'
        ),
        # AI (35): Range Position
        35: f'=IFERROR((L{r}-AH{r})/(AG{r}-AH{r}),"N/A")',
        # AJ (36): Yesterday Close
        36: (
            f'=IFERROR(INDEX(_xlfn.STOCKHISTORY({M_LOOKUP},'
            f'WORKDAY(TODAY(),-1),WORKDAY(TODAY(),-1),0,0),1,2),"N/A")'
        ),
    }


def find_broken_rows(ws_p):
    """Yield row numbers where col L (Live Price) uses the bare-ticker pattern."""
    for r in range(3, 200):
        ticker = ws_p.cell(row=r, column=4).value
        if not ticker or not isinstance(ticker, str):
            continue
        if str(ticker).strip().upper() in ("CASH", "TOTAL", "GRAND TOTAL", "GRAND TOTAL (INCL. CASH)"):
            continue
        L = ws_p.cell(row=r, column=12).value
        if looks_broken(L):
            yield r, str(ticker)


def fix_portfolio_formulas(wb, *, log=print):
    """Returns (rows_fixed, fields_fixed)."""
    if "Portfolio" not in wb.sheetnames:
        log("  No Portfolio sheet, skipping")
        return (0, 0)
    ws_p = wb["Portfolio"]

    rows_to_fix = list(find_broken_rows(ws_p))
    if not rows_to_fix:
        return (0, 0)

    rows_fixed = 0
    fields_fixed = 0
    for r, ticker in rows_to_fix:
        formulas = correct_formulas(r)
        for col, new_formula in formulas.items():
            old = ws_p.cell(row=r, column=col).value
            if old != new_formula:
                ws_p.cell(row=r, column=col, value=new_formula)
                fields_fixed += 1
        rows_fixed += 1
        log(f"  Portfolio row {r:3d} ({ticker}): rewrote price/range/signal formulas (8 fields)")

    return (rows_fixed, fields_fixed)


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--apply", action="store_true", help="Write changes back. Without this, runs as a dry-run.")
    p.add_argument("--backup", action="store_true", help="When applying, also write a timestamped backup.")
    p.add_argument("--master", default=str(MASTER), help="Path to eToro_Master.xlsx")
    args = p.parse_args()

    master = Path(args.master)
    if not master.exists():
        print(f"ERROR: {master} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {master} ...")
    wb = openpyxl.load_workbook(master)

    rows_fixed, fields_fixed = fix_portfolio_formulas(wb)

    if not rows_fixed:
        print("\nNo broken rows found.")
        return

    if args.apply:
        if args.backup:
            stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            backup = master.with_suffix(f".bak-{stamp}.xlsx")
            shutil.copy2(master, backup)
            print(f"\nWrote backup: {backup}")
        wb.save(master)
        print(f"Saved. Fixed {rows_fixed} row(s), {fields_fixed} field(s) overwritten.")
    else:
        print(f"\nDRY-RUN: would fix {rows_fixed} row(s), {fields_fixed} field(s). Re-run with --apply to commit.")


if __name__ == "__main__":
    main()
