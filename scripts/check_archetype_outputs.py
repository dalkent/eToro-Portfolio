#!/usr/bin/env python3
"""
check_archetype_outputs.py
──────────────────────────
Reads the Assumptions sheet of eToro_Master.xlsx and reports the Method label
and blended target for each of the 20 portfolio holdings. Use this to verify
whether the archetype framework fired correctly after a valuation.py run.

Usage:
  python scripts/check_archetype_outputs.py
"""
from pathlib import Path
import openpyxl

BASE_DIR = Path(__file__).parent.parent
MASTER   = BASE_DIR / "data" / "eToro_Master.xlsx"

PORTFOLIO = {
    "RKT.L":  "Mature compounder",
    "DGE.L":  "Mature compounder",
    "BATS.L": "Mature compounder",
    "KGF.L":  "Mature cyclical",
    "JD.L":   "Mature cyclical",
    "MKS.L":  "Mature cyclical",
    "INCH.L": "Mature cyclical",
    "SBRY.L": "Mature cyclical",
    "EZJ.L":  "Mature cyclical",
    "VSVS.L": "Mature cyclical",
    "NG.L":   "Yield anchor",
    "UU.L":   "Yield anchor",
    "IMB.L":  "Yield anchor",
    "TATE.L": "Yield anchor",
    "BP.L":   "Capex cyclical",
    "SHEL.L": "Capex cyclical",
    "ENOG.L": "Capex cyclical",
    "HBR.L":  "Capex cyclical",
    "VOD.L":  "Restructuring",
    "NWG.L":  "(Bank logic — no archetype)",
}


def main():
    wb = openpyxl.load_workbook(str(MASTER), data_only=True)
    ws_a = wb["Assumptions"]
    ws_t = wb["Tickers"]

    # Step 1: confirm Sub-Sector values are present on the Tickers sheet
    print("=" * 90)
    print("Tickers sheet — Sub-Sector column (col O) check")
    print("=" * 90)
    print(f"{'Ticker':<8} {'Expected archetype':<32} {'Actual Sub-Sector':<32}")
    print("-" * 90)
    sub_sector_by_ticker = {}
    for row in ws_t.iter_rows(min_row=2, max_row=ws_t.max_row, values_only=True):
        if len(row) < 6 or not row[5]:
            continue
        yf_t = str(row[5]).strip()
        if yf_t in PORTFOLIO:
            ss = str(row[14] or "").strip() if len(row) > 14 else ""
            sub_sector_by_ticker[yf_t] = ss
    for ticker, expected in PORTFOLIO.items():
        actual = sub_sector_by_ticker.get(ticker, "(not found)")
        ok = "OK" if (actual == expected or "Bank" in expected) else "MISMATCH"
        print(f"{ticker:<8} {expected:<32} {actual:<32} {ok}")

    # Step 2: Method labels + blended targets on the Assumptions sheet
    print()
    print("=" * 90)
    print("Assumptions sheet — Method + blended target check (portfolio)")
    print("=" * 90)
    print(f"{'Ticker':<8} {'Blended target':<18} {'Method':<60}")
    print("-" * 90)
    for row in ws_a.iter_rows(min_row=7, max_row=300, values_only=True):
        ticker = row[0]
        if ticker not in PORTFOLIO:
            continue
        blended = row[11]
        method = str(row[12] or "").strip()
        blended_p = f"{blended * 100:.1f}p" if blended else "(none)"
        print(f"{ticker:<8} {blended_p:<18} {method:<60}")

    # Step 3: Sector-default summary across the full non-financial universe
    print()
    print("=" * 90)
    print("Sector-default fan-out — how many tickers picked up each archetype path")
    print("=" * 90)
    from collections import Counter
    archetype_counter = Counter()
    sector_default_count = 0
    explicit_count = 0
    legacy_count = 0
    examples = {}
    for row in ws_a.iter_rows(min_row=7, max_row=300, values_only=True):
        ticker = row[0]
        if not ticker or not isinstance(ticker, str):
            continue
        method = str(row[12] or "").strip()
        if "Archetype:" in method:
            # Strip the parameter list to get just the archetype name
            archetype_name = method.split("(")[0].replace("Archetype:", "")
            archetype_counter[archetype_name] += 1
            if "sector-default" in method:
                sector_default_count += 1
            else:
                explicit_count += 1
            examples.setdefault(archetype_name, []).append(ticker)
        elif method in ("primary", "fallback", "skip", "No Valuation", "Analyst Consensus"):
            legacy_count += 1
    for arch, count in archetype_counter.most_common():
        sample = ", ".join(examples[arch][:5])
        more = f" + {count - 5} more" if count > 5 else ""
        print(f"  {arch:<22} {count:>3} tickers   ({sample}{more})")
    print()
    print(f"  Total archetype-routed: {sum(archetype_counter.values())}")
    print(f"    of which explicit Sub-Sector: {explicit_count}")
    print(f"    of which sector-default:      {sector_default_count}")
    print(f"  Legacy winsorise/financial routes: {legacy_count}")

    # Step 4: extreme-signal review — Strong Buy / Strong Sell names routed via
    # sector-default, where the default's correctness matters most for what
    # actually publishes in the Tuesday tracker.
    print()
    print("=" * 90)
    print("Extreme signals routed via sector-default (review for misclassification)")
    print("=" * 90)
    print(f"{'Ticker':<10} {'Sector':<25} {'Archetype':<22} {'Target':<10} {'Price':<10} {'VR':<6} {'Signal'}")
    print("-" * 100)
    # Need price from Tickers sheet col N (Manual Price Override) for VR calc
    price_by_ticker = {}
    for row in ws_t.iter_rows(min_row=2, max_row=ws_t.max_row, values_only=True):
        if len(row) < 14 or not row[5]:
            continue
        yf_t = str(row[5]).strip()
        price = row[13] if len(row) > 13 else None
        if price:
            try:
                price_by_ticker[yf_t] = float(price)
            except (ValueError, TypeError):
                pass
    extremes = []
    for row in ws_a.iter_rows(min_row=7, max_row=300, values_only=True):
        ticker = row[0]
        if not ticker or not isinstance(ticker, str):
            continue
        method = str(row[12] or "").strip()
        if "sector-default" not in method:
            continue
        blended_gbp = row[11]
        if not blended_gbp:
            continue
        price_p = price_by_ticker.get(ticker)
        if not price_p:
            continue
        target_p = float(blended_gbp) * 100
        vr = target_p / price_p
        if vr >= 1.25:
            signal = "Strong Buy"
        elif vr < 0.75:
            signal = "Strong Sell"
        else:
            continue  # not extreme
        archetype = method.split("(")[0].replace("Archetype:", "")
        sector_a = str(row[2] or "").strip()
        extremes.append((ticker, sector_a, archetype, target_p, price_p, vr, signal))
    extremes.sort(key=lambda x: x[5])  # sort by VR ascending (Strong Sells first)
    for ticker, sec, arch, tgt, pr, vr, sig in extremes:
        print(f"{ticker:<10} {sec:<25} {arch:<22} {tgt:>7.0f}p   {pr:>7.0f}p   {vr:.2f}   {sig}")
    if not extremes:
        print("  (none)")


if __name__ == "__main__":
    main()
