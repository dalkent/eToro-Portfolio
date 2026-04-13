#!/usr/bin/env python3
"""
sync_portfolio.py
─────────────────
Run this script whenever you make a trade on eToro (buy, sell, or partial sell).

It will:
  1. Call the eToro API for your current open positions
  2. Compare against the Portfolio sheet in eToro_Master.xlsx
  3. Handle changes:
       Full sell    → remove from Portfolio, add to Closed Positions,
                      move all dividends, add to Watchlist
       Partial sell → add Closed Positions row for units sold,
                      move proportional dividends, update Portfolio units
       New buy      → add row to Portfolio, remove from Watchlist
  4. Update Tickers sheet (In Portfolio / In Watchlist flags)

Usage:
  python sync_portfolio.py

Requirements:
  Environment variables:
    ETORO_PUBLIC_API_KEY
    ETORO_USER_KEY
"""

import os
import sys
import csv
import re
import uuid
import requests
import yfinance as yf
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent.parent
DATA_DIR   = BASE_DIR / "data"
LOGS_DIR   = BASE_DIR / "logs"
MASTER     = DATA_DIR / "eToro_Master.xlsx"
MATCH_CSV  = DATA_DIR / "etoro_portfolio_tickermatch.csv"
LOG_FILE   = LOGS_DIR / "sync_portfolio.log"

LOGS_DIR.mkdir(exist_ok=True)

# ── eToro API ─────────────────────────────────────────────────────────────────
PUBLIC_API_KEY = os.getenv("ETORO_PUBLIC_API_KEY")
USER_KEY       = os.getenv("ETORO_USER_KEY")
BASE_URL       = "https://public-api.etoro.com/api/v1"

def etoro_headers():
    return {
        "x-api-key":    PUBLIC_API_KEY,
        "x-user-key":   USER_KEY,
        "x-request-id": str(uuid.uuid4()),
        "Accept":       "application/json",
    }

# ── Logging ───────────────────────────────────────────────────────────────────
def log(msg: str):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")
    print(line)

# ── Styling helpers ───────────────────────────────────────────────────────────
INPUT_BLUE  = Font(color="FF0000FF", name="Arial", size=10)
HEADER_FONT = Font(bold=True, name="Arial", size=10)
NORMAL_FONT = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
RIGHT       = Alignment(horizontal="right",  vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")

def _set(cell, value, fmt=None, font=None, align=None):
    cell.value = value
    if font:  cell.font = font
    if align: cell.alignment = align
    if fmt:   cell.number_format = fmt

# ── Load ticker mapping (Asset ID → ticker) ───────────────────────────────────
def load_mapping() -> dict:
    """Returns {asset_id (int): ticker (str)}"""
    mapping = {}
    if not MATCH_CSV.exists():
        log(f"Warning: {MATCH_CSV} not found — will use Asset IDs as tickers")
        return mapping
    with open(MATCH_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            try:
                mapping[int(row["Asset_ID"])] = row["Ticker"].strip()
            except (KeyError, ValueError):
                continue
    log(f"Loaded {len(mapping)} ticker mappings")
    return mapping

# ── Fetch current open positions from eToro ───────────────────────────────────
def fetch_open_positions() -> dict:
    """
    Returns {ticker: {units, invested_usd, current_price, trades, asset_id}}
    """
    if not PUBLIC_API_KEY or not USER_KEY:
        log("ERROR: ETORO_PUBLIC_API_KEY / ETORO_USER_KEY not set in environment")
        sys.exit(1)

    url = f"{BASE_URL}/trading/info/real/pnl"
    resp = requests.get(url, headers=etoro_headers(), timeout=15)
    resp.raise_for_status()
    data = resp.json()

    port      = data.get("clientPortfolio", {})
    positions = port.get("positions", [])
    cash      = port.get("credit", 0)
    log(f"eToro API: {len(positions)} open positions, cash ${cash:,.2f}")

    mapping = load_mapping()

    from collections import defaultdict
    grouped = defaultdict(lambda: {"units": 0, "invested_usd": 0, "current_price": None, "trades": 0, "asset_id": None})

    for pos in positions:
        iid = pos.get("instrumentID")
        if not isinstance(iid, int):
            continue
        ticker = mapping.get(iid, f"ID_{iid}")
        u = pos.get("unrealizedPnL", {})
        grouped[ticker]["units"]        += pos.get("units", 0)
        grouped[ticker]["invested_usd"] += pos.get("amount", 0)
        grouped[ticker]["trades"]       += 1
        grouped[ticker]["asset_id"]      = iid
        if "closeRate" in u:
            grouped[ticker]["current_price"] = u["closeRate"]

    # Add cash
    grouped["CASH"] = {"units": cash, "invested_usd": cash, "current_price": 1.0, "trades": 1, "asset_id": "CASH"}

    return dict(grouped)


# ── Try to fetch closed trade details from eToro API ─────────────────────────
def fetch_closed_trade_price(ticker: str, asset_id) -> tuple:
    """
    Attempt to get sale price and date from eToro closed trades API.
    Returns (sale_price, sale_date_str) or (None, None) if unavailable.
    """
    try:
        url = f"{BASE_URL}/trading/history/positions"
        resp = requests.get(url, headers=etoro_headers(), timeout=10,
                            params={"PageSize": 50, "PageNumber": 1})
        if resp.status_code != 200:
            return None, None
        trades = resp.json().get("PublicHistoryPositions", [])
        # Find most recent closed trade matching asset_id
        for t in trades:
            if t.get("InstrumentID") == asset_id:
                price = t.get("CloseRate") or t.get("closeRate")
                date  = t.get("CloseDateTime") or t.get("ClosedDate", "")
                if date:
                    try:
                        date = datetime.fromisoformat(date[:10]).strftime("%Y-%m-%d")
                    except Exception:
                        pass
                return price, date
    except Exception as e:
        log(f"  Closed trades API unavailable ({e}) — sale price will need manual entry")
    return None, None


# ── Resolve "ID_XXXX" tickers to real tickers via Tickers sheet col E ─────────
def _resolve_id_tickers(etoro_positions: dict, ws_t) -> None:
    """
    For any position key that is still "ID_NNNN" (not mapped via CSV),
    look up NNNN in Tickers col E (eToro ID) and replace the key with
    the real ticker (col D).  Mutates etoro_positions in place.
    """
    unresolved = [t for t in list(etoro_positions.keys()) if t.startswith("ID_")]
    if not unresolved:
        return

    # Build a map: str(etoro_id) → eToro ticker from Tickers sheet
    id_to_ticker = {}
    for row in ws_t.iter_rows(min_row=3, max_row=300):
        cell_id     = row[4].value   # col E = eToro ID
        cell_ticker = row[3].value   # col D = eToro Ticker
        if cell_id is not None and cell_ticker:
            try:
                id_to_ticker[str(int(float(cell_id)))] = str(cell_ticker).strip()
            except (ValueError, TypeError):
                pass

    for id_key in unresolved:
        numeric = id_key[3:]   # strip "ID_"
        real_ticker = id_to_ticker.get(numeric)
        if real_ticker:
            etoro_positions[real_ticker] = etoro_positions.pop(id_key)
            etoro_positions[real_ticker]["asset_id"] = int(numeric)
            log(f"  Resolved {id_key} → {real_ticker} via Tickers sheet")
        else:
            log(f"  Warning: {id_key} not found in Tickers sheet (add it to etoro_portfolio_tickermatch.csv)")


# ── Read Portfolio sheet into a dict ─────────────────────────────────────────
def read_portfolio_sheet(ws) -> dict:
    """
    Returns {ticker: {row_num, units, invested, avg_buy, currency, sector,
                       company, etoro_id, trades, div_2023, div_2024, div_2025, div_2026}}
    Row 2 = headers, data from row 3 onward.
    """
    portfolio = {}
    for row in ws.iter_rows(min_row=3, max_row=200):
        ticker = row[3].value   # D = eToro Ticker
        if not ticker or str(ticker).strip().upper() in ("CASH", "TOTAL", ""):
            continue
        ticker = str(ticker).strip()
        portfolio[ticker] = {
            "row_num":  row[0].row,
            "seq_num":  row[0].value,           # A
            "company":  row[1].value,            # B
            "sector":   row[2].value,            # C
            "etoro_id": row[6].value,            # G
            "trades":   row[7].value,            # H
            "units":    row[8].value,            # I
            "avg_buy":  row[9].value,            # J
            "invested": row[10].value,           # K
            "currency": row[5].value,            # F
            "div_2023": row[17].value or 0,      # R
            "div_2024": row[18].value or 0,      # S
            "div_2025": row[19].value or 0,      # T
            "div_2026": row[20].value or 0,      # U
        }
    return portfolio


# ── Next available row in Closed Positions ────────────────────────────────────
def next_closed_row(ws_c) -> int:
    last = 2
    for row in ws_c.iter_rows(min_row=3, max_row=500):
        if row[1].value:  # B = Ticker
            last = row[0].row
    return last + 1


# ── Add a row to Closed Positions ─────────────────────────────────────────────
def add_closed_position(ws_c, seq_num, ticker, sector, units_sold,
                        invested_usd, avg_buy_local, sale_price_local,
                        sale_value_usd, div_2023, div_2024, div_2025,
                        date_sold, notes, currency):
    r = next_closed_row(ws_c)
    fmt_price = "£#,##0.00" if currency == "GBp" else "$#,##0.00"

    data = [
        (1,  seq_num,          None,            LEFT),
        (2,  ticker,           None,            LEFT),
        (3,  sector or "",     None,            LEFT),
        (4,  round(units_sold, 6), "#,##0.0000", RIGHT),
        (5,  round(invested_usd, 2), "$#,##0.00", RIGHT),
        (6,  avg_buy_local,    fmt_price,       RIGHT),
        (7,  sale_price_local, fmt_price,       RIGHT),
        (8,  round(sale_value_usd, 2) if sale_value_usd else None, "$#,##0.00", RIGHT),
        # col I = Capital ROI formula
        (10, round(div_2023 or 0, 2), "$#,##0.00", RIGHT),
        (11, round(div_2024 or 0, 2), "$#,##0.00", RIGHT),
        (12, round(div_2025 or 0, 2), "$#,##0.00", RIGHT),
        # col M = SUM formula, N = div return, O = total ROI
        (16, date_sold,        "YYYY-MM-DD",    CENTER),
        (17, notes or "",      None,            LEFT),
    ]

    for col_1based, value, fmt, align in data:
        c = ws_c.cell(row=r, column=col_1based, value=value)
        c.font  = INPUT_BLUE
        c.alignment = align
        if fmt: c.number_format = fmt

    # Formulas
    ws_c.cell(row=r, column=9,  value=f'=IFERROR((H{r}-E{r})/E{r},"N/A")')
    ws_c.cell(row=r, column=13, value=f'=SUM(J{r}:L{r})')
    ws_c.cell(row=r, column=14, value=f'=IFERROR(M{r}/E{r},0)')
    ws_c.cell(row=r, column=15, value=f'=IFERROR((H{r}-E{r}+M{r})/E{r},"N/A")')
    for col in (9, 13, 14, 15):
        ws_c.cell(row=r, column=col).font = NORMAL_FONT

    log(f"  → Closed Positions row {r}: {ticker}  {round(units_sold,4)} units  "
        f"sale_price={sale_price_local}  invested=${invested_usd:.2f}")
    return r


# ── Remove a row from Portfolio sheet ────────────────────────────────────────
def remove_portfolio_row(ws_p, row_num):
    ws_p.delete_rows(row_num, 1)
    log(f"  → Removed Portfolio row {row_num}")
    # Repair formula row references — openpyxl doesn't update them on delete_rows
    _repair_portfolio_formulas(ws_p)
    # Re-number sequence column A
    seq = 1
    for row in ws_p.iter_rows(min_row=3, max_row=500):
        if row[3].value and str(row[3].value).strip().upper() not in ("CASH","TOTAL",""):
            row[0].value = seq
            seq += 1


def _repair_portfolio_formulas(ws_p):
    """Fix formula row references after a row deletion (openpyxl doesn't adjust them)."""
    # Dynamically find the CASH row so we know where data ends.
    # Only repair data rows — summary rows (CASH, GRAND TOTAL) are handled
    # by _fix_portfolio_summary_rows and must not be touched here.
    cash_row = None
    for r in ws_p.iter_rows(min_row=3, max_row=300):
        if str(r[1].value or "").strip() == "CASH":
            cash_row = r[0].row
            break
    max_data_row = (cash_row - 1) if cash_row else 42  # fallback to old behaviour

    for row in ws_p.iter_rows(min_row=3, max_row=max_data_row):
        actual_row = row[0].row
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula = cell.value
                def replacer(m, ar=actual_row):
                    col = m.group(1)
                    row_ref = int(m.group(2))
                    # Only fix self-referential row numbers (shifted by 1 after deletion)
                    if row_ref == ar + 1:
                        return col + str(ar)
                    return m.group(0)
                new_formula = re.sub(r'(?<!\$)([A-Z]+)(\d+)(?!\$)', replacer, formula)
                if new_formula != formula:
                    cell.value = new_formula
    # Fix summary rows after the data
    _fix_portfolio_summary_rows(ws_p)


def _fix_portfolio_summary_rows(ws_p):
    """
    After any row insertion or deletion, fix:
      - Weight formula (col O) in all data rows: update $N$<gt_row>
      - GRAND TOTAL row: fix circular/stale references
    """
    cash_row = None
    gt_row   = None
    for row in ws_p.iter_rows(min_row=3, max_row=300):
        b = str(row[1].value or "").strip()
        if b == "CASH":
            cash_row = row[0].row
        if "GRAND TOTAL" in b.upper():
            gt_row = row[0].row
        if cash_row and gt_row:
            break

    if not cash_row or not gt_row:
        return

    last_data = cash_row - 1

    # Fix weight formula in data rows
    for r in range(3, last_data + 1):
        cell = ws_p.cell(row=r, column=15)
        v = cell.value
        if isinstance(v, str) and v.startswith("="):
            new_v = re.sub(r"\$?N\$\d+", f"$N${gt_row}", v)
            if new_v != v:
                cell.value = new_v

    # Fix GRAND TOTAL row formulas
    ws_p.cell(row=gt_row, column=11).value = f"=SUM(K3:K{cash_row})"
    ws_p.cell(row=gt_row, column=14).value = f"=SUM(N3:N{cash_row})"
    ws_p.cell(row=gt_row, column=16).value = f"=P{cash_row}"
    ws_p.cell(row=gt_row, column=17).value = f'=IFERROR(P{gt_row}/K{gt_row},"N/A")'
    ws_p.cell(row=gt_row, column=22).value = f"=V{cash_row}"
    ws_p.cell(row=gt_row, column=23).value = f"=IFERROR(V{gt_row}/K{gt_row},0)"
    ws_p.cell(row=gt_row, column=24).value = f'=IFERROR(P{gt_row}/K{gt_row},"N/A")'
    ws_p.cell(row=gt_row, column=25).value = f'=IFERROR((P{gt_row}+V{gt_row})/K{gt_row},"N/A")'


# ── Update Portfolio units / invested / dividends after partial sell ──────────
def update_portfolio_partial(ws_p, row_num, new_units, new_invested,
                              new_div_2023, new_div_2024, new_div_2025, new_div_2026):
    ws_p.cell(row=row_num, column=9,  value=round(new_units, 6))     # I
    ws_p.cell(row=row_num, column=11, value=round(new_invested, 2))  # K
    ws_p.cell(row=row_num, column=18, value=round(new_div_2023, 2))  # R
    ws_p.cell(row=row_num, column=19, value=round(new_div_2024, 2))  # S
    ws_p.cell(row=row_num, column=20, value=round(new_div_2025, 2))  # T
    ws_p.cell(row=row_num, column=21, value=round(new_div_2026, 2))  # U
    log(f"  → Updated Portfolio row {row_num}: {new_units:.4f} units remaining, "
        f"invested=${new_invested:.2f}")


# ── Update Tickers sheet flags ────────────────────────────────────────────────
def update_tickers_flags(ws_t, ticker: str, in_portfolio: str, in_watchlist: str):
    for row in ws_t.iter_rows(min_row=3, max_row=300):
        t = row[3].value  # D = eToro Ticker
        if t and str(t).strip() == ticker:
            row[9].value  = in_portfolio   # J
            row[10].value = in_watchlist   # K
            log(f"  → Tickers: {ticker}  In Portfolio={in_portfolio}  In Watchlist={in_watchlist}")
            return
    log(f"  → Warning: {ticker} not found in Tickers sheet")


# ── Get current market price for a ticker ────────────────────────────────────
def get_market_price(ticker: str) -> float | None:
    try:
        info = yf.Ticker(ticker).fast_info
        return getattr(info, "last_price", None) or getattr(info, "regularMarketPrice", None)
    except Exception:
        return None


# ── Find last row in Portfolio (before CASH / totals) ────────────────────────
def last_portfolio_data_row(ws_p) -> int:
    last = 2
    for row in ws_p.iter_rows(min_row=3, max_row=200):
        if row[3].value and str(row[3].value).strip().upper() not in ("CASH","TOTAL",""):
            last = row[0].row
    return last


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN SYNC LOGIC
# ═══════════════════════════════════════════════════════════════════════════════
def run_sync():
    log("=" * 60)
    log("sync_portfolio.py  starting")
    log("=" * 60)

    # ── 1. Fetch live eToro positions ─────────────────────────────────────────
    etoro_positions = fetch_open_positions()
    etoro_tickers   = set(etoro_positions.keys())

    # ── 2. Load workbook ──────────────────────────────────────────────────────
    if not MASTER.exists():
        log(f"ERROR: {MASTER} not found")
        sys.exit(1)

    wb    = openpyxl.load_workbook(str(MASTER))
    ws_p  = wb["Portfolio"]
    ws_c  = wb["Closed Positions"]
    ws_w  = wb["Watchlist"]
    ws_t  = wb["Tickers"]

    # ── 2b. Resolve any unmatched "ID_XXXX" tickers via Tickers sheet ────────
    _resolve_id_tickers(etoro_positions, ws_t)
    etoro_tickers = set(etoro_positions.keys())   # refresh after resolution

    # ── 3. Read current Portfolio sheet ──────────────────────────────────────
    current_portfolio = read_portfolio_sheet(ws_p)
    portfolio_tickers = set(current_portfolio.keys())

    log(f"Portfolio sheet: {len(portfolio_tickers)} holdings")
    log(f"eToro open:      {len(etoro_tickers) - 1} holdings + cash")

    # ── 4. Detect changes ────────────────────────────────────────────────────
    fully_sold   = portfolio_tickers - etoro_tickers - {"CASH"}
    new_buys     = (etoro_tickers - portfolio_tickers) - {"CASH"}
    still_held   = portfolio_tickers & etoro_tickers

    log(f"Fully sold:    {fully_sold or 'none'}")
    log(f"New buys:      {new_buys or 'none'}")
    log(f"Still held:    {len(still_held)}")

    today = datetime.now().strftime("%Y-%m-%d")
    closed_seq = sum(1 for row in ws_c.iter_rows(min_row=3, max_row=500) if row[1].value) + 1

    # ── 5. Handle fully sold positions ───────────────────────────────────────
    # Process in reverse row order so deletions don't shift row numbers
    for ticker in sorted(fully_sold, key=lambda t: current_portfolio[t]["row_num"], reverse=True):
        pos  = current_portfolio[ticker]
        log(f"\nFully sold: {ticker}")

        # Try to get sale price from eToro API
        sale_price, sale_date = fetch_closed_trade_price(ticker, pos["etoro_id"])
        estimated = False
        if sale_price is None:
            sale_price = get_market_price(ticker)  # best estimate
            estimated  = True
            log(f"  Sale price not in API — using market price estimate: {sale_price}")

        sale_date  = sale_date or today
        units_sold = pos["units"] or 0
        invested   = pos["invested"] or 0
        avg_buy    = pos["avg_buy"]
        currency   = pos["currency"]

        # Sale value in USD
        if sale_price and units_sold:
            if currency == "GBp":
                sale_value_usd = units_sold * (sale_price / 100) * 1.34  # approximate GBP/USD
            else:
                sale_value_usd = units_sold * sale_price
        else:
            sale_value_usd = None

        notes = "(sale price estimated — please verify)" if estimated else ""

        add_closed_position(
            ws_c, closed_seq, ticker, pos["sector"],
            units_sold, invested, avg_buy, sale_price,
            sale_value_usd,
            pos["div_2023"], pos["div_2024"], pos["div_2025"],
            sale_date, notes, currency
        )
        closed_seq += 1

        # Remove from Portfolio
        remove_portfolio_row(ws_p, pos["row_num"])
        # Re-read portfolio since rows shifted
        current_portfolio = read_portfolio_sheet(ws_p)

        # Update Tickers sheet → out of portfolio, into watchlist
        update_tickers_flags(ws_t, ticker, "No", "Yes")

        # Add to Watchlist sheet if not already there
        _add_to_watchlist_if_missing(ws_w, ws_t, ticker, pos["company"], pos["sector"], currency)

    # ── 6. Handle partially sold positions ───────────────────────────────────
    for ticker in still_held:
        pos        = current_portfolio[ticker]
        etoro_data = etoro_positions[ticker]
        old_units  = round(pos["units"] or 0, 6)
        new_units  = round(etoro_data["units"], 6)

        if new_units >= old_units * 0.999:   # no meaningful change (rounding tolerance)
            continue

        log(f"\nPartial sell: {ticker}  {old_units} → {new_units} units")

        units_sold   = old_units - new_units
        proportion   = units_sold / old_units if old_units else 0
        invested_old = pos["invested"] or 0
        invested_sold = invested_old * proportion
        invested_kept = invested_old * (1 - proportion)

        # Sale price
        sale_price, sale_date = fetch_closed_trade_price(ticker, pos["etoro_id"])
        estimated = False
        if sale_price is None:
            sale_price = etoro_data.get("current_price") or get_market_price(ticker)
            estimated  = True

        sale_date   = sale_date or today
        currency    = pos["currency"]
        if sale_price and units_sold:
            if currency == "GBp":
                sale_value_usd = units_sold * (sale_price / 100) * 1.27
            else:
                sale_value_usd = units_sold * sale_price
        else:
            sale_value_usd = None

        # Pro-rata dividends
        div_sold = {yr: round((pos[f"div_{yr}"] or 0) * proportion, 2) for yr in ("2023","2024","2025","2026")}
        div_kept = {yr: round((pos[f"div_{yr}"] or 0) * (1 - proportion), 2) for yr in ("2023","2024","2025","2026")}

        notes = f"Partial sell {proportion*100:.1f}%"
        if estimated:
            notes += " (sale price estimated — please verify)"

        add_closed_position(
            ws_c, closed_seq, ticker, pos["sector"],
            units_sold, invested_sold, pos["avg_buy"], sale_price,
            sale_value_usd,
            div_sold["2023"], div_sold["2024"], div_sold["2025"],
            sale_date, notes, currency
        )
        closed_seq += 1

        # Update Portfolio row
        update_portfolio_partial(
            ws_p, pos["row_num"],
            new_units, invested_kept,
            div_kept["2023"], div_kept["2024"], div_kept["2025"], div_kept["2026"]
        )

    # ── 7. Handle new buys ───────────────────────────────────────────────────
    for ticker in new_buys:
        etoro_data = etoro_positions[ticker]
        log(f"\nNew buy: {ticker}  units={etoro_data['units']:.4f}  "
            f"invested=${etoro_data['invested_usd']:.2f}")
        _add_to_portfolio(ws_p, ws_t, ticker, etoro_data)
        update_tickers_flags(ws_t, ticker, "Yes", "No")
        _remove_from_watchlist(ws_w, ticker)

    # ── 8. Update still-held positions (units & invested from eToro) ──────────
    current_portfolio = read_portfolio_sheet(ws_p)   # re-read after deletions
    for ticker in still_held:
        if ticker not in current_portfolio:
            continue
        pos        = current_portfolio[ticker]
        etoro_data = etoro_positions.get(ticker, {})
        old_units  = pos["units"] or 0
        new_units  = etoro_data.get("units", old_units)

        # Only update if eToro shows meaningfully different invested amount
        new_invested = etoro_data.get("invested_usd")
        if new_invested and abs((new_invested - (pos["invested"] or 0)) / max(pos["invested"] or 1, 1)) > 0.01:
            ws_p.cell(row=pos["row_num"], column=11, value=round(new_invested, 2))
            log(f"  Updated {ticker} invested: ${pos['invested']:.2f} → ${new_invested:.2f}")

    # ── 9. Update CASH row balance ────────────────────────────────────────────
    cash_balance = etoro_positions.get("CASH", {}).get("units", 0)
    if cash_balance:
        for row in ws_p.iter_rows(min_row=3, max_row=300):
            if str(row[1].value or "").strip() == "CASH":
                cash_r = row[0].row
                ws_p.cell(row=cash_r, column=11).value = round(cash_balance, 2)  # K = invested
                ws_p.cell(row=cash_r, column=14).value = round(cash_balance, 2)  # N = current value
                log(f"  Updated CASH balance → ${cash_balance:,.2f}")
                break

    # ── 10. Save ──────────────────────────────────────────────────────────────
    wb.save(str(MASTER))
    log(f"\nSaved → {MASTER}")
    log("sync_portfolio.py complete.")


# ── Helper: add a stock to Watchlist if not already present ──────────────────
def _add_to_watchlist_if_missing(ws_w, ws_t, ticker: str, company: str, sector: str, currency: str):
    # Check if already in watchlist
    for row in ws_w.iter_rows(min_row=3, max_row=300):
        if row[3].value and str(row[3].value).strip() == ticker:
            log(f"  {ticker} already in Watchlist")
            return

    # Find next empty row
    next_r = 3
    for row in ws_w.iter_rows(min_row=3, max_row=300):
        if row[3].value:
            next_r = row[0].row + 1

    # Find Yahoo ticker from Tickers sheet
    yf_ticker = ticker
    for row in ws_t.iter_rows(min_row=3, max_row=300):
        if row[3].value and str(row[3].value).strip() == ticker:
            yf_ticker = row[5].value or ticker   # F = Yahoo Finance Ticker
            break

    seq = next_r - 2  # approximate sequence number

    # Write minimal watchlist row (formulas for price/target/signal will be added)
    is_ftse = ticker.endswith(".L")
    live_price_formula = f'=IFERROR(INDEX(stockhistory("{yf_ticker}",TODAY()-7,TODAY(),0,0),1,2),"N/A")'
    price_formula      = f'=IFERROR(IF(F{next_r}="GBp",G{next_r}/100,G{next_r}),"N/A")'
    target_formula     = f'=IFERROR(INDEX(Assumptions!$H:$H,MATCH(D{next_r},Assumptions!$A:$A,0)),"N/A")'
    ratio_formula      = f'=IFERROR(IF(J{next_r}="N/A","N/A",J{next_r}/H{next_r}),"N/A")'
    signal_formula     = (f'=IF(OR(K{next_r}="N/A",K{next_r}=""),"-",'
                          f'IF(K{next_r}>=1.25,"Strong Buy",IF(K{next_r}>=1.1,"Buy",'
                          f'IF(K{next_r}>=0.9,"Fair Value",IF(K{next_r}>=0.75,"Sell","Strong Sell")))))')

    row_data = {
        1:  seq,
        2:  company or ticker,
        3:  sector or "",
        4:  ticker,
        5:  yf_ticker,
        6:  currency or ("GBp" if is_ftse else "USD"),
        7:  live_price_formula,
        8:  price_formula,
        10: target_formula,
        11: ratio_formula,
        12: signal_formula,
        13: datetime.now().strftime("%Y-%m-%d"),
    }
    for col, val in row_data.items():
        c = ws_w.cell(row=next_r, column=col, value=val)
        c.font = INPUT_BLUE if not isinstance(val, str) or not val.startswith("=") else NORMAL_FONT
        c.alignment = RIGHT if col > 5 else LEFT

    log(f"  → Added {ticker} to Watchlist (row {next_r})")


# ── Helper: repair watchlist formula refs after a row deletion ────────────────
def _repair_watchlist_formulas(ws_w):
    """
    After a row is deleted from the Watchlist, rows below shift up by 1 but
    openpyxl does not update formula references. This repairs self-referential
    formulas (those that reference row+1 instead of their own row).
    """
    # Collect all rows that have a ticker (col D)
    data_rows = []
    for row in ws_w.iter_rows(min_row=3, max_row=300):
        if row[3].value and str(row[3].value).strip().upper() not in ("", "TICKER"):
            data_rows.append(row[0].row)

    if not data_rows:
        return

    last_data = max(data_rows)

    for actual_row in data_rows:
        if actual_row == last_data:
            continue  # last row has no next row to be confused with
        stale_ref = actual_row + 1
        row_cells = list(ws_w.iter_rows(min_row=actual_row, max_row=actual_row))[0]
        for cell in row_cells:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                old = cell.value
                new = re.sub(
                    r'(?<!\$)([A-Z]+)' + str(stale_ref) + r'(?!\d)',
                    lambda m: m.group(1) + str(actual_row),
                    old
                )
                if new != old:
                    cell.value = new


# ── Helper: remove from Watchlist ─────────────────────────────────────────────
def _remove_from_watchlist(ws_w, ticker: str):
    for row in ws_w.iter_rows(min_row=3, max_row=300):
        if row[3].value and str(row[3].value).strip() == ticker:
            ws_w.delete_rows(row[0].row, 1)
            log(f"  → Removed {ticker} from Watchlist")
            _repair_watchlist_formulas(ws_w)
            return


# ── Helper: add a new stock to Portfolio sheet ────────────────────────────────
def _add_to_portfolio(ws_p, ws_t, ticker: str, etoro_data: dict):
    # Find Yahoo ticker and company info from Tickers sheet
    yf_ticker = ticker
    company   = ticker
    sector    = ""
    currency  = "GBp" if ticker.endswith(".L") else "USD"
    etoro_id  = etoro_data.get("asset_id", "")

    for row in ws_t.iter_rows(min_row=3, max_row=300):
        if row[3].value and str(row[3].value).strip() == ticker:
            yf_ticker = row[5].value or ticker   # F
            company   = row[1].value or ticker   # B
            sector    = row[7].value or ""        # H
            break

    # Try to get company name and currency from yfinance
    try:
        info = yf.Ticker(yf_ticker).info
        company  = info.get("longName") or info.get("shortName") or company
        currency = info.get("currency", currency)
        if currency == "GBp": currency = "GBp"
        elif currency == "GBP": currency = "GBp"  # yfinance returns GBp as GBP sometimes
        sector   = info.get("sector") or sector
    except Exception:
        pass

    # Insert a new row after last data row (before CASH/totals)
    last_r = last_portfolio_data_row(ws_p)
    r = last_r + 1
    ws_p.insert_rows(r)   # pushes CASH and GRAND TOTAL down by 1

    # After insert, GRAND TOTAL has shifted by 1
    gt_row_after = None
    for sr in ws_p.iter_rows(min_row=3, max_row=300):
        if "GRAND TOTAL" in str(sr[1].value or "").upper():
            gt_row_after = sr[0].row
            break
    if gt_row_after is None:
        gt_row_after = r + 2   # fallback

    # Sequence number
    seq = 1
    for row in ws_p.iter_rows(min_row=3, max_row=last_r):
        if row[3].value:
            seq += 1

    units    = round(etoro_data.get("units", 0), 6)
    invested = round(etoro_data.get("invested_usd", 0), 2)
    trades   = etoro_data.get("trades", 1)

    price_fmt  = "£#,##0.00" if currency == "GBp" else "$#,##0.00"

    # Write static fields
    statics = {
        1:  (seq,       None,           LEFT),
        2:  (company,   None,           LEFT),
        3:  (sector,    None,           LEFT),
        4:  (ticker,    None,           LEFT),
        5:  (yf_ticker, None,           LEFT),
        6:  (currency,  None,           CENTER),
        7:  (str(etoro_id), None,       CENTER),
        8:  (trades,    None,           CENTER),
        9:  (units,     "#,##0.000000", RIGHT),
        11: (invested,  "$#,##0.00",    RIGHT),
    }
    for col, (val, fmt, align) in statics.items():
        c = ws_p.cell(row=r, column=col, value=val)
        c.font = INPUT_BLUE; c.alignment = align
        if fmt: c.number_format = fmt

    # Write formulas (weight formula uses dynamically computed GRAND TOTAL row)
    formulas = {
        12: f'=IFERROR(INDEX(stockhistory("{yf_ticker}",TODAY()-7,TODAY(),0,0),1,2),"N/A")',
        13: f'=IFERROR(IF(F{r}="GBp",L{r}/100,L{r}),"N/A")',
        14: f'=IFERROR(IF(F{r}="GBp",I{r}*L{r}/100*Assumptions!$B$3,I{r}*M{r}),"N/A")',
        15: f'=IFERROR(N{r}/$N${gt_row_after},0)',
        16: f'=IFERROR(N{r}-K{r},"N/A")',
        17: f'=IFERROR(P{r}/K{r},"N/A")',
        22: f'=SUM(R{r}:U{r})',
        23: f'=IFERROR(V{r}/K{r},0)',
        24: f'=Q{r}',
        25: f'=IFERROR((P{r}+V{r})/K{r},"N/A")',
        28: f'=IFERROR(INDEX(Assumptions!$H:$H,MATCH(D{r},Assumptions!$A:$A,0)),"N/A")',
        29: f'=IFERROR(IF(AB{r}="N/A","N/A",AB{r}/M{r}),"N/A")',
        30: (f'=IF(OR(AC{r}="N/A",AC{r}=""),"-",'
             f'IF(AC{r}>=1.25,"Strong Buy",IF(AC{r}>=1.1,"Buy",'
             f'IF(AC{r}>=0.9,"Fair Value",IF(AC{r}>=0.75,"Sell","Strong Sell")))))'),
        32: (f'=IFERROR(MAX(INDEX(stockhistory("{yf_ticker}",'
             f'DATE(YEAR(TODAY())-1,MONTH(TODAY()),DAY(TODAY())),'
             f'TODAY(),1,0),0,2)),"N/A")'),
        33: (f'=IFERROR(MIN(INDEX(stockhistory("{yf_ticker}",'
             f'DATE(YEAR(TODAY())-1,MONTH(TODAY()),DAY(TODAY())),'
             f'TODAY(),1,0),0,2)),"N/A")'),
        34: f'=IFERROR((L{r}-AG{r})/(AF{r}-AG{r}),"N/A")',
    }
    for col, formula in formulas.items():
        c = ws_p.cell(row=r, column=col, value=formula)
        c.font = NORMAL_FONT; c.alignment = RIGHT
        c.number_format = price_fmt if col in (12, 13, 28) else ("0.00%" if col in (15,17,23,24,25) else "General")

    # Fix GRAND TOTAL and all weight formulas to account for the new row
    _fix_portfolio_summary_rows(ws_p)

    log(f"  → Added {ticker} to Portfolio (row {r})  units={units}  invested=${invested}")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    run_sync()
