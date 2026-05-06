"""Audit and clean up the Tickers sheet in eToro_Master.xlsx.

Three jobs
----------
1. Find duplicate rows with the same eToro Ticker (col D). For each duplicate
   group, keep the most-populated row and delete the others. Without this, the
   watchlist's MATCH(ticker, Tickers!D:D, 0) often matches the *first* row,
   which can be the broken/incomplete duplicate -> blank live prices.

2. Report rows missing an eToro Asset ID (col E). This is what the orphan
   resolver in sync_portfolio.py uses, so missing IDs cause new positions to
   stay stuck as ID_NNNN.

3. Report rows missing a manual price override (col N). The watchlist's
   STOCKHISTORY-driven price formula falls back to col N when the live data
   call fails - so missing overrides means a transient Excel error becomes a
   visible blank.

Usage
-----
    python scripts/audit_tickers_sheet.py            # dry-run
    python scripts/audit_tickers_sheet.py --apply    # delete duplicate rows
    python scripts/audit_tickers_sheet.py --apply --backup
"""
from __future__ import annotations
import argparse
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import _envloader  # noqa: F401

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
MASTER = BASE_DIR / "data" / "eToro_Master.xlsx"


# Tickers sheet column layout (1-indexed):
# A=#  B=Company  C=FTSE Ticker  D=eToro Ticker  E=eToro Asset ID
# F=Yahoo Ticker  G=Market  H=Sector  I=Asset Type  J=In Portfolio
# K=In Watchlist  L=Notes  M=Microsoft Ticker (formula)  N=Manual Price Override


def _row_score(row_data, ticker):
    """Higher = more populated. The 'best' row wins."""
    score = 0
    company = (row_data.get("company") or "").strip()
    if company and company.upper() != ticker.upper():
        score += 10  # real company name
    if (row_data.get("sector") or "").strip():
        score += 5
    if row_data.get("asset_id") is not None:
        score += 3
    if row_data.get("override") is not None:
        score += 2
    # Tiebreak: higher row number (newer)
    score += row_data["row"] / 1000.0
    return score


def _load_tickers_rows(ws_t):
    """Read all data rows from the Tickers sheet into a list of dicts."""
    rows = []
    for r in range(3, 300):
        ticker = ws_t.cell(row=r, column=4).value
        if not ticker or not isinstance(ticker, str):
            continue
        rows.append({
            "row": r,
            "ticker": ticker.strip(),
            "company": ws_t.cell(row=r, column=2).value or "",
            "ftse_ticker": ws_t.cell(row=r, column=3).value or "",
            "asset_id": ws_t.cell(row=r, column=5).value,
            "yahoo": ws_t.cell(row=r, column=6).value or "",
            "market": ws_t.cell(row=r, column=7).value or "",
            "sector": ws_t.cell(row=r, column=8).value or "",
            "asset_type": ws_t.cell(row=r, column=9).value or "",
            "in_pf": ws_t.cell(row=r, column=10).value or "",
            "in_wl": ws_t.cell(row=r, column=11).value or "",
            "notes": ws_t.cell(row=r, column=12).value,
            "ms_ticker": ws_t.cell(row=r, column=13).value,
            "override": ws_t.cell(row=r, column=14).value,
        })
    return rows


def _merge_into_winner(ws_t, winner_row, losers):
    """Before deleting losers, copy any non-empty fields from losers into winner
    if the winner's field is empty. Stops us losing data when the broken row
    happens to have an override but the otherwise-complete row doesn't."""
    transferred = []
    for col_idx, key in [
        (2, "company"), (5, "asset_id"), (8, "sector"), (10, "in_pf"),
        (11, "in_wl"), (14, "override"),
    ]:
        cur = ws_t.cell(row=winner_row, column=col_idx).value
        if cur in (None, "", winner_row):  # winner_row check is defensive
            for L in losers:
                v = L.get(key)
                if v not in (None, ""):
                    ws_t.cell(row=winner_row, column=col_idx, value=v)
                    transferred.append(f"{key}={v!r}")
                    break
    return transferred


def _repair_self_refs_after_deletion(ws_t, deleted_rows_sorted_desc):
    """When we delete row N, rows below shift up by 1 (and so on for cumulative
    deletions). openpyxl does NOT update formula references that point to a
    specific row number. The Microsoft Ticker formula in column M is the main
    self-referential one (=IF(G{r}=...,SUBSTITUTE(C{r},...))). Fix it for every
    surviving row to point at its own current row."""
    for row in ws_t.iter_rows(min_row=3, max_row=300):
        actual_row = row[0].row
        ms_cell = ws_t.cell(row=actual_row, column=13)
        v = ms_cell.value
        if isinstance(v, str) and v.startswith("="):
            # Replace any [A-Z]+\d+ pattern that's not the actual_row with actual_row,
            # but only for the columns this formula uses (G, C, D)
            new = re.sub(
                r'(?<!\$)([CDG])(\d+)(?!\d)',
                lambda m: f'{m.group(1)}{actual_row}',
                v,
            )
            if new != v:
                ms_cell.value = new


def find_duplicates(rows):
    """Return {ticker: [rows]} for tickers with more than one row."""
    by_ticker = defaultdict(list)
    for row in rows:
        by_ticker[row["ticker"]].append(row)
    return {t: rs for t, rs in by_ticker.items() if len(rs) > 1}


def dedupe_tickers(wb, *, log=print):
    """Returns (deleted_count, merged_count)."""
    if "Tickers" not in wb.sheetnames:
        log("  audit_tickers: no Tickers sheet, skipping")
        return (0, 0)
    ws_t = wb["Tickers"]
    rows = _load_tickers_rows(ws_t)
    dups = find_duplicates(rows)
    if not dups:
        return (0, 0)

    # Pick winner per group, then delete losers (bottom-up so row numbers
    # don't shift mid-loop)
    losers_to_delete = []
    merged = 0
    for ticker, group in dups.items():
        scored = sorted(group, key=lambda r: -_row_score(r, ticker))
        winner = scored[0]
        losers = scored[1:]
        transfers = _merge_into_winner(ws_t, winner["row"], losers)
        if transfers:
            merged += 1
            log(f"  Tickers: {ticker} - merged into row {winner['row']}: {', '.join(transfers)}")
        for L in losers:
            losers_to_delete.append((L["row"], ticker, winner["row"]))

    # Delete bottom-up
    losers_to_delete.sort(reverse=True)
    deleted_rows = []
    for r, ticker, winner_row in losers_to_delete:
        ws_t.delete_rows(r, 1)
        deleted_rows.append(r)
        log(f"  Tickers: deleted row {r} (duplicate {ticker}, winner is row {winner_row})")

    if deleted_rows:
        _repair_self_refs_after_deletion(ws_t, sorted(deleted_rows, reverse=True))

    return (len(deleted_rows), merged)


def report_hygiene(wb, *, log=print):
    """Report rows missing eToro Asset ID and manual price override."""
    if "Tickers" not in wb.sheetnames:
        return
    ws_t = wb["Tickers"]
    rows = _load_tickers_rows(ws_t)

    missing_asset_id = [r["ticker"] for r in rows if r["asset_id"] is None]
    missing_override = [r["ticker"] for r in rows if r["override"] is None]

    if missing_asset_id:
        log(f"  {len(missing_asset_id)} row(s) missing eToro Asset ID (col E).")
        log(f"    First 10: {', '.join(missing_asset_id[:10])}")
        if len(missing_asset_id) > 10:
            log(f"    ... and {len(missing_asset_id)-10} more")
    if missing_override:
        log(f"  {len(missing_override)} row(s) missing manual price override (col N).")
        log(f"    Tickers: {', '.join(missing_override)}")


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--apply", action="store_true", help="Write changes back. Without this, runs as a dry-run.")
    p.add_argument("--backup", action="store_true", help="When applying, also write a timestamped backup.")
    p.add_argument("--master", default=str(MASTER), help="Path to eToro_Master.xlsx")
    args = p.parse_args()

    master = Path(args.master)
    if not master.exists():
        print(f"ERROR: {master} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {master} ...")
    wb = openpyxl.load_workbook(master)

    print("\nDuplicate analysis:")
    deleted, merged = dedupe_tickers(wb)
    if not deleted:
        print("  No duplicates found.")

    print("\nHygiene checks:")
    report_hygiene(wb)

    if args.apply and deleted:
        if args.backup:
            stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            backup = master.with_suffix(f".bak-{stamp}.xlsx")
            shutil.copy2(master, backup)
            print(f"\nWrote backup: {backup}")
        wb.save(master)
        print(f"Saved. Deleted {deleted} duplicate row(s), merged {merged}.")
    elif deleted:
        print(f"\nDRY-RUN: would delete {deleted} duplicate row(s), merge {merged}. Re-run with --apply.")
    else:
        print("\nNo changes proposed.")


if __name__ == "__main__":
    main()
