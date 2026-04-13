#!/usr/bin/env python3
"""
valuation.py
────────────
Consolidated valuation script — replaces daily_ftse_valuation.py + intl_valuation.py.

Reads ALL tickers from the "Yahoo Finance Ticker" column of the Tickers sheet
in eToro_Master.xlsx, so adding a new ticker to the sheet is all you need to do.

For each ticker:
  • FTSE (.L)  → sector-routed valuation, output in GBP (£)
       – Banks          : Two-stage DDM (45%) + P/B Excess Returns (35%) + Earnings Cap (20%)
       – Gen Insurance  : P/Tangible Book adjusted for Combined Ratio
       – Life Insurance : EPS capitalisation (Ke)
       – Asset Mgmt     : Justified P/B (60%) + Earnings Cap (40%)
       – PE/Alternatives: NAV × 1.10 premium; fallback EPS/Ke
       – Capital Markets: EV/EBITDA @ 22×
       – Other fin.     : DDM + EPV + Residual Income
       – Non-financial  : DCF + DDM + EPV (3-model blend)
  • US / Intl  → DCF where profitable, else analyst consensus, output in USD ($)
  • Skip       → BTC, ETFs, CASH (no fundamental valuation possible)

Writes results to:
  • Assumptions sheet in eToro_Master.xlsx  (cols E, I, J, K, L, M, N)
  • Reports/ftse_report.csv
  • Reports/intl_report.csv

Usage:
  python valuation.py

Optional env var:
  FTSE_VALUATION_DISCORD_WEBHOOK  — Discord notification on completion
"""

import sys, os, time, csv, requests
import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR     = Path(__file__).parent.parent
DATA_DIR     = BASE_DIR / "data"
REPORTS_DIR  = BASE_DIR / "Reports"
LOGS_DIR     = BASE_DIR / "logs"
ONEDRIVE_DIR = Path(r"C:\Users\Neil\OneDrive\Personal\Etoro")
MASTER       = DATA_DIR / "eToro_Master.xlsx"
LOG_FILE     = LOGS_DIR / "valuation_run.log"

REPORTS_DIR.mkdir(exist_ok=True)
LOGS_DIR.mkdir(exist_ok=True)

DISCORD_WEBHOOK = os.getenv("FTSE_VALUATION_DISCORD_WEBHOOK")

# ── Logging ───────────────────────────────────────────────────────────────────
def log(msg: str):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")
    print(line)

# ── Styling ───────────────────────────────────────────────────────────────────
INPUT_BLUE = Font(color="FF0000FF", name="Arial", size=10)
UPDATED_GREEN = Font(color="FF059669", name="Arial", size=10)
NORMAL_FONT   = Font(name="Arial", size=10)
RIGHT  = Alignment(horizontal="right",  vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")

def _set_blue(ws, row, col, value, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = INPUT_BLUE
    c.alignment = RIGHT
    if fmt: c.number_format = fmt

# ── Constants ─────────────────────────────────────────────────────────────────
WACC_DEFAULT     = 0.09
GROWTH_5Y        = 0.05
TERMINAL_GROWTH  = 0.025
PE_MULTIPLE      = 15           # conservative normalised P/E for EPV
GBP_USD          = 1.34         # fallback — script reads live rate from Assumptions!B3
EUR_GBP          = 0.845        # fallback — script fetches live rate from yfinance (EURGBP=X)

SKIP_ASSET_TYPES = {"ETF", "Crypto", "Cash", "Bond"}
SKIP_TICKERS     = {"BTC", "CASH", "LQDE.L"}   # no fundamental valuation

# Financial sub-sector routing — all use Ke (cost of equity) not WACC
# Maps from yfinance industry string → display sub-sector used for routing
FIN_INDUSTRY_MAP = {
    "Banks - Regional":                  "Banks",
    "Banks - Diversified":               "Banks",
    "Insurance - Life":                  "Life Insurance",
    "Insurance - Diversified":           "General Insurance",
    "Insurance - Property & Casualty":   "General Insurance",
    "Insurance - Specialty":             "General Insurance",
    "Asset Management":                  "Asset Management",
    "Financial Data & Stock Exchanges":  "Capital Markets",
    "Mortgage Finance":                  "Financial Services",
}

# Ticker-level overrides (take precedence over industry map)
FIN_TICKER_MAP = {
    "LLOY.L": "Banks",           "BARC.L": "Banks",
    "HSBA.L": "Banks",           "STAN.L": "Banks",
    "NWG.L":  "Banks",
    "ADM.L":  "General Insurance", "AV.L":  "General Insurance",
    "BEZ.L":  "General Insurance", "HSX.L": "General Insurance",
    "PRU.L":  "General Insurance",
    "SLA.L":  "Life Insurance",  "LGEN.L": "Life Insurance",
    "PHX.L":  "Life Insurance",  "JUP.L":  "Asset Management",
    "MNG.L":  "Asset Management","SDR.L":  "Asset Management",
    "ICG.L":  "PE/Alternatives", "3IN.L":  "PE/Alternatives",
    "LSE.L":  "Capital Markets",
}

# All financial display sub-sectors
FIN_DISPLAY_SECTORS = {
    "Banks", "General Insurance", "Life Insurance",
    "Asset Management", "PE/Alternatives", "Capital Markets",
    "Insurance", "Financial Services",
}

# Sectors where FCF / DCF is not meaningful — use specialist financial models instead
SKIP_DCF_SECTORS = FIN_DISPLAY_SECTORS

# Bank-specific model constants
BANK_G1              = 0.035    # near-term dividend growth for banks
BANK_G2              = 0.025    # terminal growth (matches TERMINAL_GROWTH)
BANK_EM_PREMIUM      = {"HSBA.L": 0.010, "STAN.L": 0.010}   # extra country risk
BANK_MULTIPLIERS     = {"LLOY.L": 0.92, "BARC.L": 0.95}      # structural risk haircuts
BANK_WEIGHTS         = {"ddm": 0.45, "pb": 0.35, "epv": 0.20}

# Capital Markets EV/EBITDA multiple
CM_EV_EBITDA         = 22.0

# FCF-to-earnings conversion by sector (for fallback FCF estimate)
FCF_CONVERSION = {
    "Utilities":          0.85,
    "Consumer Defensive": 0.80,
    "Energy":             0.70,
    "Basic Materials":    0.65,
    "Industrials":        0.75,
}

# Sectors where fallback FCF is too unreliable to use in DCF blending.
# For these sectors, DCF is only included when primary FCF data is available.
# Basic Materials (miners) have highly volatile, cycle-dependent FCF — the
# EPS × conversion proxy produces nonsense when EPS is depressed or distorted.
SKIP_FALLBACK_DCF_SECTORS = {"Basic Materials"}

# Sectors where DDM is unreliable and should be excluded from blending.
# Commodity miners/traders pay variable distributions tied to commodity cycles,
# not stable growing dividends. DDM (which assumes perpetual dividend growth)
# produces misleadingly low valuations when the reported dividendRate is
# depressed or doesn't capture the full distribution (e.g. Glencore pays in
# USD special + base distributions that yfinance may underreport for .L tickers).
# For these sectors, only EPV and primary-method DCF are used.
SKIP_DDM_SECTORS = {"Basic Materials"}

# ── Read GBP/USD from Assumptions sheet ──────────────────────────────────────
def read_gbp_usd(wb) -> float:
    try:
        val = wb["Assumptions"]["B3"].value
        return float(val) if val else GBP_USD
    except Exception:
        return GBP_USD

# ── Fetch live EUR/GBP rate from yfinance ────────────────────────────────────
def fetch_eur_gbp() -> float:
    try:
        rate = yf.Ticker("EURGBP=X").fast_info.last_price
        if rate and 0.5 < rate < 1.5:
            log(f"EUR/GBP = {rate:.4f} (live)")
            return float(rate)
    except Exception:
        pass
    log(f"EUR/GBP = {EUR_GBP} (fallback)")
    return EUR_GBP

# ── Load tickers from Tickers sheet ──────────────────────────────────────────
def load_tickers(wb) -> list[dict]:
    """
    Returns list of dicts:
      {yf_ticker, etoro_ticker, company, sector, market, asset_type, in_portfolio, in_watchlist}
    """
    ws = wb["Tickers"]
    tickers = []
    for row in ws.iter_rows(min_row=3, max_row=400, values_only=True):
        if not row[5]:   # F = Yahoo Finance Ticker — skip if blank
            continue
        yf_t     = str(row[5]).strip()
        asset_t  = str(row[8] or "").strip()   # I = Asset Type
        if asset_t in SKIP_ASSET_TYPES or yf_t in SKIP_TICKERS:
            continue
        tickers.append({
            "yf_ticker":    yf_t,
            "etoro_ticker": str(row[3] or yf_t).strip(),   # D
            "company":      str(row[1] or "").strip(),      # B
            "sector":       str(row[7] or "").strip(),      # H
            "market":       str(row[6] or "").strip(),      # G
            "asset_type":   asset_t,
            "in_portfolio": str(row[9] or "").strip(),      # J
            "in_watchlist": str(row[10] or "").strip(),     # K
        })
    log(f"Loaded {len(tickers)} tickers from Tickers sheet (skipped ETF/Crypto/Cash/Bond/BTC)")
    return tickers

# ── FCF helper ────────────────────────────────────────────────────────────────
def get_fcf_per_share(stock: yf.Ticker, sector: str, shares: float, eps: float) -> tuple:
    """
    Returns (fcf_per_share, method, years_used)
    method: 'primary' | 'fallback' | 'skip'
    """
    if sector in SKIP_DCF_SECTORS:
        return np.nan, "skip", 0
    try:
        cf = stock.cashflow
        if cf is None or cf.empty:
            raise ValueError("No cashflow data")
        ocf_label = next((l for l in cf.index if "operating" in l.lower() and "cash" in l.lower()), None)
        capex_label = next((l for l in cf.index if "capital" in l.lower() and "expenditure" in l.lower()), None)
        if not ocf_label:
            raise ValueError("OCF row not found")
        ocf_vals   = cf.loc[ocf_label].dropna().values[:3]
        capex_vals = cf.loc[capex_label].dropna().values[:3] if capex_label else np.zeros(len(ocf_vals))
        if len(ocf_vals) == 0:
            raise ValueError("No OCF values")
        fcf_vals = ocf_vals[:len(capex_vals)] - np.abs(capex_vals[:len(ocf_vals)])
        if any(fcf_vals < 0) or shares <= 0:
            raise ValueError("Negative FCF")
        fcf_norm = np.mean(fcf_vals) / shares
        return fcf_norm, "primary", len(fcf_vals)
    except Exception:
        pass
    # Fallback: EPS × conversion ratio
    ratio = FCF_CONVERSION.get(sector)
    if ratio and eps and eps > 0:
        return eps * ratio, "fallback", 0
    return np.nan, "skip", 0

# ── DCF valuation ─────────────────────────────────────────────────────────────
def dcf_value(fcf: float, wacc: float, g1: float = GROWTH_5Y, g2: float = TERMINAL_GROWTH) -> float:
    pv = 0.0
    for yr in range(1, 6):
        pv += fcf * (1 + g1) ** yr / (1 + wacc) ** yr
    terminal = fcf * (1 + g1) ** 5 * (1 + g2) / (wacc - g2)
    pv += terminal / (1 + wacc) ** 5
    return pv

# ── DDM valuation ─────────────────────────────────────────────────────────────
def ddm_value(div: float, wacc: float, g2: float = TERMINAL_GROWTH) -> float:
    if wacc <= g2 or div <= 0:
        return np.nan
    return div * (1 + g2) / (wacc - g2)

# ── EPV / earnings-power valuation (forward PE) ───────────────────────────────
def epv_value(forward_eps: float, pe_multiple: float = PE_MULTIPLE) -> float:
    if forward_eps and forward_eps > 0:
        return forward_eps * pe_multiple
    return np.nan

# ── WACC estimate from beta ───────────────────────────────────────────────────
def estimate_wacc(beta: float, is_ftse: bool) -> float:
    rf   = 0.049 if is_ftse else 0.045  # UK 10yr / US 10yr (updated Q2 2026: US 10yr ~4.44%)
    erp  = 0.05                          # equity risk premium
    wacc = rf + max(0.5, min(beta or 1.0, 2.5)) * erp
    return round(wacc, 4)

# ── Cost of equity (Ke) — used for all financial sub-sectors ─────────────────
def estimate_ke(beta: float, is_ftse: bool) -> float:
    """
    For financial stocks, discount at Ke (equity rate only), not WACC.
    Leverage in banks/insurers is operational, not a capital-structure choice,
    so blending in a cost-of-debt would understate the required return.
    """
    rf  = 0.049 if is_ftse else 0.045  # US 10yr updated Q2 2026: ~4.44%
    erp = 0.05
    return round(rf + max(0.5, min(beta or 1.0, 2.5)) * erp, 4)

# ── Determine financial sub-sector ───────────────────────────────────────────
def get_fin_subsector(yf_ticker: str, yf_sector: str, yf_industry: str) -> str | None:
    """
    Returns a financial display sub-sector string, or None if not a financial stock.
    Ticker-level map takes precedence; falls back to industry map; then broad sector check.
    """
    if yf_ticker in FIN_TICKER_MAP:
        return FIN_TICKER_MAP[yf_ticker]
    mapped = FIN_INDUSTRY_MAP.get(yf_industry or "")
    if mapped:
        return mapped
    broad = (yf_sector or "").lower()
    if "bank" in broad:
        return "Banks"
    if "insurance" in broad:
        return "General Insurance"
    if "financial" in broad:
        return "Financial Services"
    return None

# ── 3-year average diluted EPS ────────────────────────────────────────────────
def get_eps_3yr_avg(stock: yf.Ticker, fallback_eps) -> float:
    try:
        inc = stock.income_stmt
        if inc is None or inc.empty:
            raise ValueError("No income statement")
        eps_label = next(
            (l for l in inc.index if "diluted" in l.lower() and "eps" in l.lower()), None
        )
        if eps_label is None:
            eps_label = next((l for l in inc.index if "eps" in l.lower()), None)
        if eps_label is None:
            raise ValueError("No EPS row")
        vals = inc.loc[eps_label].iloc[:3].dropna()
        if len(vals) == 0:
            raise ValueError("Empty EPS")
        return float(np.mean(vals))
    except Exception:
        return float(fallback_eps) if fallback_eps and not np.isnan(float(fallback_eps or np.nan)) else np.nan

# ══════════════════════════════════════════════════════════════════════════════
# FINANCIAL SECTOR VALUATION MODELS
# ══════════════════════════════════════════════════════════════════════════════

def val_banks(yf_ticker: str, ke: float, div: float, roe: float,
              book: float, eps3: float) -> tuple:
    """
    Three-method bank valuation.
    Returns (ddm_val, pb_val, epv_val, blended_target, method_label)

    Method 1 — Two-stage DDM at Ke  (weight 45%)
    Method 2 — P/B Excess Returns   (weight 35%)
      justified_pb = 1 + (ROE − Ke) / (Ke − g),  clamped [0.3, 3.0]
    Method 3 — Earnings capitalisation on 3yr avg EPS / Ke  (weight 20%)
    """
    g1 = BANK_G1
    g2 = BANK_G2
    # Apply EM country-risk premium on top of Ke for HSBC / Standard Chartered
    ke_adj = ke + BANK_EM_PREMIUM.get(yf_ticker, 0.0)

    # Method 1: Two-stage DDM
    if div and div > 0 and ke_adj > g2:
        pv_divs  = sum(div * (1 + g1)**t / (1 + ke_adj)**t for t in range(1, 6))
        div5     = div * (1 + g1)**5
        terminal = (div5 * (1 + g2)) / (ke_adj - g2)
        ddm_val  = pv_divs + terminal / (1 + ke_adj)**5
    else:
        ddm_val = np.nan

    # Method 2: P/B Excess Returns
    if (roe and not np.isnan(roe) and book and not np.isnan(book)
            and book > 0 and ke_adj > g2):
        justified_pb = max(0.3, min(3.0, 1 + (roe - ke_adj) / (ke_adj - g2)))
        pb_val = justified_pb * book
    else:
        pb_val = np.nan

    # Method 3: Earnings capitalisation
    if eps3 and not np.isnan(eps3) and eps3 > 0 and ke_adj > 0:
        epv_val = eps3 / ke_adj
    else:
        epv_val = np.nan

    # Weighted blend — re-weight if any method is missing
    candidates = {"ddm": (BANK_WEIGHTS["ddm"], ddm_val),
                  "pb":  (BANK_WEIGHTS["pb"],  pb_val),
                  "epv": (BANK_WEIGHTS["epv"], epv_val)}
    valid = {k: (w, v) for k, (w, v) in candidates.items()
             if v is not None and not np.isnan(v) and v > 0}
    if valid:
        total_w = sum(w for w, _ in valid.values())
        target  = sum(w * v for w, v in valid.values()) / total_w
        target *= BANK_MULTIPLIERS.get(yf_ticker, 1.0)
        methods = "+".join(k.upper() for k in valid)
    else:
        target  = np.nan
        methods = "No Valuation"

    return ddm_val, pb_val, epv_val, target, f"Bank:{methods}"


def val_general_insurance(ke: float, roe: float, book: float) -> tuple:
    """
    P/Tangible Book adjusted for underwriting quality.
    justified_ptb = (ROE / Ke) + cr_adjustment,  clamped [0.5, 3.0]
    Combined Ratio not available in yfinance — cr_adjustment defaults to 0.
    """
    if (roe and not np.isnan(roe) and book and not np.isnan(book)
            and book > 0 and ke > 0):
        cr_adj        = 0.0   # Combined Ratio unavailable from yfinance
        justified_ptb = max(0.5, min(3.0, (roe / ke) + cr_adj))
        val = justified_ptb * book
        return val, val, f"GI:PTB({justified_ptb:.2f}x)"
    return np.nan, np.nan, "GI:NoData"


def val_life_insurance(ke: float, eps3: float) -> tuple:
    """
    Embedded Value not available in yfinance — fall back to EPS / Ke.
    """
    if eps3 and not np.isnan(eps3) and eps3 > 0 and ke > 0:
        val = eps3 / ke
        return val, f"Life:EPSCap"
    return np.nan, "Life:NoData"


def val_asset_management(ke: float, roe: float, book: float, eps3: float) -> tuple:
    """
    Method 1 — Justified P/B  (weight 60%):  (ROE / Ke) × Book,  clamped [0.5, 3.0]
    Method 2 — Earnings cap   (weight 40%):  EPS_3yr / Ke
    """
    # Method 1
    if (roe and not np.isnan(roe) and book and not np.isnan(book)
            and book > 0 and ke > 0):
        justified_pb = max(0.5, min(3.0, roe / ke))
        pb_val = justified_pb * book
    else:
        pb_val = np.nan

    # Method 2
    epv_val = (eps3 / ke) if (eps3 and not np.isnan(eps3) and eps3 > 0 and ke > 0) else np.nan

    candidates = {"pb": (0.60, pb_val), "epv": (0.40, epv_val)}
    valid = {k: (w, v) for k, (w, v) in candidates.items()
             if v is not None and not np.isnan(v) and v > 0}
    if valid:
        total_w = sum(w for w, _ in valid.values())
        target  = sum(w * v for w, v in valid.values()) / total_w
        methods = "+".join(k.upper() for k in valid)
    else:
        target  = np.nan
        methods = "NoData"

    return pb_val, epv_val, target, f"AM:{methods}"


def val_pe_alternatives(book: float, ke: float, eps3: float) -> tuple:
    """
    NAV-based: Book Value × 1.10 quality premium.
    Fallback: EPS_3yr / Ke.
    """
    if book and not np.isnan(book) and book > 0:
        val = book * 1.10
        return val, "PE:NAV"
    if eps3 and not np.isnan(eps3) and eps3 > 0 and ke > 0:
        return eps3 / ke, "PE:EPSCap"
    return np.nan, "PE:NoData"


def val_capital_markets(ebitda: float, net_debt: float, shares: float) -> tuple:
    """
    EV/EBITDA @ 22× for exchange/data businesses with recurring revenue.
    Target = (EBITDA × 22 − Net Debt) / Shares
    """
    if (ebitda and not np.isnan(ebitda) and ebitda > 0
            and shares and not np.isnan(shares) and shares > 0):
        ev           = ebitda * CM_EV_EBITDA
        equity_value = ev - float(net_debt if net_debt and not np.isnan(net_debt) else 0)
        val          = equity_value / shares
        return (val if val > 0 else np.nan), "CM:EV/EBITDA"
    return np.nan, "CM:NoData"


def val_other_financial(ke: float, div: float, eps: float, book: float, roe: float) -> tuple:
    """
    Catch-all for financial stocks not matched by a specific sub-sector.
    Blends: DDM + EPV (EPS/Ke) + Residual Income Model.
    Residual Income: B0 × (1 + (ROE − Ke) / (Ke − g))
    """
    g = TERMINAL_GROWTH
    # DDM
    if div and div > 0 and ke > g:
        ddm = div * (1 + g) / (ke - g)
    else:
        ddm = np.nan
    # EPV
    epv = (eps / ke) if (eps and not np.isnan(eps) and eps > 0 and ke > 0) else np.nan
    # Residual Income
    if (book and not np.isnan(book) and book > 0 and roe and not np.isnan(roe)
            and ke > g):
        rim = book * (1 + (roe - ke) / (ke - g))
        rim = rim if rim > 0 else np.nan
    else:
        rim = np.nan

    vals = [v for v in [ddm, epv, rim] if v is not None and not np.isnan(v) and v > 0]
    target = float(np.mean(vals)) if vals else np.nan
    return ddm, epv, rim, target, "Fin:DDM+EPV+RIM"

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN VALUATION LOOP
# ═══════════════════════════════════════════════════════════════════════════════
def value_ticker(row: dict, gbp_usd: float, eur_gbp: float = EUR_GBP) -> dict | None:
    """
    Fetch data from yfinance and run valuations for one ticker.
    Returns a result dict or None on failure.
    """
    yf_t    = row["yf_ticker"]
    sector  = row["sector"]
    is_ftse = yf_t.endswith(".L")

    try:
        stock = yf.Ticker(yf_t)
        info  = stock.info
        if not info or not info.get("regularMarketPrice"):
            log(f"  {yf_t}: no price data — skipping")
            return None
    except Exception as e:
        log(f"  {yf_t}: yfinance error — {e}")
        return None

    price       = info.get("currentPrice") or info.get("regularMarketPrice") or 0
    beta        = info.get("beta") or 1.0
    fwd_eps     = info.get("forwardEps")
    trail_eps   = info.get("trailingEps")
    div_rate    = info.get("dividendRate") or 0
    fwd_pe      = info.get("forwardPE")
    shares      = info.get("sharesOutstanding") or 0
    sector      = info.get("sector") or sector
    industry    = info.get("industry") or ""
    company     = info.get("longName") or info.get("shortName") or row["company"]
    currency    = info.get("currency", "GBp" if is_ftse else "USD")

    # ── Helper: convert a yfinance price/target from its native currency to GBP ──
    # Most FTSE .L stocks quote in GBp (pence); divide by 100 → GBP.
    # A few .L stocks (e.g. IHG.L, MTLN.L) quote in USD or EUR on yfinance.
    def _to_gbp(value, src_currency=currency):
        """Convert a yfinance per-share value to GBP based on its source currency."""
        if value is None or (isinstance(value, float) and np.isnan(value)):
            return value
        # Check ORIGINAL case first: yfinance uses "GBp" for pence, "GBP" for pounds
        if src_currency == "GBp" or src_currency == "GBX" or src_currency == "GBx":
            return value / 100              # pence → pounds
        src = (src_currency or "GBP").upper()
        if src == "GBP":
            return value                    # already pounds
        elif src == "USD":
            return value / gbp_usd          # USD → GBP
        elif src == "EUR":
            return value * eur_gbp          # EUR → GBP
        else:
            log(f"  WARNING: unknown currency '{src_currency}' for {yf_t}, treating as GBp")
            return value / 100              # fallback: assume pence
    book_value  = info.get("bookValue")          # used by financial models
    roe         = info.get("returnOnEquity")      # used by financial models
    ebitda      = info.get("ebitda")              # used by Capital Markets model
    total_debt  = info.get("totalDebt") or 0
    total_cash  = info.get("totalCash") or 0
    net_debt    = total_debt - total_cash

    wacc = estimate_wacc(beta, is_ftse)
    ke   = estimate_ke(beta, is_ftse)            # cost of equity for financial stocks

    # ── Determine if this is a financial sub-sector stock ─────────────────────
    fin_subsector = get_fin_subsector(yf_t, sector, industry) if is_ftse else None

    # ── Fetch 3-year average EPS (used by financial models) ───────────────────
    if fin_subsector:
        eps3 = get_eps_3yr_avg(stock, trail_eps)
    else:
        eps3 = np.nan

    analyst_target_raw = info.get("targetMeanPrice")

    # ══════════════════════════════════════════════════════════════════════════
    # FINANCIAL SECTOR ROUTING (FTSE only)
    # ══════════════════════════════════════════════════════════════════════════
    if is_ftse and fin_subsector:
        financial_currency = (info.get("financialCurrency") or "GBP").upper()
        price_out = round(_to_gbp(price), 4)   # native currency → GBP

        # book_value and eps3 are in GBP for .L stocks (yfinance returns GBP, not GBp)
        # div_rate is also in GBP (same convention as EPS)
        # analyst target is in the same currency as currentPrice → convert to GBP
        analyst_target = round(_to_gbp(analyst_target_raw), 4) if analyst_target_raw else None

        # Convert book_value / eps3 if company reports in USD
        if financial_currency == "USD":
            bv_gbp  = (book_value / gbp_usd) if book_value else None
            e3_gbp  = (eps3 / gbp_usd)       if eps3 and not np.isnan(eps3) else np.nan
        else:
            bv_gbp  = book_value
            e3_gbp  = eps3

        # ── Route to sub-sector model ──────────────────────────────────────
        dcf_out = None   # financial stocks don't use DCF
        fcf_method = fin_subsector

        if fin_subsector == "Banks":
            ddm_val, pb_val, epv_val, target, fcf_method = val_banks(
                yf_t, ke, div_rate or 0,
                roe or np.nan, bv_gbp or np.nan, e3_gbp
            )
            ddm_out = round(float(ddm_val), 4) if ddm_val and not np.isnan(ddm_val) and ddm_val > 0 else None
            epv_out = round(float(pb_val),  4) if pb_val  and not np.isnan(pb_val)  and pb_val  > 0 else None
            # col K (EPV slot) holds the bank EPV (EPS cap), col J (DDM) holds DDM,
            # we use a separate note in FCF_Flag; col I unused → store pb_val
            # Reassign for column mapping: I=DDM, J=P/B, K=EarningsCap
            ddm_out = round(float(ddm_val), 4) if ddm_val and not np.isnan(ddm_val) and ddm_val > 0 else None
            epv_out = round(float(pb_val),  4) if pb_val  and not np.isnan(pb_val)  and pb_val  > 0 else None

        elif fin_subsector == "General Insurance":
            gi_val, target, fcf_method = val_general_insurance(ke, roe or np.nan, bv_gbp or np.nan)
            ddm_out = None
            epv_out = round(float(gi_val), 4) if gi_val and not np.isnan(gi_val) and gi_val > 0 else None

        elif fin_subsector == "Life Insurance":
            target, fcf_method = val_life_insurance(ke, e3_gbp)
            ddm_out = None
            epv_out = round(float(target), 4) if target and not np.isnan(target) and target > 0 else None

        elif fin_subsector == "Asset Management":
            pb_val, epv_val, target, fcf_method = val_asset_management(
                ke, roe or np.nan, bv_gbp or np.nan, e3_gbp
            )
            ddm_out = round(float(pb_val),  4) if pb_val  and not np.isnan(pb_val)  and pb_val  > 0 else None
            epv_out = round(float(epv_val), 4) if epv_val and not np.isnan(epv_val) and epv_val > 0 else None

        elif fin_subsector == "PE/Alternatives":
            target, fcf_method = val_pe_alternatives(bv_gbp or np.nan, ke, e3_gbp)
            ddm_out = None
            epv_out = round(float(target), 4) if target and not np.isnan(target) and target > 0 else None

        elif fin_subsector == "Capital Markets":
            # ebitda and net_debt: if USD company, convert
            ebitda_gbp   = (ebitda   / gbp_usd) if ebitda   and financial_currency == "USD" else ebitda
            net_debt_gbp = (net_debt / gbp_usd) if net_debt and financial_currency == "USD" else net_debt
            target, fcf_method = val_capital_markets(
                ebitda_gbp or np.nan, net_debt_gbp or np.nan, shares
            )
            ddm_out = None
            epv_out = round(float(target), 4) if target and not np.isnan(target) and target > 0 else None

        else:  # Financial Services catch-all
            ddm_val, epv_val, rim_val, target, fcf_method = val_other_financial(
                ke, div_rate or 0, trail_eps or np.nan, bv_gbp or np.nan, roe or np.nan
            )
            ddm_out = round(float(ddm_val), 4) if ddm_val and not np.isnan(ddm_val) and ddm_val > 0 else None
            epv_out = round(float(epv_val), 4) if epv_val and not np.isnan(epv_val) and epv_val > 0 else None

        # Fallback to analyst consensus if model returned nothing
        if target is None or (isinstance(target, float) and np.isnan(target)):
            target = analyst_target
            if not target:
                fcf_method = "No Valuation"
            else:
                fcf_method += "+Analyst"

        def clean_gbp(v):
            return round(float(v), 4) if v is not None and not np.isnan(float(v or np.nan)) and float(v) > 0 else None

        tgt_out      = clean_gbp(target)
        out_currency = "GBP"

    # ══════════════════════════════════════════════════════════════════════════
    # STANDARD (NON-FINANCIAL) VALUATION  — FTSE + US/Intl
    # ══════════════════════════════════════════════════════════════════════════
    else:
        # ── FCF / DCF ─────────────────────────────────────────────────────────
        # FIX: pass shares (actual count), not shares/price — the old code inflated
        # FCF per share by ~500x for FTSE stocks (price in pence) and ~100x for US stocks.
        fcf, fcf_method, fcf_yrs = get_fcf_per_share(stock, sector, shares, trail_eps or 0)
        dcf_val = np.nan
        if not np.isnan(fcf) and fcf > 0:
            dcf_val = dcf_value(fcf, wacc)

        # ── DDM ───────────────────────────────────────────────────────────────
        ddm_val = ddm_value(div_rate, wacc) if div_rate and div_rate > 0 else np.nan

        # ── EPV ───────────────────────────────────────────────────────────────
        # Use forward PE * eps OR normalised PE multiple * eps
        if fwd_pe and fwd_pe > 0 and fwd_eps and fwd_eps > 0:
            epv_val = fwd_eps * min(fwd_pe, PE_MULTIPLE * 1.5)   # cap at 1.5× normal PE
        else:
            epv_val = epv_value(fwd_eps or trail_eps)

        # ── Currency normalisation + blended target ───────────────────────────
        #
        # yfinance unit conventions for .L (FTSE) stocks:
        #   currentPrice          → GBp (pence)     ← divide by 100 → GBP
        #   targetMeanPrice       → GBp (pence)     ← divide by 100 → GBP
        #   dividendRate          → GBP (pounds)    ← already GBP, no division
        #   forwardEps/trailingEps→ GBP (pounds)    ← already GBP, no division
        #   cashflow statement    → company financial currency (GBP or USD)
        #       primary FCF from USD companies → divide by gbp_usd → GBP
        #       fallback FCF (trail_eps-based) → already GBP
        #
        # For non-FTSE (US/intl): all per-share data is in USD, no conversion needed.

        analyst_target = round(_to_gbp(analyst_target_raw), 4) if (analyst_target_raw and is_ftse) else analyst_target_raw

        if is_ftse:
            financial_currency = (info.get("financialCurrency") or "GBP").upper()
            price_out = round(_to_gbp(price), 4)   # native currency → GBP

            # Convert DCF to GBP based on company's reporting currency
            if not np.isnan(dcf_val) and dcf_val > 0:
                if fcf_method == "primary" and financial_currency == "USD":
                    dcf_val_gbp = dcf_val / gbp_usd
                elif fcf_method == "primary" and financial_currency == "EUR":
                    dcf_val_gbp = dcf_val * eur_gbp
                else:
                    dcf_val_gbp = dcf_val   # GBP or fallback EPS-based (already GBP)
            else:
                dcf_val_gbp = np.nan

            # For sectors where fallback FCF is unreliable (e.g. Basic Materials),
            # suppress the DCF entirely unless primary cashflow data was used.
            if sector in SKIP_FALLBACK_DCF_SECTORS and fcf_method == "fallback":
                dcf_val_gbp = np.nan

            # For sectors where DDM is unreliable (e.g. Basic Materials — commodity
            # miners/traders with variable distributions), suppress DDM from blending.
            if sector in SKIP_DDM_SECTORS:
                ddm_val = np.nan

            # Blend in GBP — outlier-remove symmetrically: drop values > 3× median or < median/3
            vals = [v for v in [dcf_val_gbp, ddm_val, epv_val] if not np.isnan(v) and v > 0]
            if vals:
                if len(vals) > 1:
                    med = float(np.median(vals))
                    vals = [v for v in vals if (med / 3) <= v <= (med * 3)]
                target = float(np.mean(vals)) if vals else (analyst_target or np.nan)
            elif analyst_target:
                target = analyst_target
                fcf_method = "Analyst Consensus"
            else:
                target = np.nan
                fcf_method = "No Valuation"

            def clean_gbp(v):
                return round(float(v), 4) if v is not None and not np.isnan(v) and v > 0 else None

            dcf_out = clean_gbp(dcf_val_gbp)
            ddm_out = clean_gbp(ddm_val)
            epv_out = clean_gbp(epv_val)
            tgt_out = clean_gbp(target) if target and not np.isnan(target) else None
            out_currency = "GBP"

        else:
            # Non-FTSE: all metrics in USD — only the shares bug fix above was needed
            analyst_target = analyst_target_raw
            vals = [v for v in [dcf_val, ddm_val, epv_val] if not np.isnan(v) and v > 0]
            if vals:
                if len(vals) > 1:
                    med = float(np.median(vals))
                    vals = [v for v in vals if v <= med * 3]
                target = float(np.mean(vals)) if vals else (analyst_target or np.nan)
            elif analyst_target:
                target = analyst_target
                fcf_method = "Analyst Consensus"
            else:
                target = np.nan
                fcf_method = "No Valuation"

            def clean(v):
                return round(v, 4) if not np.isnan(v) and v > 0 else None
            price_out  = round(price, 4)
            dcf_out    = clean(dcf_val)
            ddm_out    = clean(ddm_val)
            epv_out    = clean(epv_val)
            tgt_out    = clean(target) if not np.isnan(target) else None
            out_currency = "USD"

    value_ratio = round(tgt_out / price_out, 2) if tgt_out and price_out else None

    # Signal
    if value_ratio:
        if   value_ratio >= 1.25: signal = "Strong Buy"
        elif value_ratio >= 1.10: signal = "Buy"
        elif value_ratio >= 0.90: signal = "Fair Value"
        elif value_ratio >= 0.75: signal = "Sell"
        else:                     signal = "Strong Sell"
    else:
        signal = "No Signal"

    # Sanity check: if target < price, signal must never be Buy or Strong Buy
    if value_ratio and tgt_out and price_out and tgt_out < price_out:
        assert signal not in ("Strong Buy", "Buy"), (
            f"Signal inversion detected for {yf_t}: target={tgt_out} < price={price_out} "
            f"but signal='{signal}' (value_ratio={value_ratio}). "
            f"Check that value_ratio = target/price, not price/target."
        )

    yield_ = round(div_rate / price * 100, 2) if div_rate and price else 0
    subsector_label = fin_subsector or ""

    log(f"  {yf_t:<12} [{subsector_label or 'Standard':<20}] price={price_out}  target={tgt_out}  ratio={value_ratio}  → {signal}")

    return {
        "Ticker":        yf_t,
        "etoro_ticker":  row["etoro_ticker"],
        "Name":          company,
        "Sector":        sector,
        "Fin_Subsector": subsector_label,
        "Beta":          round(info.get("beta"), 3) if info.get("beta") else None,
        "WACC":          f"{wacc*100:.2f}%",
        "Ke":            f"{ke*100:.2f}%",
        "Current Price": price_out,
        "raw_price":     round(_to_gbp(price) * 100, 4) if (is_ftse and currency and currency != "GBp") else price,  # always GBp for FTSE, USD for US
        "price_currency": currency,
        "Target Price":  tgt_out,
        "Value_Ratio":   value_ratio,
        "Signal":        signal,
        "Int_Val_DCF":   dcf_out,
        "Int Val Div":   ddm_out,
        "Int Val PE":    epv_out,
        "Analyst_Target": analyst_target_raw,
        "Yield":         f"{yield_:.2f}%",
        "FCF_Flag":      fcf_method,
        "Currency":      out_currency,
        "is_ftse":       is_ftse,
    }


# ── Update Assumptions sheet ─────────────────────────────────────────────────
def update_assumptions(wb, results: list[dict]):
    ws = wb["Assumptions"]
    gbp_usd = read_gbp_usd(wb)

    # Build ticker → row map
    ticker_row = {}
    for row in ws.iter_rows(min_row=7, max_row=300):
        if row[0].value:
            ticker_row[str(row[0].value).strip()] = row[0].row

    now_str   = datetime.now().strftime("%Y-%m-%d %H:%M")
    updated   = 0
    skipped   = []

    # ── Step 1: Roll current signal → Previous Signal (col P) BEFORE updating ──
    # This preserves what the signal was before today's run, enabling Daily Summary
    # to highlight changes. Only roll if Current Signal (col Q) is already populated.
    rolled = 0
    for row in ws.iter_rows(min_row=7, max_row=300):
        q_cell = row[16] if len(row) > 16 else None   # col Q = index 16 (1-based col 17)
        p_cell = row[15] if len(row) > 15 else None   # col P = index 15 (1-based col 16)
        if q_cell and p_cell and q_cell.value:
            p_cell.value = q_cell.value
            rolled += 1
    log(f"Assumptions: rolled {rolled} signals P←Q (prev←current)")

    for r in results:
        etoro_t = r["etoro_ticker"]
        row_num = ticker_row.get(etoro_t)
        if not row_num:
            skipped.append(etoro_t)
            continue

        is_ftse  = r["is_ftse"]
        fmt      = "£#,##0.00" if is_ftse else "$#,##0.00"

        # col B = Company Name
        name = r.get("Name", "")
        if name:
            c = ws.cell(row=row_num, column=2, value=name)
            c.font = NORMAL_FONT

        # col C = Sector (prefer financial sub-sector label for financial stocks)
        sector_label = r.get("Fin_Subsector") or r.get("Sector", "")
        if sector_label:
            c = ws.cell(row=row_num, column=3, value=sector_label)
            c.font = NORMAL_FONT

        # col D = Beta
        beta_val = r.get("Beta")
        if beta_val is not None:
            c = ws.cell(row=row_num, column=4, value=beta_val)
            c.font = NORMAL_FONT
            c.number_format = "0.000"

        wacc_str = r.get("WACC","")
        try:
            wacc_val = float(wacc_str.replace("%","")) / 100
            _set_blue(ws, row_num, 5, wacc_val, "0.00%")  # E
        except Exception:
            pass

        for col, key in [(9,"Int_Val_DCF"), (10,"Int Val Div"), (11,"Int Val PE")]:
            v = r.get(key)
            if v is not None:
                _set_blue(ws, row_num, col, v, fmt)

        # col L = Python/Analyst target
        tgt = r.get("Target Price")
        if tgt is not None:
            _set_blue(ws, row_num, 12, tgt, fmt)

        # col M = FCF method
        c = ws.cell(row=row_num, column=13, value=r.get("FCF_Flag",""))
        c.font = INPUT_BLUE; c.alignment = CENTER

        # col N = last updated
        c = ws.cell(row=row_num, column=14, value=now_str)
        c.font = UPDATED_GREEN; c.alignment = CENTER

        # ── Step 2: Write Current Signal (col Q) AFTER updating values ──
        signal = r.get("Signal", "")
        if signal:
            c = ws.cell(row=row_num, column=17, value=signal)
            c.font  = UPDATED_GREEN
            c.alignment = CENTER

        updated += 1

    log(f"Assumptions: updated {updated} rows  |  skipped (not in sheet): {skipped[:10]}")


# ── Save CSVs ─────────────────────────────────────────────────────────────────
def save_csvs(results: list[dict]):
    ftse_rows = [r for r in results if r["is_ftse"]]
    intl_rows = [r for r in results if not r["is_ftse"]]

    ftse_cols = ["Ticker","Name","Fin_Subsector","WACC","Ke","Int Val Div","Int Val PE","Int_Val_DCF",
                 "Target Price","Current Price","Value_Ratio","Yield","Signal","FCF_Flag"]
    intl_cols = ["Ticker","Name","Sector","Current Price","Target Price","Value_Ratio",
                 "Signal","FCF_Flag","Yield","Analyst_Target","WACC"]

    def _write(path, rows, cols):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)
        log(f"Saved {path.name} ({len(rows)} rows)")

    _write(REPORTS_DIR / "ftse_report.csv",  ftse_rows, ftse_cols)
    _write(REPORTS_DIR / "intl_report.csv",  intl_rows, intl_cols)

    # Mirror to OneDrive if available
    try:
        ONEDRIVE_DIR.mkdir(parents=True, exist_ok=True)
        _write(ONEDRIVE_DIR / "ftse_report.csv", ftse_rows, ftse_cols)
        _write(ONEDRIVE_DIR / "intl_report.csv", intl_rows, intl_cols)
    except Exception as e:
        log(f"OneDrive save skipped: {e}")


# ── Discord notification ──────────────────────────────────────────────────────
def notify_discord(results: list[dict]):
    if not DISCORD_WEBHOOK:
        return
    strong_buys  = [r["Ticker"] for r in results if r["Signal"] == "Strong Buy"]
    strong_sells = [r["Ticker"] for r in results if r["Signal"] == "Strong Sell"]
    msg = (
        f"**Valuation Run Complete** — {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        f"Tickers valued: {len(results)}\n"
        f"Strong Buys ({len(strong_buys)}): {', '.join(strong_buys[:10])}\n"
        f"Strong Sells ({len(strong_sells)}): {', '.join(strong_sells[:10])}"
    )
    try:
        requests.post(DISCORD_WEBHOOK, json={"content": msg, "username": "Valuation Bot"}, timeout=10)
    except Exception as e:
        log(f"Discord notification failed: {e}")


# ── Fallback price updater ────────────────────────────────────────────────────
# Maps eToro tickers that need a different Yahoo Finance symbol for price lookup
PRICE_YF_OVERRIDE = {
    "BTC":  "BTC-USD",
    "ETH":  "ETH-USD",
    "SOL":  "SOL-USD",
}

def update_fallback_prices(wb, results: list[dict]):
    """
    Writes live prices into column N (Manual Price Override) of the Tickers sheet
    for every ticker — both those processed by valuation.py and those skipped
    (ETFs, Crypto, Bonds, Cash).

    The Portfolio L formula checks Tickers!N first before falling back to
    STOCKHISTORY, so these values ensure prices display correctly even when
    Microsoft's data feed is unavailable or returns bad data.

    Price units written to Tickers N:
      FTSE (.L) stocks → GBp (pence)  — same unit as STOCKHISTORY returns
      US / intl stocks → USD
      Crypto (BTC etc.)→ USD
      ETFs (.L)        → GBp (pence)

    Results from the valuation loop already carry raw_price (the yfinance price
    before any unit conversion), so no extra network calls are needed for those
    tickers.  Skipped tickers (ETFs, Crypto, Bonds) are fetched separately.
    """
    ws = wb["Tickers"]

    # ── Build lookup: yf_ticker → (row_number, market, asset_type) ────────────
    ticker_info = {}   # yf_ticker (col F) → (row, market, asset_type)
    for row in ws.iter_rows(min_row=3, max_row=400):
        yf_t    = str(row[5].value or "").strip()   # col F = Yahoo Finance Ticker
        market  = str(row[6].value or "").strip()   # col G = Market
        asset_t = str(row[8].value or "").strip()   # col I = Asset Type
        etoro_t = str(row[3].value or "").strip()   # col D = eToro Ticker
        if yf_t:
            ticker_info[yf_t] = {
                "row":      row[5].row,
                "market":   market,
                "asset":    asset_t,
                "etoro":    etoro_t,
            }

    if not ticker_info:
        log("update_fallback_prices: no tickers found in Tickers sheet")
        return

    written = 0
    fetched = 0
    skipped = 0

    # ── Step 1: write prices for tickers already processed by valuation loop ──
    # raw_price = original yfinance price (GBp for FTSE, USD for US/intl)
    results_map = {r["Ticker"]: r["raw_price"] for r in results if r.get("raw_price")}
    for yf_t, info in ticker_info.items():
        if yf_t in results_map:
            raw = results_map[yf_t]
            if raw and raw > 0:
                ws.cell(row=info["row"], column=14).value = round(float(raw), 4)
                written += 1

    log(f"  Fallback prices from valuation results: {written} tickers written")

    # ── Step 2: fetch prices for SKIPPED tickers (ETF, Crypto, Bond, Cash) ────
    skipped_rows = {
        yf_t: info for yf_t, info in ticker_info.items()
        if yf_t not in results_map
    }
    log(f"  Fetching prices for {len(skipped_rows)} skipped tickers (ETFs, Crypto, etc.)...")

    for yf_t, info in skipped_rows.items():
        etoro_t = info["etoro"]
        market  = info["market"]
        asset_t = info["asset"]

        # Resolve the actual yfinance symbol (handle crypto overrides)
        yf_sym = PRICE_YF_OVERRIDE.get(etoro_t) or PRICE_YF_OVERRIDE.get(yf_t) or yf_t
        if not yf_sym:
            skipped += 1
            continue

        try:
            stock = yf.Ticker(yf_sym)
            price = stock.fast_info.last_price
            if not price or price <= 0:
                skipped += 1
                continue

            # For FTSE ETFs/bonds: yfinance returns GBp (pence) — write as-is
            # For US/crypto: yfinance returns USD — write as-is
            ws.cell(row=info["row"], column=14).value = round(float(price), 4)
            fetched += 1
            time.sleep(0.2)   # polite delay for skipped tickers

        except Exception as e:
            log(f"  Price fetch failed for {yf_t} ({yf_sym}): {e}")
            skipped += 1

    log(f"  Fallback prices fetched for skipped tickers: {fetched} written, {skipped} unavailable")
    log(f"  Tickers N column updated: {written + fetched} total prices written")


# ── Update Tickers sheet metadata (Company Name + Sector) ────────────────────
def update_tickers_metadata(wb, results: list[dict]):
    """
    Writes Company Name (col B) and Sector (col H) back to the Tickers sheet
    for all tickers — processed tickers use data already fetched by the valuation
    loop; skipped tickers (ETFs, Crypto, Bonds) are fetched separately.
    """
    ws = wb["Tickers"]

    # Build lookup: yf_ticker → row_number
    ticker_rows = {}
    for row in ws.iter_rows(min_row=3, max_row=400):
        yf_t = str(row[5].value or "").strip()   # col F = Yahoo Finance Ticker
        if yf_t:
            ticker_rows[yf_t] = row[5].row

    if not ticker_rows:
        log("update_tickers_metadata: no tickers found in Tickers sheet")
        return

    updated = 0

    # ── Step 1: write metadata for tickers already processed by valuation loop ──
    results_map = {r["Ticker"]: r for r in results}
    for yf_t, row_num in ticker_rows.items():
        r = results_map.get(yf_t)
        if not r:
            continue
        name   = r.get("Name", "")
        sector = r.get("Sector", "")
        if name:
            c = ws.cell(row=row_num, column=2, value=name)
            c.font = NORMAL_FONT
        if sector:
            c = ws.cell(row=row_num, column=8, value=sector)
            c.font = NORMAL_FONT
        if name or sector:
            updated += 1

    log(f"  Tickers metadata: {updated} rows updated from valuation results")

    # ── Step 2: fetch metadata for skipped tickers (ETFs, Crypto, Bonds, Cash) ──
    skipped_tickers = [yf_t for yf_t in ticker_rows if yf_t not in results_map]
    log(f"  Fetching metadata for {len(skipped_tickers)} skipped tickers...")
    fetched = 0
    for yf_t in skipped_tickers:
        row_num = ticker_rows[yf_t]
        try:
            stock = yf.Ticker(yf_t)
            info  = stock.info or {}
            name   = info.get("longName") or info.get("shortName") or ""
            sector = info.get("sector") or ""
            if name:
                c = ws.cell(row=row_num, column=2, value=name)
                c.font = NORMAL_FONT
            if sector:
                c = ws.cell(row=row_num, column=8, value=sector)
                c.font = NORMAL_FONT
            if name or sector:
                fetched += 1
            time.sleep(0.3)
        except Exception as e:
            log(f"  Metadata fetch failed for {yf_t}: {e}")

    log(f"  Tickers metadata: {fetched} skipped tickers updated from yfinance")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    log("=" * 60)
    log("valuation.py  starting")
    log("=" * 60)

    if not MASTER.exists():
        log(f"ERROR: {MASTER} not found")
        sys.exit(1)

    wb      = openpyxl.load_workbook(str(MASTER))
    gbp_usd = read_gbp_usd(wb)
    log(f"GBP/USD = {gbp_usd}")
    eur_gbp = fetch_eur_gbp()

    tickers = load_tickers(wb)
    log(f"Processing {len(tickers)} tickers...\n")

    results = []
    for i, row in enumerate(tickers, 1):
        log(f"[{i}/{len(tickers)}] {row['yf_ticker']}")
        result = value_ticker(row, gbp_usd, eur_gbp)
        if result:
            results.append(result)
        time.sleep(0.4)   # be polite to Yahoo Finance

    log(f"\nValuation complete: {len(results)}/{len(tickers)} tickers valued")

    # Update Assumptions sheet
    update_assumptions(wb, results)

    # Write live prices to Tickers!N (Manual Price Override) for ALL tickers.
    # Processed tickers use the price already fetched above (no extra calls).
    # Skipped tickers (ETFs, Crypto, Bonds) are fetched separately.
    # Portfolio L reads Tickers!N first, so this acts as a reliable fallback
    # whenever STOCKHISTORY returns bad/stale data.
    log("Writing fallback prices to Tickers!N...")
    update_fallback_prices(wb, results)

    # Write Company Name and Sector back to Tickers sheet for all tickers
    log("Writing Company Name and Sector to Tickers sheet...")
    update_tickers_metadata(wb, results)

    # Save workbook
    wb.save(str(MASTER))
    log(f"Saved → {MASTER}")

    # Save CSVs
    save_csvs(results)

    # Discord
    notify_discord(results)

    # Summary
    signals = {}
    for r in results:
        signals[r["Signal"]] = signals.get(r["Signal"], 0) + 1
    log("\nSignal summary:")
    for sig, count in sorted(signals.items(), key=lambda x: x[1], reverse=True):
        log(f"  {sig:<15} {count}")

    log("\nvaluation.py complete.")


if __name__ == "__main__":
    main()
