#!/usr/bin/env python3
"""
snapshot_sectors.py

Reads eToro_Master.xlsx (or most-recent .bak if live file is corrupt), aggregates
covered stocks by sector, and appends one row per sector per snapshot date to
sector_history.csv. Also appends per-stock rows to per_stock_history.csv.

Idempotent: if a row for (snapshot_date, sector) already exists, it is replaced
rather than duplicated. Re-running on the same day = same output.

Usage:
    python snapshot_sectors.py                # snapshot today
    python snapshot_sectors.py --dry-run      # print, don't write
    python snapshot_sectors.py --date 2026-04-25  # backfill an older date

Outputs land in:
    data/sector_history/sector_history.csv
    data/sector_history/per_stock_history.csv

Owner: Neil Daley. Methodology source-of-truth: the eToro Master spreadsheet.
"""

from __future__ import annotations

import argparse
import csv
import glob
import json
import os
import statistics
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import date as _date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    # openpyxl is only required by the xlsx fallback path; the primary JSON path
    # works without it.
    openpyxl = None  # type: ignore


# ---------------------------------------------------------------------------
# Paths (resolved relative to this script's location)
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
DATA_DIR = PROJECT_ROOT / "data"
HISTORY_DIR = DATA_DIR / "sector_history"
SECTOR_HISTORY_CSV = HISTORY_DIR / "sector_history.csv"
PER_STOCK_HISTORY_CSV = HISTORY_DIR / "per_stock_history.csv"

LIVE_JSON = DATA_DIR / "etoro_master.json"
LIVE_XLSX = DATA_DIR / "eToro_Master.xlsx"
BAK_GLOB = str(DATA_DIR / "eToro_Master.bak-*.xlsx")

# Sectors that don't have valuation models and so don't belong in the aggregate
EXCLUDE_SECTORS = {"Corp Bonds", "Crypto"}

# Output schema
SECTOR_FIELDS = [
    "snapshot_date", "sector", "n",
    "mean_value_ratio", "median_value_ratio", "std_value_ratio",
    "min_vr", "max_vr",
    "n_strong_buy", "n_buy", "n_fair_value", "n_sell", "n_strong_sell",
    "pct_buy_or_strong_buy", "pct_sell_or_strong_sell",
    "source_file",
]
PER_STOCK_FIELDS = [
    "snapshot_date", "ticker", "name", "sector",
    "value_ratio", "signal", "source_file",
]


# ---------------------------------------------------------------------------
# File loading with corrupt-file fallback
# ---------------------------------------------------------------------------
def is_valid_xlsx(path: Path) -> bool:
    """Quick check: does this file open as a zip (xlsx is a zip)?"""
    try:
        with zipfile.ZipFile(path, "r") as z:
            z.namelist()
        return True
    except (zipfile.BadZipFile, OSError):
        return False


def _load_json_robust(path: Path) -> dict:
    """Parse etoro_master.json, tolerating trailing null bytes / extra data from
    older non-atomic writes. Mirrors the raw_decode approach used by build_site.py."""
    with open(path, encoding="utf-8") as f:
        raw = f.read()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        decoder = json.JSONDecoder()
        d, _ = decoder.raw_decode(raw)
        return d


def is_valid_json(path: Path) -> bool:
    try:
        _load_json_robust(path)
        return True
    except Exception:
        return False


def pick_source_file() -> Path:
    """Prefer the canonical JSON; fall back to live xlsx; then most recent .bak."""
    if LIVE_JSON.exists() and is_valid_json(LIVE_JSON):
        return LIVE_JSON
    if LIVE_JSON.exists():
        print(f"WARN: {LIVE_JSON.name} is corrupt, falling back to xlsx")
    else:
        print(f"WARN: {LIVE_JSON.name} not found, falling back to xlsx")
    if openpyxl is None:
        sys.exit("ERROR: JSON unavailable and openpyxl not installed. "
                 "Run: pip install openpyxl --break-system-packages")
    if LIVE_XLSX.exists() and is_valid_xlsx(LIVE_XLSX):
        return LIVE_XLSX
    print(f"WARN: {LIVE_XLSX.name} is missing or corrupt, falling back to backup")
    candidates = sorted(glob.glob(BAK_GLOB), reverse=True)
    for c in candidates:
        p = Path(c)
        if is_valid_xlsx(p):
            print(f"  using: {p.name}")
            return p
    sys.exit("ERROR: No valid etoro_master.json, eToro_Master.xlsx, or .bak-*.xlsx found")


# ---------------------------------------------------------------------------
# Spreadsheet extraction
# ---------------------------------------------------------------------------
def _extract_rows_from_json(json_path: Path) -> list[dict]:
    """Read Portfolio + Watchlist from etoro_master.json. Output shape matches
    the xlsx extractor."""
    data = _load_json_robust(json_path)
    sheets = data.get("sheets", {})
    out: list[dict] = []
    for obj in sheets.get("portfolio", {}).get("objects", []):
        ticker = (obj.get("eToro Ticker") or "").strip()
        if not ticker:
            continue
        out.append({
            "ticker": ticker,
            "name":   obj.get("Company Name") or "",
            "sector": obj.get("Sector") or "",
            "value_ratio": obj.get("Value Ratio"),
            "signal": obj.get("Signal"),
        })
    for obj in sheets.get("watchlist", {}).get("objects", []):
        ticker = (obj.get("eToro Ticker") or "").strip()
        if not ticker:
            continue
        out.append({
            "ticker": ticker,
            "name":   obj.get("Company / Name") or "",
            "sector": obj.get("Sector") or "",
            "value_ratio": obj.get("Value Ratio"),
            "signal": obj.get("Signal"),
        })
    # Coerce value_ratio strings to floats where possible (JSON cache stringifies numerics)
    for r in out:
        vr = r["value_ratio"]
        if isinstance(vr, str):
            try:
                r["value_ratio"] = float(vr.replace(",", ""))
            except ValueError:
                r["value_ratio"] = None
    return out


def _extract_rows_from_xlsx(xlsx_path: Path) -> list[dict]:
    """Fallback xlsx reader. Only used when the JSON cache is unavailable."""
    if openpyxl is None:
        sys.exit("openpyxl not installed and JSON unavailable. Run: pip install openpyxl")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    out: list[dict] = []
    # Portfolio sheet: Sector=col 2, eToro Ticker=col 3, Name=col 1, Value Ratio=col 29, Signal=col 30
    if "Portfolio" in wb.sheetnames:
        ws = wb["Portfolio"]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2:  # skip title + header
                continue
            if not row[0] or not row[3]:
                continue
            out.append({
                "ticker": row[3],
                "name": row[1],
                "sector": row[2],
                "value_ratio": row[29],
                "signal": row[30],
            })
    # Watchlist sheet: Sector=2, eToro Ticker=3, Name=1, Value Ratio=11, Signal=12
    if "Watchlist" in wb.sheetnames:
        ws = wb["Watchlist"]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2:
                continue
            if not row[0] or not row[3]:
                continue
            out.append({
                "ticker": row[3],
                "name": row[1],
                "sector": row[2],
                "value_ratio": row[11],
                "signal": row[12],
            })
    wb.close()
    return out


def extract_rows(source_path: Path) -> list[dict]:
    """Dispatch to JSON or xlsx extractor based on file extension."""
    if source_path.suffix.lower() == ".json":
        return _extract_rows_from_json(source_path)
    return _extract_rows_from_xlsx(source_path)


def is_usable(r: dict) -> bool:
    vr = r.get("value_ratio")
    if not isinstance(vr, (int, float)) or vr is None or vr <= 0:
        return False
    if r.get("sector") in EXCLUDE_SECTORS:
        return False
    return True


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------
def aggregate_by_sector(rows: list[dict], snapshot_date: str, source_name: str) -> list[dict]:
    by_sector = defaultdict(list)
    for r in rows:
        by_sector[r["sector"]].append(r)
    results = []
    for sector, srows in by_sector.items():
        vrs = [float(r["value_ratio"]) for r in srows]
        sig_counter = Counter((r["signal"] or "").strip() for r in srows)
        n = len(srows)
        n_strong_buy = sig_counter.get("Strong Buy", 0)
        n_buy = sig_counter.get("Buy", 0)
        n_fv = sig_counter.get("Fair Value", 0)
        n_sell = sig_counter.get("Sell", 0)
        n_strong_sell = sig_counter.get("Strong Sell", 0)
        results.append({
            "snapshot_date": snapshot_date,
            "sector": sector,
            "n": n,
            "mean_value_ratio": round(statistics.mean(vrs), 3),
            "median_value_ratio": round(statistics.median(vrs), 3),
            "std_value_ratio": round(statistics.pstdev(vrs) if n > 1 else 0, 3),
            "min_vr": round(min(vrs), 3),
            "max_vr": round(max(vrs), 3),
            "n_strong_buy": n_strong_buy,
            "n_buy": n_buy,
            "n_fair_value": n_fv,
            "n_sell": n_sell,
            "n_strong_sell": n_strong_sell,
            "pct_buy_or_strong_buy": round((n_strong_buy + n_buy) / n, 3),
            "pct_sell_or_strong_sell": round((n_strong_sell + n_sell) / n, 3),
            "source_file": source_name,
        })
    results.sort(key=lambda r: -r["mean_value_ratio"])  # cheapest first
    return results


def per_stock_rows(rows: list[dict], snapshot_date: str, source_name: str) -> list[dict]:
    out = []
    for r in rows:
        out.append({
            "snapshot_date": snapshot_date,
            "ticker": r["ticker"],
            "name": r["name"],
            "sector": r["sector"],
            "value_ratio": round(float(r["value_ratio"]), 3),
            "signal": r["signal"],
            "source_file": source_name,
        })
    return out


# ---------------------------------------------------------------------------
# Idempotent CSV append: replace any rows matching this snapshot_date
# ---------------------------------------------------------------------------
def upsert_csv(csv_path: Path, fieldnames: list[str], new_rows: list[dict],
               snapshot_date: str) -> tuple[int, int]:
    """Replace all rows for snapshot_date with new_rows. Returns (kept, written)."""
    existing: list[dict] = []
    if csv_path.exists():
        with open(csv_path, newline="") as f:
            existing = [r for r in csv.DictReader(f) if r.get("snapshot_date") != snapshot_date]
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in existing:
            # tolerate older files missing newer columns
            w.writerow({k: r.get(k, "") for k in fieldnames})
        for r in new_rows:
            w.writerow(r)
    return len(existing), len(new_rows)


# ---------------------------------------------------------------------------
# Pretty-print
# ---------------------------------------------------------------------------
def print_snapshot(results: list[dict], snapshot_date: str, source_name: str):
    print(f"\n=== SECTOR VALUATION SNAPSHOT — {snapshot_date} ===")
    print(f"Source: {source_name}")
    print("Value Ratio = blended target / live price. Higher = cheaper.\n")
    print(f"{'Sector':<24} {'n':>3}  {'Mean VR':>8} {'Med VR':>8} {'σ VR':>6}  "
          f"{'Min':>6} {'Max':>6}  {'%Buy':>5} {'%Sell':>6}")
    print("-" * 95)
    for r in results:
        flag = " ⚠" if r["n"] < 5 else ""
        print(f"{r['sector']:<24} {r['n']:>3}  {r['mean_value_ratio']:>8.3f} "
              f"{r['median_value_ratio']:>8.3f} {r['std_value_ratio']:>6.3f}  "
              f"{r['min_vr']:>6.2f} {r['max_vr']:>6.2f}  "
              f"{r['pct_buy_or_strong_buy']*100:>4.0f}% "
              f"{r['pct_sell_or_strong_sell']*100:>5.0f}%{flag}")
    n_total = sum(r["n"] for r in results)
    print(f"\nUniverse: {n_total} stocks across {len(results)} sectors "
          f"(excl. {', '.join(sorted(EXCLUDE_SECTORS))})")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Snapshot sector valuations to CSV history.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Print snapshot but don't write to CSVs")
    ap.add_argument("--date", default=_date.today().isoformat(),
                    help="Snapshot date (YYYY-MM-DD), defaults to today")
    args = ap.parse_args()

    src = pick_source_file()
    rows_raw = extract_rows(src)
    rows = [r for r in rows_raw if is_usable(r)]

    sector_results = aggregate_by_sector(rows, args.date, src.name)
    stock_results = per_stock_rows(rows, args.date, src.name)
    print_snapshot(sector_results, args.date, src.name)
    print(f"\nUsable rows: {len(rows)} / {len(rows_raw)} "
          f"(filtered out non-numeric VR or excluded sectors)")

    if args.dry_run:
        print("\n[dry-run] No files written.")
        return

    kept_s, wrote_s = upsert_csv(SECTOR_HISTORY_CSV, SECTOR_FIELDS, sector_results, args.date)
    kept_p, wrote_p = upsert_csv(PER_STOCK_HISTORY_CSV, PER_STOCK_FIELDS, stock_results, args.date)
    print(f"\nWrote {wrote_s} sector rows + {wrote_p} per-stock rows")
    print(f"  sector_history.csv: kept {kept_s} prior rows + wrote {wrote_s}")
    print(f"  per_stock_history.csv: kept {kept_p} prior rows + wrote {wrote_p}")
    print(f"\nFiles updated:")
    print(f"  {SECTOR_HISTORY_CSV}")
    print(f"  {PER_STOCK_HISTORY_CSV}")


if __name__ == "__main__":
    main()
