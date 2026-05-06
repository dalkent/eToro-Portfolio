#!/usr/bin/env python3
"""
reconcile_etoro_closed.py
─────────────────────────
Pulls the full closed-trades history from eToro's public API
(`/trading/history/positions`) and diffs it against the Closed Positions sheet
in `data/eToro_Master.xlsx`. Prints:

  • totals from the API vs the sheet
  • trades that are on the API but missing from the sheet
  • trades on the sheet that the API doesn't know about
  • a CSV dump of the API's view at data/etoro_closed_positions_api.csv

This is READ-ONLY against the Excel — it doesn't auto-edit the sheet. Review
the report first, then if you want to auto-populate, we'll add that next.

Environment:
    ETORO_PUBLIC_API_KEY
    ETORO_USER_KEY
"""

import csv
import os
import sys
import uuid
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# Load etoro.env from project root before reading os.getenv anywhere below.
sys.path.insert(0, str(Path(__file__).resolve().parent))
import _envloader  # noqa: F401  (side-effect import: populates os.environ)

import openpyxl
import requests

BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data"
LOGS_DIR = BASE_DIR / "logs"
MASTER   = DATA_DIR / "eToro_Master.xlsx"
MATCH_CSV= DATA_DIR / "etoro_portfolio_tickermatch.csv"
OUT_CSV  = DATA_DIR / "etoro_closed_positions_api.csv"
LOG_FILE = LOGS_DIR / "reconcile_etoro_closed.log"

LOGS_DIR.mkdir(exist_ok=True)


def _load_env_file(path: Path):
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        os.environ.setdefault(k.strip(), v.strip())


# Load etoro.env if vars aren't already in the environment
_load_env_file(BASE_DIR / "etoro.env")

PUBLIC_API_KEY = os.getenv("ETORO_PUBLIC_API_KEY")
USER_KEY       = os.getenv("ETORO_USER_KEY")
BASE_URL       = "https://public-api.etoro.com/api/v1"


def log(msg: str):
    line = f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {msg}"
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")
    print(line)


def etoro_headers():
    return {
        "x-api-key":    PUBLIC_API_KEY,
        "x-user-key":   USER_KEY,
        "x-request-id": str(uuid.uuid4()),
        "Accept":       "application/json",
    }


def load_id_to_ticker() -> dict:
    """{asset_id (int): ticker (str)} from the mapping CSV."""
    m = {}
    if not MATCH_CSV.exists():
        log(f"  WARN: {MATCH_CSV} not found — API trades will show Asset IDs instead of tickers")
        return m
    with open(MATCH_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            try:
                m[int(row["Asset_ID"])] = row["Ticker"].strip()
            except (KeyError, ValueError):
                continue
    return m


def fetch_all_closed() -> list:
    """Paginate through /trading/history/positions and return all rows."""
    if not PUBLIC_API_KEY or not USER_KEY:
        log("ERROR: ETORO_PUBLIC_API_KEY / ETORO_USER_KEY not set")
        sys.exit(1)

    # eToro's documented closed-trades endpoint is /trading/real/history.
    # (/trading/history/positions returns 404 — older path that no longer resolves.)
    # Try a few candidate paths to be resilient to API versioning.
    candidate_paths = [
        "/trading/real/history",
        "/trading/history",
        "/trading/history/positions",
    ]
    all_rows = []
    working_path = None
    for path in candidate_paths:
        url = f"{BASE_URL}{path}"
        probe = requests.get(url, headers=etoro_headers(), timeout=15,
                             params={"PageSize": 10, "PageNumber": 1})
        if probe.status_code == 200:
            working_path = path
            log(f"  using endpoint: {path}")
            # include the probe's payload so we don't refetch page 1
            data = probe.json()
            batch = (data.get("PublicHistoryPositions")
                     or data.get("publicHistoryPositions")
                     or data.get("positions")
                     or data.get("history")
                     or data.get("items")
                     or [])
            if isinstance(batch, list):
                all_rows.extend(batch)
            break
        log(f"  probe {path}: HTTP {probe.status_code} — {probe.text[:200]!r}")

    if not working_path:
        log("  no working closed-trades endpoint found — exiting")
        return []

    # Now paginate from page 2 onwards at full size
    url = f"{BASE_URL}{working_path}"
    page = 2
    while page < 200:
        resp = requests.get(url, headers=etoro_headers(), timeout=20,
                            params={"PageSize": 100, "PageNumber": page})
        if resp.status_code != 200:
            log(f"  page {page}: HTTP {resp.status_code} — stopping")
            break
        data = resp.json()
        batch = (data.get("PublicHistoryPositions")
                 or data.get("publicHistoryPositions")
                 or data.get("positions")
                 or data.get("history")
                 or data.get("items")
                 or [])
        if not batch:
            break
        all_rows.extend(batch)
        log(f"  page {page}: +{len(batch)} (total {len(all_rows)})")
        if len(batch) < 100:
            break
        page += 1
    return all_rows


def normalise_close(t: dict, id_to_ticker: dict) -> dict:
    """Turn one API row into a flat dict with consistent keys."""
    iid = t.get("InstrumentID") or t.get("instrumentID")
    ticker = id_to_ticker.get(int(iid), f"ID_{iid}") if iid else "?"
    open_rate  = t.get("OpenRate")  or t.get("openRate")
    close_rate = t.get("CloseRate") or t.get("closeRate")
    units      = t.get("Units")     or t.get("units")      or 0
    amount     = t.get("Amount")    or t.get("amount")     or 0
    net_pnl    = t.get("NetProfit") or t.get("netProfit")  or t.get("Profit") or t.get("profit")
    close_dt   = (t.get("CloseDateTime") or t.get("closeDateTime") or
                  t.get("ClosedDate") or "")
    open_dt    = (t.get("OpenDateTime") or t.get("openDateTime") or
                  t.get("OpenedDate") or "")
    return {
        "ticker":     ticker,
        "asset_id":   int(iid) if iid else None,
        "units":      float(units),
        "invested":   float(amount),      # USD
        "open_rate":  float(open_rate)  if open_rate  else None,
        "close_rate": float(close_rate) if close_rate else None,
        "net_pnl":    float(net_pnl)    if net_pnl is not None else None,
        "open_date":  (open_dt  or "")[:10],
        "close_date": (close_dt or "")[:10],
    }


def load_sheet_closed() -> list:
    """Read existing Closed Positions sheet rows."""
    if not MASTER.exists():
        return []
    wb = openpyxl.load_workbook(str(MASTER), data_only=True)
    if "Closed Positions" not in wb.sheetnames:
        return []
    ws = wb["Closed Positions"]
    rows = []
    for row in ws.iter_rows(min_row=3, max_row=500, values_only=True):
        ticker = str(row[1] or "").strip()
        if not ticker:
            continue
        # Skip summary/footer rows like "TOTALS", "GRAND TOTAL"
        if ticker.upper() in {"TOTALS", "TOTAL", "GRAND TOTAL", "SUMMARY"}:
            continue
        try:
            units    = float(row[3] or 0)
            invested = float(row[4] or 0)
        except Exception:
            continue
        date_sold = row[16]
        if isinstance(date_sold, datetime):
            date_sold = date_sold.strftime("%Y-%m-%d")
        else:
            date_sold = str(date_sold or "")[:10]
        rows.append({
            "ticker":     ticker,
            "units":      units,
            "invested":   invested,
            "close_date": date_sold,
        })
    return rows


def main():
    log("Pulling closed trades from eToro API ...")
    raw = fetch_all_closed()
    log(f"  API returned {len(raw)} closed trades total")

    id_to_ticker = load_id_to_ticker()
    api_rows = [normalise_close(t, id_to_ticker) for t in raw]

    # Group API closures by ticker to compare against sheet (sheet rows are
    # already aggregated per ticker)
    api_by_ticker = defaultdict(lambda: {"units": 0.0, "invested": 0.0,
                                         "net_pnl": 0.0, "count": 0,
                                         "last_close": ""})
    api_total_pnl = 0.0
    api_total_invested = 0.0
    for r in api_rows:
        g = api_by_ticker[r["ticker"]]
        g["units"]       += r["units"]
        g["invested"]    += r["invested"]
        if r["net_pnl"] is not None:
            g["net_pnl"] += r["net_pnl"]
            api_total_pnl += r["net_pnl"]
        g["count"]       += 1
        if r["close_date"] > g["last_close"]:
            g["last_close"] = r["close_date"]
        api_total_invested += r["invested"]

    sheet_rows = load_sheet_closed()
    sheet_by_ticker = defaultdict(lambda: {"units": 0.0, "invested": 0.0})
    sheet_total_invested = 0.0
    for s in sheet_rows:
        sheet_by_ticker[s["ticker"]]["units"]    += s["units"]
        sheet_by_ticker[s["ticker"]]["invested"] += s["invested"]
        sheet_total_invested += s["invested"]

    # ── Report ───────────────────────────────────────────────────────────────
    log("")
    log("── SUMMARY ──────────────────────────────────────────────")
    log(f"  API:   {len(api_rows)} trades across {len(api_by_ticker)} tickers | "
        f"invested ${api_total_invested:,.2f} | net P&L ${api_total_pnl:,.2f}")
    log(f"  Sheet: {len(sheet_rows)} rows across {len(sheet_by_ticker)} tickers | "
        f"invested ${sheet_total_invested:,.2f}")

    api_tickers   = set(api_by_ticker)
    sheet_tickers = set(sheet_by_ticker)
    only_api    = sorted(api_tickers - sheet_tickers)
    only_sheet  = sorted(sheet_tickers - api_tickers)
    both        = sorted(api_tickers & sheet_tickers)

    if only_api:
        log("")
        log("── On API, missing from sheet ───────────────────────────")
        for t in only_api:
            g = api_by_ticker[t]
            log(f"  {t:<10} units={g['units']:>10.4f}  invested=${g['invested']:>9,.2f}  "
                f"pnl=${g['net_pnl']:>9,.2f}  last={g['last_close']}  trades={g['count']}")
    if only_sheet:
        log("")
        log("── On sheet, no matching API ticker ─────────────────────")
        for t in only_sheet:
            g = sheet_by_ticker[t]
            log(f"  {t:<10} units={g['units']:>10.4f}  invested=${g['invested']:>9,.2f}")

    mismatches = []
    TOL = 1.0  # $1 tolerance on invested
    for t in both:
        a = api_by_ticker[t]
        s = sheet_by_ticker[t]
        if abs(a["invested"] - s["invested"]) > TOL or abs(a["units"] - s["units"]) > 0.01:
            mismatches.append((t, a, s))
    if mismatches:
        log("")
        log("── Ticker present in both, but values differ ────────────")
        for t, a, s in mismatches:
            log(f"  {t:<10} API units={a['units']:>10.4f} $={a['invested']:>9,.2f}  |  "
                f"sheet units={s['units']:>10.4f} $={s['invested']:>9,.2f}")

    # ── Dump full API view to CSV ────────────────────────────────────────────
    OUT_CSV.parent.mkdir(exist_ok=True)
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["ticker", "asset_id", "units", "invested_usd",
                    "open_rate", "close_rate", "net_pnl_usd",
                    "open_date", "close_date"])
        for r in sorted(api_rows, key=lambda x: (x["ticker"], x["close_date"])):
            w.writerow([r["ticker"], r["asset_id"], r["units"], r["invested"],
                        r["open_rate"], r["close_rate"], r["net_pnl"],
                        r["open_date"], r["close_date"]])
    log("")
    log(f"  Full API view written to {OUT_CSV}")


if __name__ == "__main__":
    main()
