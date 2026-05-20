"""Export eToro_Master.xlsx → markdown pages in the vault + a JSON cache.

Purpose:
  - Obsidian gets a clean, human-readable snapshot of every sheet.
  - Scheduled tasks can read the JSON cache instead of the xlsx directly,
    avoiding Excel file-lock headaches.

Triggered automatically at the end of run_all.py (after valuations + dashboard
generation), so the vault + JSON are always in sync with the latest prices.

Outputs:
  Vault/Personal/Finance/eToro/Portfolio.md
  Vault/Personal/Finance/eToro/Watchlist.md
  Vault/Personal/Finance/eToro/Closed Positions.md
  Vault/Personal/Finance/eToro/Tickers.md
  Vault/Personal/Finance/eToro/Daily Summary.md
  Vault/Personal/Finance/eToro/Assumptions.md
  data/etoro_master.json   (structured cache for dashboards)
"""
from __future__ import annotations

import json
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing openpyxl. Install with:  pip install openpyxl")

BASE = Path(__file__).parent.parent
DATA_DIR = BASE / "data"
XLSX_PATH = DATA_DIR / "eToro_Master.xlsx"
VAULT_ROOT = Path(os.environ.get("VAULT_ROOT", r"C:\Users\Neil\My Drive\Daley's Brain"))
VAULT_DIR = VAULT_ROOT / "Personal" / "Finance" / "eToro"

# (sheet_name, json_key, header_row_1based)
# Daily Summary has a multi-section visual layout and is skipped — open the xlsx.
SHEETS_TO_EXPORT = [
    ("Portfolio",        "portfolio",         2),
    ("Watchlist",        "watchlist",         2),
    ("Closed Positions", "closed_positions",  2),
    ("Tickers",          "tickers",           2),
    ("Assumptions",      "assumptions",       1),
]


def _load_env():
    env = BASE / "etoro.env"
    if not env.exists():
        return
    for line in env.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        os.environ.setdefault(k.strip(), v.strip())


def _cell(value) -> str:
    """Format a cell value for markdown output."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return f"{int(value)}"
        return f"{value:.4f}".rstrip("0").rstrip(".")
    s = str(value).replace("\n", " ").replace("|", "\\|").strip()
    return s


def _sheet_to_rows(ws, header_row: int) -> tuple[list[str], list[list[str]]]:
    """Extract header + data rows from a worksheet starting at header_row (1-based)."""
    rows = list(ws.iter_rows(values_only=True))
    if header_row - 1 >= len(rows):
        return [], []
    raw_header = rows[header_row - 1]
    # Trim trailing Nones in header
    while raw_header and raw_header[-1] in (None, ""):
        raw_header = raw_header[:-1]
    header = [_cell(c) for c in raw_header]
    n = len(header)
    data = []
    for r in rows[header_row:]:
        if not any(c not in (None, "") for c in r[:n]):
            continue
        data.append([_cell(r[i] if i < len(r) else None) for i in range(n)])
    return header, data


def _render_markdown(title: str, header: list[str], data: list[list[str]], generated: str) -> str:
    yaml = (
        "---\n"
        f"date: {datetime.now().strftime('%Y-%m-%d')}\n"
        "tags:\n"
        "  - etoro\n"
        "  - auto-generated\n"
        f"source: eToro_Master.xlsx\n"
        f"generated: {generated}\n"
        "---\n\n"
    )
    body = [yaml, f"# {title}\n\n", f"> Auto-exported from `eToro_Master.xlsx`. Do not edit by hand — changes will be overwritten.\n\n"]
    body.append(f"**{len(data)} rows · updated {generated}**\n\n")
    if not header:
        body.append("_No data._\n")
        return "".join(body)
    body.append("| " + " | ".join(header) + " |\n")
    body.append("| " + " | ".join("---" for _ in header) + " |\n")
    for row in data:
        body.append("| " + " | ".join(row) + " |\n")
    return "".join(body)


def _parse_assumptions(ws) -> dict:
    """Parse the Assumptions sheet into two parts:

        rates       -- simple key/value pairs (e.g. GBP/USD)
        valuations  -- per-ticker valuation table with name-keyed row objects

    The sheet layout is a banner (rows 1-2), a key/value pair (rows 3-4),
    a "PER-TICKER VALUATIONS" header (row 5), a column header (row 6), then
    ticker rows from row 7 onwards.
    """
    rows = list(ws.iter_rows(values_only=True))
    rates: dict = {}
    valuations: list[dict] = []
    val_headers: list[str] = []

    i, n = 0, len(rows)
    in_val_section = False
    while i < n:
        row = rows[i] or ()
        first = _cell(row[0]) if row else ""
        if not in_val_section and first.upper().startswith("PER-TICKER"):
            in_val_section = True
            # next non-empty row is the header
            j = i + 1
            while j < n and (not rows[j] or not any(c for c in rows[j] if c is not None)):
                j += 1
            if j < n:
                raw_hdr = rows[j]
                while raw_hdr and raw_hdr[-1] in (None, ""):
                    raw_hdr = raw_hdr[:-1]
                val_headers = [_cell(c) for c in raw_hdr]
                i = j + 1
                continue
        if in_val_section and val_headers:
            # data row in valuations table
            if row and any(c is not None and c != "" for c in row[:len(val_headers)]):
                obj = {val_headers[k]: (row[k] if k < len(row) else None) for k in range(len(val_headers))}
                # drop rows with no ticker
                if str(obj.get(val_headers[0]) or "").strip():
                    valuations.append(obj)
        elif not in_val_section and row:
            key = _cell(row[0])
            if key and not key.startswith("#"):
                val = row[1] if len(row) >= 2 else None
                if val is not None:
                    rates[key] = val if not isinstance(val, float) else round(val, 6)
        i += 1

    # Strip banner/section labels from rates (keep only plausible key/val pairs).
    rates.pop("VALUATION ASSUMPTIONS", None)
    rates.pop("GBP/USD Rate", None)
    return {"rates": rates, "valuations": valuations, "headers": val_headers}


def main() -> None:
    _load_env()
    if not XLSX_PATH.exists():
        sys.exit(f"Missing xlsx at {XLSX_PATH}")

    VAULT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(str(XLSX_PATH), read_only=True, data_only=True)
    generated = datetime.now().astimezone().isoformat(timespec="seconds")

    cache: dict = {
        "generated_at": generated,
        "sheets": {},
        "assumptions": {},
    }

    for sheet_name, key, header_row in SHEETS_TO_EXPORT:
        if sheet_name not in wb.sheetnames:
            print(f"  skip: {sheet_name} not found")
            continue
        ws = wb[sheet_name]

        if sheet_name == "Assumptions":
            parsed = _parse_assumptions(ws)
            cache["assumptions"] = {
                "rates":      parsed["rates"],
                "valuations": [
                    {k: (v if not isinstance(v, datetime) else v.strftime("%Y-%m-%d"))
                     for k, v in obj.items()}
                    for obj in parsed["valuations"]
                ],
            }
            # Markdown: two sections — Rates (key/value) then the Valuations table.
            rates_items = list(parsed["rates"].items())
            rates_md = "| Key | Value |\n| --- | --- |\n" + "".join(
                f"| {k} | {_cell(v)} |\n" for k, v in rates_items
            )
            val_headers = parsed["headers"]
            val_rows = [[_cell(obj.get(h)) for h in val_headers] for obj in parsed["valuations"]]
            val_md = ""
            if val_headers:
                val_md = "| " + " | ".join(val_headers) + " |\n"
                val_md += "| " + " | ".join("---" for _ in val_headers) + " |\n"
                for row in val_rows:
                    val_md += "| " + " | ".join(row) + " |\n"
            yaml = (
                "---\n"
                f"date: {datetime.now().strftime('%Y-%m-%d')}\n"
                "tags:\n  - etoro\n  - auto-generated\n"
                "source: eToro_Master.xlsx\n"
                f"generated: {generated}\n---\n\n"
            )
            md = (
                yaml + "# Assumptions\n\n"
                "> Auto-exported from `eToro_Master.xlsx`. Do not edit by hand — changes will be overwritten.\n\n"
                "## Rates\n\n" + rates_md + "\n"
                "## Per-Ticker Valuations\n\n" + val_md
            )
            data = parsed["valuations"]
        else:
            header, data = _sheet_to_rows(ws, header_row)
            # Emit name-keyed objects for easy JSON consumption in dashboards.
            objects = [
                {h: (row[i] if i < len(row) else "") for i, h in enumerate(header)}
                for row in data
            ]
            cache["sheets"][key] = {
                "headers":   header,
                "rows":      data,
                "objects":   objects,
                "row_count": len(data),
            }
            md = _render_markdown(sheet_name, header, data, generated)

        out_path = VAULT_DIR / f"{sheet_name}.md"
        out_path.write_text(md, encoding="utf-8")
        print(f"  wrote {out_path.name}  ({len(data) if sheet_name != 'Assumptions' else len(cache['assumptions'])} rows)")

    # JSON cache for dashboard consumers - atomic write:
    #   1. serialise to a string
    #   2. validate the string round-trips through json.loads (catches partial dumps)
    #   3. write to a .tmp file in the same directory
    #   4. os.replace the .tmp over the live file (atomic on POSIX & Windows)
    # This prevents the silent-truncation bug where a killed process leaves a
    # half-written JSON that build_site.py then "repairs" with stale data.
    import os as _os
    json_path = DATA_DIR / "etoro_master.json"
    tmp_path = json_path.with_suffix(".json.tmp")
    payload = json.dumps(cache, indent=2, default=str, ensure_ascii=False)
    try:
        json.loads(payload)  # sanity check before we touch the live file
    except json.JSONDecodeError as e:
        sys.exit(f"ERROR: built JSON failed self-validation ({e}). Live file left untouched at {json_path}.")
    tmp_path.write_text(payload, encoding="utf-8")
    # round-trip the file we just wrote, in case the disk wrote a short version
    try:
        with open(tmp_path, encoding="utf-8") as _f:
            json.load(_f)
    except json.JSONDecodeError as e:
        sys.exit(f"ERROR: tmp JSON at {tmp_path} failed re-parse ({e}). Live file left untouched.")
    _os.replace(str(tmp_path), str(json_path))
    print(f"  wrote {json_path}  ({len(payload):,} chars, atomically replaced)")


if __name__ == "__main__":
    main()
