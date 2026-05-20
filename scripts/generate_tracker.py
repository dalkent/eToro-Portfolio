#!/usr/bin/env python3
"""
generate_tracker.py
───────────────────
Generates the weekly FTSE Valuation Tracker article for Substack (paid tier).

Reads eToro_Master.xlsx for:
  - Assumptions sheet: blended targets, signals, model outputs
  - Portfolio sheet: current holdings
  - Tickers sheet: full FTSE universe

Fetches live prices from Yahoo Finance via yfinance.

Outputs:
  - Drafts/YYYY-MM-DD FTSE Valuation Tracker.md  (Substack-ready markdown)

Usage:
  python scripts/generate_tracker.py
  python scripts/generate_tracker.py --date 2026-04-08
  python scripts/generate_tracker.py --no-fetch   (skip Yahoo Finance, use placeholder prices)
"""

import os
import sys
import argparse
from pathlib import Path
from datetime import datetime, date
from collections import defaultdict

import openpyxl

BASE_DIR    = Path(__file__).parent.parent
DATA_DIR    = BASE_DIR / "data"
MASTER      = DATA_DIR / "eToro_Master.xlsx"
# Per-ticker computed_signal snapshots by date. Used to detect real week-over-week
# signal changes. The Assumptions sheet's prev_signal/curr_signal columns are static
# and never updated, so they cannot be used for week-on-week comparisons.
SIGNAL_HISTORY = DATA_DIR / "signal_history.json"

# Cross-platform path resolution via paths.py. VAULT_ROOT and SITE_REPO env
# vars still win if set explicitly.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from paths import VAULT_DIR as _PATHS_VAULT_DIR, SITE_REPO as _PATHS_SITE_REPO  # noqa: E402

VAULT_ROOT  = Path(os.environ.get("VAULT_ROOT", str(_PATHS_VAULT_DIR)))
DRAFTS_DIR  = VAULT_ROOT / "Projects" / "eToro & Investing" / "Drafts"

# Mirror the public-site filter from daleyvaluations-site/scripts/build_site.py
# so signal counts in the tracker match the website exactly.
EXCLUDED_TICKERS = {"PSH.L", "III.L"}

# Share the same price cache as daleyvaluations-site so the tracker and the
# website see IDENTICAL prices on any given run. The site's build_site.py also
# reads/writes this file with TTL=1h. Whichever runs first does the live fetch;
# the other consumes the cache.
SITE_REPO        = Path(os.environ.get("SITE_REPO", str(_PATHS_SITE_REPO)))
PRICE_CACHE_FILE = SITE_REPO / ".price_cache.json"
PRICE_CACHE_TTL  = 60  # minutes — must match build_site.py's PRICE_CACHE_TTL_HOURS=1

# ── Signal logic ─────────────────────────────────────────────────────────────

SIGNAL_ORDER = {
    "Strong Buy": 0, "Buy": 1, "Fair Value": 2, "Sell": 3, "Strong Sell": 4,
    "No Signal": 5, "N/A": 6, "": 6, None: 6,
}

SIGNAL_EMOJI = {
    "Strong Buy": "🟢", "Buy": "🟩", "Fair Value": "🟡",
    "Sell": "🔶", "Strong Sell": "🔴",
}

def compute_signal(vr):
    if vr is None:
        return "N/A"
    if vr >= 1.25:
        return "Strong Buy"
    if vr >= 1.10:
        return "Buy"
    if vr >= 0.90:
        return "Fair Value"
    if vr >= 0.75:
        return "Sell"
    return "Strong Sell"


# ── Load Excel ───────────────────────────────────────────────────────────────

def load_master():
    print(f"Reading {MASTER} ...")
    wb = openpyxl.load_workbook(str(MASTER), data_only=True)

    # GBP/USD rate
    ws_a = wb["Assumptions"]
    gbpusd = 1.34
    for row in ws_a.iter_rows(min_row=3, max_row=6, values_only=True):
        if row[0] == "GBP/USD" and row[1]:
            try:
                gbpusd = float(row[1])
            except Exception:
                pass
            break

    # Portfolio tickers
    ws_p = wb["Portfolio"]
    portfolio_tickers = set()
    for row in ws_p.iter_rows(min_row=3, max_row=300, values_only=True):
        yahoo = str(row[4] or "").strip()
        if yahoo and yahoo.endswith(".L"):
            portfolio_tickers.add(yahoo)

    # Tickers sheet — yahoo ticker, Market, Sector, Asset Type per row
    # (Market col G=index 6, Sector col H=index 7, Asset Type col I=index 8)
    ws_t = wb["Tickers"]
    ticker_meta: dict[str, dict] = {}
    for row in ws_t.iter_rows(min_row=2, max_row=400, values_only=True):
        etoro = str(row[3] or "").strip()
        yahoo = str(row[5] or "").strip()
        if not yahoo:
            continue
        meta = {
            "market":     str(row[6] or "").strip() if len(row) > 6 else "",
            "sector":     str(row[7] or "").strip() if len(row) > 7 else "",
            "asset_type": str(row[8] or "").strip() if len(row) > 8 else "",
        }
        # Index by both yahoo and eToro ticker — Assumptions sheet uses yahoo,
        # but be defensive in case there's a naming mismatch on a row.
        ticker_meta[yahoo] = meta
        if etoro and etoro != yahoo:
            ticker_meta.setdefault(etoro, meta)

    # Assumptions - all FTSE valuations
    stocks = []
    for row in ws_a.iter_rows(min_row=7, max_row=300, values_only=True):
        ticker = row[0]
        if not ticker or not isinstance(ticker, str) or not ticker.endswith(".L"):
            continue

        company     = str(row[1] or "").strip()
        sector      = str(row[2] or "").strip()
        beta        = float(row[3]) if row[3] is not None else None
        wacc        = float(row[4]) if row[4] is not None else None
        val1        = float(row[8]) if row[8] is not None else None
        val2        = float(row[9]) if row[9] is not None else None
        val3        = float(row[10]) if row[10] is not None else None
        blended_gbp = float(row[11]) if row[11] is not None else None
        model       = str(row[12] or "").strip()
        updated     = str(row[13] or "").strip()
        prev_signal = str(row[15] or "").strip()
        curr_signal = str(row[16] or "").strip()

        # Convert blended target from GBP to pence
        blended_p = round(blended_gbp * 100, 1) if blended_gbp else None

        in_portfolio = ticker in portfolio_tickers

        # Pull Tickers-sheet metadata for the public filter
        meta = ticker_meta.get(ticker, {})
        market     = meta.get("market", "")
        asset_type = meta.get("asset_type", "")
        # Prefer the Tickers-sheet sector (Assumptions sometimes shows fin sub-sector)
        sector_t   = meta.get("sector") or sector

        stocks.append({
            "ticker": ticker,
            "company": company,
            "sector": sector,
            "sector_t": sector_t,           # Tickers-sheet sector (used by filter)
            "market": market,
            "asset_type": asset_type,
            "beta": beta,
            "wacc": wacc,
            "val1": val1,
            "val2": val2,
            "val3": val3,
            "blended_gbp": blended_gbp,
            "blended_p": blended_p,
            "model": model,
            "updated": updated,
            "prev_signal": prev_signal,
            "curr_signal": curr_signal,
            "in_portfolio": in_portfolio,
            "live_price_p": None,  # populated by fetch
            "value_ratio": None,
            "computed_signal": None,
        })

    return stocks, gbpusd, portfolio_tickers


# Mirror site's filter_public() — exclude tickers the public website doesn't show
def filter_public(stocks):
    """Return the publishable subset matching daleyvaluations.com's filter_public().
    Six conditions, all must hold: market==FTSE, asset_type==Equity,
    sector!='Corp Bonds', current signal populated, model!='No Valuation',
    not in EXCLUDED_TICKERS."""
    out = []
    for s in stocks:
        if s.get("market") != "FTSE":              continue
        if s.get("asset_type") != "Equity":        continue
        if s.get("sector_t") == "Corp Bonds":      continue
        if not s.get("curr_signal"):               continue
        if (s.get("model") or "").strip().lower() == "no valuation":
            continue
        if s.get("ticker") in EXCLUDED_TICKERS:    continue
        out.append(s)
    return out


# ── Load via site's build_site.py (preferred) ───────────────────────────────
# This guarantees the tracker's universe + filters + prices match the public
# website exactly, since both consume the same loader/filter/price-cache.

def load_via_site():
    """Load via daleyvaluations-site/scripts/build_site.py for full consistency.
    Returns (stocks, gbpusd, portfolio_tickers) using the same dict shape that
    load_master returns, so the rest of the tracker code is unchanged.
    Returns (None, None, None) if the site repo isn't available (fall back to xlsx).
    """
    site_scripts = SITE_REPO / "scripts"
    if not site_scripts.exists():
        return None, None, None
    try:
        sys.path.insert(0, str(site_scripts))
        import build_site  # type: ignore
        # Force a fresh import in case of stale module cache
        import importlib
        importlib.reload(build_site)
    except Exception as e:
        print(f"  Could not import build_site ({e}) — falling back to xlsx loader")
        return None, None, None

    print(f"  Loading via daleyvaluations-site loader for full consistency …")
    try:
        data = build_site.load_data(build_site.DEFAULT_DATA_FILE)
        held = build_site.load_held_tickers(build_site.DEFAULT_PORTFOLIO_FILE)
        all_recs = build_site.join_records(data)
        public = build_site.filter_public(all_recs)
        # Use the SHARED cache-aware fetch (writes the cache the site reads).
        prices = build_site.fetch_live_prices(public, force_refresh=False)
        public = build_site.apply_live_prices(public, prices)
    except Exception as e:
        print(f"  build_site loader failed ({e}) — falling back to xlsx loader")
        return None, None, None

    # Adapt site records → tracker's expected dict shape.
    # Critically: copy the SITE's own value_ratio across so signal classification
    # uses the exact same number the website used (no precision drift from
    # re-rounding pence values before computing the ratio).
    stocks = []
    for r in public:
        tk = r["ticker"]
        blended_p = round(r["blended_target"] * 100, 1) if r.get("blended_target") else None
        live_price = r.get("live_price")
        live_price_p = round(float(live_price), 2) if live_price is not None else None
        # Site's value_ratio comes from apply_live_prices and is the authoritative
        # number used by the public site's signal counts.
        site_vr = r.get("value_ratio")
        stocks.append({
            "ticker": tk,
            "company": r.get("company") or "",
            "sector": r.get("sector") or "",
            "sector_t": r.get("sector") or "",
            "market": r.get("market") or "",
            "asset_type": r.get("asset_type") or "",
            "beta": r.get("beta"),
            "wacc": r.get("wacc"),
            "val1": r.get("val_dcf"),
            "val2": r.get("val_ddm"),
            "val3": r.get("val_epv"),
            "blended_gbp": r.get("blended_target"),
            "blended_p": blended_p,
            "model": r.get("model_method") or "",
            "updated": r.get("last_updated") or "",
            "prev_signal": r.get("prev_signal") or "",
            "curr_signal": r.get("current_signal") or "",
            "in_portfolio": (r.get("yahoo_ticker") or "").upper() in (held or set()),
            "live_price_p": live_price_p,
            # Keep both: raw (for classification) and pre-rounded (for display).
            "value_ratio_raw": site_vr,
            "value_ratio": round(site_vr, 3) if site_vr is not None else None,
            "computed_signal": None,
        })

    rates = (data.get("assumptions") or {}).get("rates") or {}
    try:
        gbpusd = float(rates.get("GBP/USD") or 1.34)
    except (ValueError, TypeError):
        gbpusd = 1.34

    print(f"  Loaded {len(stocks)} publishable FTSE equities via site loader (matches website universe).")
    return stocks, gbpusd, held


# ── Fetch live prices ────────────────────────────────────────────────────────

def _read_shared_price_cache():
    """Return {ticker: price} dict if the site's price cache is fresh, else None."""
    if not PRICE_CACHE_FILE.exists():
        return None
    try:
        import json as _json
        cache = _json.loads(PRICE_CACHE_FILE.read_text(encoding="utf-8"))
        cached_at = datetime.fromisoformat(cache.get("cached_at", "1970-01-01"))
        age_min = (datetime.now() - cached_at).total_seconds() / 60
        if age_min >= PRICE_CACHE_TTL:
            return None
        prices = cache.get("prices") or {}
        if not prices:
            return None
        print(f"  Using shared price cache from {PRICE_CACHE_FILE.parent.name}/.price_cache.json (age: {age_min:.1f} min, {len(prices)} prices)")
        return prices
    except Exception as e:  # noqa: BLE001
        print(f"  Cache read failed (non-fatal): {e}")
        return None


def _write_shared_price_cache(prices):
    """Write {ticker: price} to the shared cache so the site sees the same numbers."""
    try:
        import json as _json
        PRICE_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
        PRICE_CACHE_FILE.write_text(
            _json.dumps({"cached_at": datetime.now().isoformat(), "prices": prices}, indent=2),
            encoding="utf-8",
        )
    except Exception as e:  # noqa: BLE001
        print(f"  Cache write failed (non-fatal): {e}")


def fetch_prices(stocks):
    """Populate live_price_p (in pence) on each stock dict.
    Prefers the shared price cache (same one daleyvaluations-site uses); falls
    back to a live yfinance batch fetch if the cache is stale or missing.
    Writes the cache after a fresh fetch so the site sees identical prices.
    """
    candidates = [s for s in stocks if s["blended_p"] is not None]

    # 1. Try the shared cache
    cache = _read_shared_price_cache()
    if cache:
        applied = 0
        for s in candidates:
            v = cache.get(s["ticker"])
            if v is not None:
                s["live_price_p"] = round(float(v), 2)
                applied += 1
        if applied > 0:
            missing = len(candidates) - applied
            print(f"  Applied {applied}/{len(candidates)} prices from shared cache " +
                  (f"({missing} missing — will fetch individually)" if missing else ""))
            if missing == 0:
                return
            # Fall through to fill in the few missing ones individually.

    # 2. Live fetch (covers all tickers, or just the few missing from cache)
    try:
        import yfinance as yf
    except ImportError:
        print("Warning: yfinance not installed. Run: pip install yfinance")
        return

    targets = [s for s in candidates if s["live_price_p"] is None]
    print(f"  Fetching {len(targets)} prices live from Yahoo Finance ...")
    for s in targets:
        try:
            hist = yf.Ticker(s["ticker"]).history(period="2d")
            if not hist.empty:
                s["live_price_p"] = round(float(hist["Close"].iloc[-1]), 2)
        except Exception as e:  # noqa: BLE001
            print(f"  Failed to fetch {s['ticker']}: {e}")
    fetched = sum(1 for s in candidates if s["live_price_p"] is not None)
    print(f"  Total prices populated: {fetched}/{len(candidates)}")

    # 3. Update the shared cache so the site reads the same numbers next run
    if not cache and fetched > 0:
        prices = {s["ticker"]: s["live_price_p"] for s in candidates if s["live_price_p"] is not None}
        _write_shared_price_cache(prices)
        print(f"  Wrote {len(prices)} prices to shared cache ({PRICE_CACHE_FILE})")


def compute_signals(stocks):
    """Compute value ratios and signals from live prices and blended targets.

    Three concerns matter for site-tracker consistency:
    (1) When the record was loaded via load_via_site, prefer the SITE's exact
        value_ratio (computed by build_site.apply_live_prices) — it uses the
        unrounded target * 100 / pence-price and avoids precision drift.
    (2) Classify on the RAW value ratio (not the rounded one), matching site's
        signal_for() byte-for-byte.
    (3) Round only for DISPLAY, never before classification.
    """
    for s in stocks:
        # Path 1: site already supplied a raw value_ratio — trust it
        site_raw = s.get("value_ratio_raw")
        if site_raw is not None:
            s["computed_signal"] = compute_signal(site_raw)
            s["value_ratio"]     = round(site_raw, 3)
            continue
        # Path 2: xlsx fallback — compute from blended_p / live_price_p
        if s["live_price_p"] and s["blended_p"] and s["live_price_p"] > 0:
            raw_vr = s["blended_p"] / s["live_price_p"]
            s["computed_signal"] = compute_signal(raw_vr)
            s["value_ratio"]     = round(raw_vr, 3)
        else:
            s["computed_signal"] = s.get("curr_signal", "N/A")


# ── Markdown generation ──────────────────────────────────────────────────────

def fmt_price(p):
    if p is None:
        return "[PENDING]"
    if p >= 100:
        return f"{p:,.0f}p"
    return f"{p:.1f}p"

def fmt_vr(vr):
    if vr is None:
        return "[PENDING]"
    return f"{vr:.2f}"

def fmt_signal(sig):
    emoji = SIGNAL_EMOJI.get(sig, "")
    if emoji:
        return f"{emoji} {sig}"
    return sig or "N/A"


def _load_prior_snapshot(today: date):
    """Return (prior_date_str, {ticker: signal}) for the most recent snapshot
    strictly before `today`, or (None, {}) if none exists.
    """
    import json as _json
    if not SIGNAL_HISTORY.exists():
        return None, {}
    try:
        data = _json.loads(SIGNAL_HISTORY.read_text(encoding="utf-8"))
    except Exception as e:  # noqa: BLE001
        print(f"  WARNING: could not read signal_history.json ({e}) - no prior snapshot")
        return None, {}
    snaps = data.get("snapshots") or {}
    today_iso = today.isoformat()
    prior_dates = sorted([d for d in snaps.keys() if d < today_iso])
    if not prior_dates:
        return None, {}
    chosen = prior_dates[-1]
    return chosen, snaps[chosen]


def _save_snapshot(today: date, ticker_signals: dict):
    """Persist today's {ticker: computed_signal} snapshot into signal_history.json."""
    import json as _json
    data = {"snapshots": {}}
    if SIGNAL_HISTORY.exists():
        try:
            data = _json.loads(SIGNAL_HISTORY.read_text(encoding="utf-8"))
        except Exception:  # noqa: BLE001
            data = {"snapshots": {}}
    data.setdefault("snapshots", {})
    data["snapshots"][today.isoformat()] = ticker_signals
    SIGNAL_HISTORY.write_text(_json.dumps(data, indent=2, sort_keys=True), encoding="utf-8")


def generate_markdown(stocks, tracker_date):
    """Generate the full Substack article as markdown."""
    lines = []
    w = lines.append  # shorthand

    # Filter out investment trusts with "No Valuation" and absurd value ratios (data issues)
    valid = [s for s in stocks
             if s["computed_signal"] not in ("N/A", "No Signal", "", None)
             and s.get("model", "") not in ("No Valuation",)
             and (s["value_ratio"] is None or s["value_ratio"] < 10)]  # VR > 10 = data error
    portfolio = [s for s in valid if s["in_portfolio"]]
    non_portfolio = [s for s in valid if not s["in_portfolio"]]

    # Signal changes - compare today's computed_signal against the most recent
    # snapshot in signal_history.json (NOT the static prev/curr columns).
    prior_date, prior_signals = _load_prior_snapshot(tracker_date)
    changes = []
    if prior_signals:
        for s in valid:
            prev = prior_signals.get(s["ticker"])
            curr = s["computed_signal"]
            if prev and curr and prev != curr and prev not in ("No Signal", "N/A", ""):
                # Attach the prev signal so downstream rendering can use it
                s_copy = dict(s)
                s_copy["prev_signal_real"] = prev
                changes.append(s_copy)
        print(f"  Detected {len(changes)} signal change(s) vs snapshot {prior_date}")
    else:
        print(f"  No prior snapshot available - skipping signal-change detection")

    # Counts
    strong_buys = [s for s in valid if s["computed_signal"] == "Strong Buy"]
    strong_sells = [s for s in valid if s["computed_signal"] == "Strong Sell"]

    # Group portfolio by signal
    port_by_signal = defaultdict(list)
    for s in portfolio:
        port_by_signal[s["computed_signal"]].append(s)

    # Group non-portfolio strong signals
    np_strong_buy = [s for s in non_portfolio if s["computed_signal"] == "Strong Buy"]
    np_strong_sell = [s for s in non_portfolio if s["computed_signal"] == "Strong Sell"]

    # Sector heatmap
    sector_signals = defaultdict(lambda: defaultdict(int))
    for s in valid:
        sector_signals[s["sector"]][s["computed_signal"]] += 1

    # Cross-platform: %-d (Linux/Mac) and %#d (Windows) both strip the leading
    # zero, but neither works on the other OS. Format with leading zero, then
    # lstrip it. Result: "28 April 2026", "7 May 2026".
    date_str = tracker_date.strftime("%d %B %Y").lstrip("0")
    week_str = tracker_date.strftime("%Y-%m-%d")

    # ── Header ───────────────────────────────────────────────────────────
    w(f"# FTSE Valuation Tracker - Week of {date_str}")
    w("")
    w("*Updated every Tuesday. All valuations from my proprietary DCF/DDM/EPV models.*")
    w("")
    w("---")
    w("")

    # ── Section 1: Summary ───────────────────────────────────────────────
    w("## This Week at a Glance")
    w("")
    w(f"- **{len(changes)} signal change{'s' if len(changes) != 1 else ''}** this week")
    if changes:
        for c in changes:
            w(f"  - {c['company']} ({c['ticker']}): {c['prev_signal_real']} -> {c['computed_signal']}")
    w(f"- **{len(strong_buys)} Strong Buy** signals across the FTSE universe")
    w(f"- **{len(strong_sells)} Strong Sell** signals - names my models say are overvalued")
    w("")

    # ── Section 2: Signal changes ────────────────────────────────────────
    w("## Signal Changes This Week")
    w("")
    if changes:
        w("*Stocks where the signal moved from last week.*")
        w("")
        w("| Company | Ticker | Sector | Previous | New | Target | Live Price | VR |")
        w("|---|---|---|---|---|---|---|---|")
        for c in sorted(changes, key=lambda x: SIGNAL_ORDER.get(x["computed_signal"], 9)):
            w(f"| {c['company']} | {c['ticker']} | {c['sector']} | "
              f"{fmt_signal(c['prev_signal_real'])} | {fmt_signal(c['computed_signal'])} | "
              f"{fmt_price(c['blended_p'])} | {fmt_price(c['live_price_p'])} | {fmt_vr(c['value_ratio'])} |")
    else:
        w("No signal changes this week. All valuations stable at current prices.")
    w("")
    w("---")
    w("")

    # ── Section 3: Portfolio table ───────────────────────────────────────
    w("## My Portfolio - Current Signals")
    w("")
    w("*These are the FTSE stocks I hold in my live eToro portfolio. See all positions at "
      "[etoro.com/people/dalkent13](https://www.etoro.com/people/dalkent13).*")
    w("")

    signal_labels = ["Strong Buy", "Buy", "Fair Value", "Sell", "Strong Sell"]
    for sig in signal_labels:
        group = port_by_signal.get(sig, [])
        if not group:
            continue
        w(f"### {fmt_signal(sig)}")
        w("")
        w("| Company | Ticker | Sector | Target | Live Price | VR | Signal |")
        w("|---|---|---|---|---|---|---|")
        for s in sorted(group, key=lambda x: -(x["value_ratio"] or 0)):
            w(f"| {s['company']} | {s['ticker']} | {s['sector']} | "
              f"{fmt_price(s['blended_p'])} | {fmt_price(s['live_price_p'])} | "
              f"{fmt_vr(s['value_ratio'])} | {fmt_signal(s['computed_signal'])} |")
        w("")

    w("---")
    w("")

    # ── Section 4: Beyond portfolio ──────────────────────────────────────
    w("## Beyond My Portfolio - FTSE Strong Signals")
    w("")
    w("*FTSE stocks I don't currently hold where the models are flagging extreme valuations.*")
    w("")

    if np_strong_buy:
        w(f"### {fmt_signal('Strong Buy')} - Not in Portfolio")
        w("")
        w("| Company | Ticker | Sector | Target | Live Price | VR |")
        w("|---|---|---|---|---|---|")
        for s in sorted(np_strong_buy, key=lambda x: -(x["value_ratio"] or 0)):
            w(f"| {s['company']} | {s['ticker']} | {s['sector']} | "
              f"{fmt_price(s['blended_p'])} | {fmt_price(s['live_price_p'])} | {fmt_vr(s['value_ratio'])} |")
        w("")

    if np_strong_sell:
        w(f"### {fmt_signal('Strong Sell')} - Not in Portfolio")
        w("")
        w("| Company | Ticker | Sector | Target | Live Price | VR |")
        w("|---|---|---|---|---|---|")
        for s in sorted(np_strong_sell, key=lambda x: (x["value_ratio"] or 99)):
            w(f"| {s['company']} | {s['ticker']} | {s['sector']} | "
              f"{fmt_price(s['blended_p'])} | {fmt_price(s['live_price_p'])} | {fmt_vr(s['value_ratio'])} |")
        w("")

    w("---")
    w("")

    # ── Section 5: Sector heatmap ────────────────────────────────────────
    w("## Sector Heatmap")
    w("")
    w("| Sector | Strong Buy | Buy | Fair Value | Sell | Strong Sell |")
    w("|---|---|---|---|---|---|")
    for sector in sorted(sector_signals.keys()):
        counts = sector_signals[sector]
        w(f"| {sector} | {counts.get('Strong Buy', 0)} | {counts.get('Buy', 0)} | "
          f"{counts.get('Fair Value', 0)} | {counts.get('Sell', 0)} | {counts.get('Strong Sell', 0)} |")
    w("")

    # Cheapest / most expensive
    sector_buy_pct = {}
    for sector, counts in sector_signals.items():
        total = sum(counts.values())
        buys = counts.get("Strong Buy", 0) + counts.get("Buy", 0)
        if total > 0:
            sector_buy_pct[sector] = buys / total
    cheapest = sorted(sector_buy_pct.items(), key=lambda x: -x[1])[:2]
    most_exp_pct = {}
    for sector, counts in sector_signals.items():
        total = sum(counts.values())
        sells = counts.get("Strong Sell", 0) + counts.get("Sell", 0)
        if total > 0:
            most_exp_pct[sector] = sells / total
    expensive = sorted(most_exp_pct.items(), key=lambda x: -x[1])[:2]

    if cheapest:
        w(f"**Cheapest sectors:** {', '.join(s for s, _ in cheapest)}")
    if expensive:
        w(f"**Most expensive sectors:** {', '.join(s for s, _ in expensive)}")
    w("")
    w("---")
    w("")

    # ── Section 6: Approaching boundary ──────────────────────────────────
    w("## Approaching the Boundary")
    w("")
    w("*Stocks close to a signal change - value ratio within 5% of a threshold.*")
    w("")

    boundaries = [1.25, 1.10, 0.90, 0.75]
    boundary_names = {1.25: "Strong Buy/Buy", 1.10: "Buy/Fair Value",
                      0.90: "Fair Value/Sell", 0.75: "Sell/Strong Sell"}
    near_boundary = []
    for s in valid:
        vr = s["value_ratio"]
        if vr is None:
            continue
        for b in boundaries:
            if abs(vr - b) / b <= 0.05:
                direction = "upgrade" if vr < b else "downgrade"
                near_boundary.append({
                    **s,
                    "boundary": b,
                    "boundary_name": boundary_names[b],
                    "direction": direction,
                })
                break

    if near_boundary:
        w("| Company | Ticker | Signal | VR | Nearest Boundary | Direction |")
        w("|---|---|---|---|---|---|")
        for n in sorted(near_boundary, key=lambda x: abs(x["value_ratio"] - x["boundary"])):
            w(f"| {n['company']} | {n['ticker']} | {fmt_signal(n['computed_signal'])} | "
              f"{fmt_vr(n['value_ratio'])} | {n['boundary_name']} ({n['boundary']:.2f}) | "
              f"Potential {n['direction']} |")
    else:
        w("No stocks currently within 5% of a signal boundary.")
    w("")
    w("---")
    w("")

    # ── Methodology ──────────────────────────────────────────────────────
    w("## Methodology")
    w("")
    w("All valuations use my three-model framework: **DCF** (Discounted Cash Flow), "
      "**DDM** (Dividend Discount Model), and **EPV** (Earnings Power Value). "
      "Models are blended with sector-specific weights. Banks use DDM + P/B Excess Returns "
      "+ EPS Capitalisation (no DCF).")
    w("")
    w("**Current parameters:** UK risk-free rate 4.9% | Equity risk premium 5.0% | Terminal growth 2.5%")
    w("")
    w("Full methodology: [How to Value a Company](https://dalkent13.substack.com/p/how-to-value-a-company)")
    w("")
    w("---")
    w("")
    w("*Not financial advice. These are my personal views based on my own valuation models. "
      "Always do your own research before investing.*")
    w("")
    w("*Neil Daley - CFA Charterholder - "
      "[eToro](https://www.etoro.com/people/dalkent13) - "
      "[X/Twitter](https://x.com/Dalkent13)*")

    return "\n".join(lines)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate FTSE Valuation Tracker for Substack")
    parser.add_argument("--date", type=str, default=None,
                        help="Publication date (YYYY-MM-DD). Defaults to today.")
    parser.add_argument("--no-fetch", action="store_true",
                        help="Skip Yahoo Finance price fetch (use for testing)")
    args = parser.parse_args()

    tracker_date = datetime.strptime(args.date, "%Y-%m-%d").date() if args.date else date.today()

    # Preferred path: load + filter + fetch via the site's own build_site.py.
    # This guarantees the tracker's universe, signal mix, and per-ticker prices
    # match the public website (daleyvaluations.com) exactly, byte-for-byte.
    stocks, gbpusd, portfolio_tickers = (None, None, None)
    if not args.no_fetch:
        stocks, gbpusd, portfolio_tickers = load_via_site()
        if stocks is not None:
            print(f"Loaded {len(stocks)} FTSE stocks via site (matches daleyvaluations.com). GBP/USD: {gbpusd}")

    # Fallback path: read the xlsx directly (used when --no-fetch or site repo missing).
    if stocks is None:
        stocks, gbpusd, portfolio_tickers = load_master()
        print(f"Loaded {len(stocks)} FTSE stocks via xlsx fallback. {len(portfolio_tickers)} in portfolio. GBP/USD: {gbpusd}")
        before = len(stocks)
        stocks = filter_public(stocks)
        print(f"  Filtered to {len(stocks)} publishable FTSE equities (dropped {before - len(stocks)})")
        if args.no_fetch:
            print("Skipping price fetch (--no-fetch).")
        else:
            fetch_prices(stocks)

    compute_signals(stocks)

    md = generate_markdown(stocks, tracker_date)

    # Write output
    DRAFTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"{tracker_date.isoformat()} FTSE Valuation Tracker.md"
    output_path = DRAFTS_DIR / filename
    output_path.write_text(md, encoding="utf-8")
    print(f"\nTracker written to: {output_path}")

    # Persist today's computed_signal per ticker for next week's diff.
    # Use the SAME "valid" filter generate_markdown uses, so snapshots are clean.
    today_snapshot = {
        s["ticker"]: s["computed_signal"]
        for s in stocks
        if s.get("computed_signal") not in ("N/A", "No Signal", "", None)
        and s.get("model", "") not in ("No Valuation",)
        and (s.get("value_ratio") is None or s.get("value_ratio", 0) < 10)
    }
    _save_snapshot(tracker_date, today_snapshot)
    print(f"  Snapshot saved: {len(today_snapshot)} tickers -> signal_history.json")
    # Full signal-mix breakdown — should match daleyvaluations.com exactly when
    # loaded via load_via_site (the default path).
    from collections import Counter as _Counter
    mix = _Counter(s.get("computed_signal", "N/A") for s in stocks)
    print("  Signal mix:")
    for k in ("Strong Buy", "Buy", "Fair Value", "Sell", "Strong Sell", "N/A"):
        if mix.get(k, 0) or k != "N/A":
            print(f"    {k:13s}: {mix.get(k, 0)}")
    print(f"    {'TOTAL':13s}: {sum(mix.values())}")


if __name__ == "__main__":
    main()
