"""Backfill missing eToro Asset IDs in the Tickers sheet column E.

How it works
------------
1. Load all rows from Tickers sheet where col E (eToro Asset ID) is empty.
2. Fetch eToro's full instrument catalogue once
   (https://api.etorostatic.com/sapi/instrumentsmetadata/V1.1/instruments).
3. Build a SymbolFull -> InstrumentID lookup. SymbolFull matches the eToro
   Ticker (col D) directly: 'BRBY.L', 'GSK.L', 'PLTR', 'HOOD', etc.
   For known mismatches like BT.A.L -> BT.L, an alias map is consulted.
4. For each missing-ID row, find a match and write the InstrumentID into col E.
5. Report any tickers that had no match so they can be checked manually.

Usage
-----
    python scripts/backfill_asset_ids.py            # dry-run
    python scripts/backfill_asset_ids.py --apply
    python scripts/backfill_asset_ids.py --apply --backup
"""
from __future__ import annotations
import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import _envloader  # noqa: F401

import openpyxl
import requests

BASE_DIR = Path(__file__).resolve().parent.parent
MASTER = BASE_DIR / "data" / "eToro_Master.xlsx"
META_URL = "https://api.etorostatic.com/sapi/instrumentsmetadata/V1.1/instruments"

# Tickers col layout: A=#  B=Company  C=FTSE Ticker  D=eToro Ticker  E=eToro Asset ID  ...

# Known alias mappings: our local eToro Ticker (col D) -> eToro's catalogue SymbolFull.
# eToro's symbol convention sometimes differs from Yahoo / locally-recorded ticker form.
SYMBOL_ALIASES = {
    "BT.A.L": "BT.L",   # BT Group: locally "BT.A.L", eToro catalogue "BT.L"
}


def fetch_instruments(timeout=30, *, log=print):
    log(f"Fetching eToro instrument catalogue from {META_URL} ...")
    r = requests.get(META_URL, timeout=timeout)
    r.raise_for_status()
    items = r.json().get("InstrumentDisplayDatas", [])
    log(f"  Got {len(items)} instruments")
    return items


def build_symbol_lookup(items):
    """Return {SymbolFull: InstrumentID}. If duplicates exist, keep the lowest
    InstrumentID (oldest, more likely the canonical listing)."""
    lookup = {}
    for it in items:
        sym = it.get("SymbolFull")
        iid = it.get("InstrumentID")
        if not sym or iid is None:
            continue
        existing = lookup.get(sym)
        if existing is None or iid < existing:
            lookup[sym] = iid
    return lookup


def resolve_id(ticker, lookup):
    """Try direct match, then alias map. Returns (InstrumentID, alias_used) or (None, None)."""
    if ticker in lookup:
        return lookup[ticker], None
    alias = SYMBOL_ALIASES.get(ticker)
    if alias and alias in lookup:
        return lookup[alias], alias
    return None, None


def find_missing_id_rows(ws_t):
    """Yield (row_num, ticker, company) for Tickers rows missing col E."""
    for r in range(3, 300):
        ticker = ws_t.cell(row=r, column=4).value
        if not ticker or not isinstance(ticker, str):
            continue
        ticker = ticker.strip()
        if not ticker:
            continue
        asset_id = ws_t.cell(row=r, column=5).value
        if asset_id is not None:
            continue
        company = ws_t.cell(row=r, column=2).value or ""
        yield r, ticker, str(company)


def backfill(wb, *, instruments=None, log=print):
    """In-place. Returns (resolved, unresolved_tickers)."""
    if "Tickers" not in wb.sheetnames:
        log("  No Tickers sheet, skipping")
        return (0, [])
    ws_t = wb["Tickers"]
    missing = list(find_missing_id_rows(ws_t))
    if not missing:
        log("  No rows are missing an eToro Asset ID.")
        return (0, [])
    log(f"  Found {len(missing)} row(s) missing an eToro Asset ID.")

    if instruments is None:
        instruments = fetch_instruments(log=log)
    lookup = build_symbol_lookup(instruments)
    log(f"  Built lookup of {len(lookup)} SymbolFull -> InstrumentID entries.")

    resolved = 0
    unresolved = []
    for row_num, ticker, company in missing:
        iid, alias = resolve_id(ticker, lookup)
        if iid is None:
            unresolved.append(ticker)
            log(f"  Row {row_num:3d}: {ticker:<10} ({company!r}) - NO MATCH")
            continue
        ws_t.cell(row=row_num, column=5, value=int(iid))
        resolved += 1
        via = f" (via alias {alias})" if alias else ""
        log(f"  Row {row_num:3d}: {ticker:<10} ({company!r}) -> {iid}{via}")

    return (resolved, unresolved)


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--apply", action="store_true", help="Write changes back. Without this, runs as a dry-run.")
    p.add_argument("--backup", action="store_true", help="When applying, also write a timestamped backup.")
    p.add_argument("--master", default=str(MASTER), help="Path to eToro_Master.xlsx")
    args = p.parse_args()

    master = Path(args.master)
    if not master.exists():
        print(f"ERROR: {master} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {master} ...")
    wb = openpyxl.load_workbook(master)

    resolved, unresolved = backfill(wb)

    if args.apply and resolved:
        if args.backup:
            stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            backup = master.with_suffix(f".bak-{stamp}.xlsx")
            shutil.copy2(master, backup)
            print(f"\nWrote backup: {backup}")
        wb.save(master)
        print(f"Saved. Resolved {resolved} asset ID(s).")
        if unresolved:
            print(f"\n{len(unresolved)} ticker(s) had no match in eToro's catalogue:")
            print("  " + ", ".join(unresolved))
            print("These either don't exist on eToro, are delisted, or use a different symbol. Check manually.")
    elif resolved:
        print(f"\nDRY-RUN: would resolve {resolved} asset ID(s).")
        if unresolved:
            print(f"  {len(unresolved)} unmatched: {', '.join(unresolved)}")
        print("Re-run with --apply to commit.")
    else:
        if unresolved:
            print(f"\nNothing to write. {len(unresolved)} ticker(s) had no match: {', '.join(unresolved)}")
        else:
            print("\nNothing to do.")


if __name__ == "__main__":
    main()
