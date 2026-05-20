#!/usr/bin/env python3
"""
set_archetypes.py
─────────────────
One-shot script to write non-financial archetype assignments to the Sub-Sector
column (col O) of the Tickers sheet in eToro_Master.xlsx.

Default behaviour is a DRY RUN: prints a diff of current vs proposed values
and exits. Pass --commit to actually write.

Always saves a dated backup to data/backups/ before any write.

Usage:
  python scripts/set_archetypes.py            # dry run, prints diff
  python scripts/set_archetypes.py --commit   # writes after backup

Rolling back: copy the backup file back over eToro_Master.xlsx.
"""

import argparse
import shutil
import sys
from datetime import date
from pathlib import Path

import openpyxl

BASE_DIR    = Path(__file__).parent.parent
DATA_DIR    = BASE_DIR / "data"
MASTER      = DATA_DIR / "eToro_Master.xlsx"
BACKUP_DIR  = DATA_DIR / "backups"

# Non-financial archetype assignments for the 20 portfolio holdings classified
# in 2026-05-12 Archetype Assignment Proposal.md.
# FXPO and GLEN deferred pending data-quality investigation.
# NWG keeps its existing bank-sector logic.
ARCHETYPES = {
    # Mature compounder
    "RKT.L":  "Mature compounder",
    "DGE.L":  "Mature compounder",
    "BATS.L": "Mature compounder",
    # Mature cyclical
    "KGF.L":  "Mature cyclical",
    "JD.L":   "Mature cyclical",
    "MKS.L":  "Mature cyclical",
    "INCH.L": "Mature cyclical",
    "SBRY.L": "Mature cyclical",
    "EZJ.L":  "Mature cyclical",
    "VSVS.L": "Mature cyclical",
    # Yield anchor
    "NG.L":   "Yield anchor",
    "UU.L":   "Yield anchor",
    "IMB.L":  "Yield anchor",
    "TATE.L": "Yield anchor",
    # Capex cyclical
    "BP.L":   "Capex cyclical",
    "SHEL.L": "Capex cyclical",
    "ENOG.L": "Capex cyclical",
    "HBR.L":  "Capex cyclical",
    # Restructuring
    "VOD.L":  "Restructuring",
    # Non-portfolio explicit overrides — quality compounders that the
    # sector-default would misclassify. Added 2026-05-12 after the
    # sector-default rollout flagged HLMA/DPLM/RMV as Strong Sell under
    # Mature cyclical / Yield anchor defaults, when the businesses are
    # serial-acquirer compounders or software compounders. RR included on
    # Neil's judgement that Rolls-Royce is no longer in its recovery-cyclical
    # phase and is now best modelled as a quality compounder.
    "HLMA.L": "Mature compounder",
    "DPLM.L": "Mature compounder",
    "RMV.L":  "Mature compounder",
    "RR.L":   "Mature compounder",
}

# Sub-Sector column on the Tickers sheet (col O = 15, 1-indexed in openpyxl).
SUB_SECTOR_COL = 15
# Yahoo Finance Ticker column (col F = 6).
YF_TICKER_COL = 6


def load_sheet():
    if not MASTER.exists():
        print(f"ERROR: {MASTER} not found.", file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(str(MASTER), data_only=False)
    if "Tickers" not in wb.sheetnames:
        print("ERROR: Tickers sheet not found in workbook.", file=sys.stderr)
        sys.exit(1)
    return wb, wb["Tickers"]


def build_diff(ws):
    """Walk the Tickers sheet and return a list of (ticker, current, proposed, row)
    for every ticker in ARCHETYPES. Missing tickers are reported separately.
    """
    found = {}
    for row in range(2, ws.max_row + 1):
        yf_ticker = ws.cell(row=row, column=YF_TICKER_COL).value
        if not yf_ticker:
            continue
        yf_ticker = str(yf_ticker).strip()
        if yf_ticker in ARCHETYPES:
            current = ws.cell(row=row, column=SUB_SECTOR_COL).value
            current = str(current).strip() if current is not None else ""
            found[yf_ticker] = {
                "row": row,
                "current": current,
                "proposed": ARCHETYPES[yf_ticker],
            }
    missing = [t for t in ARCHETYPES if t not in found]
    return found, missing


def print_diff(found, missing):
    print()
    print(f"{'Ticker':<10} {'Row':<5} {'Current Sub-Sector':<30} {'Proposed':<25} {'Change'}")
    print("-" * 90)
    for ticker in sorted(found.keys()):
        info = found[ticker]
        cur = info["current"] or "(blank)"
        prop = info["proposed"]
        change = "WRITE" if cur != prop else "no-op"
        print(f"{ticker:<10} {info['row']:<5} {cur:<30} {prop:<25} {change}")
    print()
    if missing:
        print(f"WARNING: {len(missing)} ticker(s) not found in Tickers sheet:")
        for t in missing:
            print(f"  - {t}")
        print()
    n_writes = sum(1 for info in found.values() if info["current"] != info["proposed"])
    n_noop = sum(1 for info in found.values() if info["current"] == info["proposed"])
    print(f"Summary: {n_writes} writes, {n_noop} already correct, {len(missing)} missing.")


def take_backup():
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = date.today().isoformat()
    backup = BACKUP_DIR / f"eToro_Master_{stamp}.xlsx"
    # If a backup for today already exists, suffix with a counter so we never
    # overwrite an earlier same-day backup.
    counter = 1
    while backup.exists():
        backup = BACKUP_DIR / f"eToro_Master_{stamp}_{counter}.xlsx"
        counter += 1
    shutil.copy2(MASTER, backup)
    print(f"Backup saved: {backup}")
    return backup


def apply_writes(wb, ws, found):
    n = 0
    for ticker, info in found.items():
        if info["current"] == info["proposed"]:
            continue
        ws.cell(row=info["row"], column=SUB_SECTOR_COL).value = info["proposed"]
        n += 1
    wb.save(str(MASTER))
    return n


def main():
    parser = argparse.ArgumentParser(description=__doc__.strip().splitlines()[0])
    parser.add_argument("--commit", action="store_true",
                        help="Actually write the changes after taking a backup. "
                             "Without this flag the script only prints a diff.")
    args = parser.parse_args()

    wb, ws = load_sheet()
    found, missing = build_diff(ws)
    print_diff(found, missing)

    if not args.commit:
        print()
        print("DRY RUN — no changes made.")
        print("Re-run with --commit to apply the writes.")
        return 0

    n_writes = sum(1 for info in found.values() if info["current"] != info["proposed"])
    if n_writes == 0:
        print("Nothing to write. Exiting.")
        return 0

    print(f"Committing {n_writes} write(s) to {MASTER} …")
    backup = take_backup()
    n_written = apply_writes(wb, ws, found)
    print(f"Done. {n_written} cells updated on Tickers sheet, Sub-Sector column.")
    print(f"Rollback: copy {backup} over {MASTER} if needed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
