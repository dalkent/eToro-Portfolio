"""One-off and reusable fixer for orphan ID_NNNN rows in the Watchlist sheet.

What this fixes
---------------
Some Watchlist rows can end up with placeholder tickers like "ID_9272" because
they were added to the sheet before any mapping (Tickers sheet or
etoro_portfolio_tickermatch.csv) contained the eToro Asset ID. These rows
persist forever with no name/sector and clutter the watchlist.

Resolution sources, in order:
  1. Tickers sheet column E (eToro Asset ID) -> column D (eToro Ticker)
  2. etoro_portfolio_tickermatch.csv (Asset_ID,Ticker[,Ticker,Market,Asset])

If the CSV resolves a row that the Tickers sheet didn't, the missing row is
also appended to the Tickers sheet so the mapping is permanent and the rest
of the workflow (In Portfolio / In Watchlist flags) sees it.

For each resolved orphan:
  (a) delete the row, if the ticker is already held in the Portfolio sheet
  (b) otherwise rewrite ticker/yahoo/company/sector cells in place

Usage
-----
    python scripts/fix_watchlist_orphans.py            # dry-run
    python scripts/fix_watchlist_orphans.py --apply
    python scripts/fix_watchlist_orphans.py --apply --backup
"""
from __future__ import annotations
import argparse
import csv
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import _envloader  # noqa: F401

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
MASTER = BASE_DIR / "data" / "eToro_Master.xlsx"
MATCH_CSV = BASE_DIR / "data" / "etoro_portfolio_tickermatch.csv"


# Watchlist:    A=#  B=Company  C=Sector  D=eToro Ticker  E=Yahoo Ticker  F=Currency
# Tickers:      A=#  B=Company  C=FTSE Ticker  D=eToro Ticker  E=eToro Asset ID
#               F=Yahoo Ticker  G=Market  H=Sector  I=Asset Type  J=In Portfolio
#               K=In Watchlist  L=Notes  M=Microsoft Ticker (formula)  N=Manual Override
# Portfolio:    A=#  B=Company  C=Sector  D=eToro Ticker  ...


def _build_tickers_lookup(ws_t):
    """Return {numeric_id_str: dict of column values keyed by header} from Tickers sheet."""
    lookup = {}
    for row in ws_t.iter_rows(min_row=3, max_row=300, values_only=False):
        cell_id = row[4].value
        cell_ticker = row[3].value
        if cell_id is None or not cell_ticker:
            continue
        try:
            id_str = str(int(float(cell_id)))
        except (ValueError, TypeError):
            continue
        lookup[id_str] = {
            "ticker": str(cell_ticker).strip(),
            "company": str(row[1].value or "").strip(),
            "sector": str(row[7].value or "").strip(),
            "yahoo": str(row[5].value or cell_ticker).strip(),
            "market": str(row[6].value or "").strip(),
            "asset_type": str(row[8].value or "").strip(),
            "ftse_ticker": str(row[2].value or "").strip(),
        }
    return lookup


def _build_csv_lookup(csv_path):
    """Return {numeric_id_str: {ticker, market, asset_type}} from tickermatch.csv.

    The CSV has columns: Asset_ID, Ticker, Ticker, Market, Asset
    Note: the second 'Ticker' column overrides the first when read by DictReader,
    so row['Ticker'] gives us the *real* ticker (e.g. PLTR not 'ID 7991').
    """
    lookup = {}
    if not csv_path.exists():
        return lookup
    with open(csv_path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            try:
                id_str = str(int(row["Asset_ID"]))
            except (KeyError, ValueError):
                continue
            ticker = (row.get("Ticker") or "").strip()
            if not ticker or ticker.startswith("ID "):
                continue
            lookup[id_str] = {
                "ticker": ticker,
                "market": (row.get("Market") or "").strip(),
                "asset_type": (row.get("Asset") or "").strip(),
            }
    return lookup


def _portfolio_tickers(ws_p):
    held = set()
    for row in ws_p.iter_rows(min_row=3, max_row=200, values_only=False):
        ticker = row[3].value
        if not ticker:
            continue
        s = str(ticker).strip().upper()
        if s in ("", "CASH", "TOTAL", "GRAND TOTAL", "GRAND TOTAL (INCL. CASH)"):
            continue
        held.add(s)
    return held


def _last_data_row(ws, ticker_col_idx=3):
    last = 2
    for row in ws.iter_rows(min_row=3, max_row=300):
        v = row[ticker_col_idx].value
        if v and str(v).strip().upper() not in ("", "TICKER"):
            last = row[0].row
    return last


def _append_to_tickers(ws_t, *, ticker, company, sector, asset_id, yahoo,
                      market, asset_type, ftse_ticker, in_portfolio, in_watchlist,
                      log=print):
    """Append a new row to the Tickers sheet. Sets the Microsoft ticker formula."""
    next_r = _last_data_row(ws_t) + 1
    # Get current max # in column A
    max_num = 0
    for row in ws_t.iter_rows(min_row=3, max_row=next_r, values_only=False):
        if isinstance(row[0].value, (int, float)):
            max_num = max(max_num, int(row[0].value))
    new_num = max_num + 1

    ws_t.cell(row=next_r, column=1, value=new_num)
    ws_t.cell(row=next_r, column=2, value=company)
    ws_t.cell(row=next_r, column=3, value=ftse_ticker or None)
    ws_t.cell(row=next_r, column=4, value=ticker)
    ws_t.cell(row=next_r, column=5, value=int(asset_id))
    ws_t.cell(row=next_r, column=6, value=yahoo or ticker)
    ws_t.cell(row=next_r, column=7, value=market or "NASDQ")
    ws_t.cell(row=next_r, column=8, value=sector or "")
    ws_t.cell(row=next_r, column=9, value=asset_type or "Equity")
    ws_t.cell(row=next_r, column=10, value=in_portfolio)
    ws_t.cell(row=next_r, column=11, value=in_watchlist)
    # Notes (L) blank
    # Microsoft ticker formula in M (col 13) - same pattern as the existing rows
    ms_formula = (f'=IF(G{next_r}="FTSE","XLON:"&SUBSTITUTE(C{next_r},".L",""),'
                  f'IF(G{next_r}="NYSE","XNYS:"&D{next_r},'
                  f'IF(G{next_r}="NASDQ","XNAS:"&D{next_r},'
                  f'IF(G{next_r}="INT","CRYPTO:"&D{next_r},D{next_r}))))')
    ws_t.cell(row=next_r, column=13, value=ms_formula)
    log(f"  Tickers sheet: appended row {next_r} for {ticker} (asset_id={asset_id})")
    return next_r


def _repair_watchlist_formulas(ws_w):
    data_rows = []
    for row in ws_w.iter_rows(min_row=3, max_row=300):
        if row[3].value and str(row[3].value).strip().upper() not in ("", "TICKER"):
            data_rows.append(row[0].row)
    if not data_rows:
        return
    last_data = max(data_rows)
    for actual_row in data_rows:
        if actual_row == last_data:
            continue
        stale_ref = actual_row + 1
        cells = list(ws_w.iter_rows(min_row=actual_row, max_row=actual_row))[0]
        for cell in cells:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                old = cell.value
                new = re.sub(
                    r'(?<!\$)([A-Z]+)' + str(stale_ref) + r'(?!\d)',
                    lambda m: m.group(1) + str(actual_row),
                    old,
                )
                if new != old:
                    cell.value = new


def find_orphans(ws_w):
    """Return list of (row_num, raw_id_str, numeric_id_str)."""
    orphans = []
    for row in ws_w.iter_rows(min_row=3, max_row=300):
        ticker = row[3].value
        if not ticker:
            continue
        s = str(ticker).strip()
        m = re.match(r'^ID[_ ](\d+)$', s)
        if m:
            orphans.append((row[0].row, s, m.group(1)))
    return orphans


def fix_watchlist_orphans(wb, *, csv_path=MATCH_CSV, log=print):
    """In-place fix on the workbook.

    Returns (deleted, renamed, unresolved).
    """
    needed = ("Watchlist", "Tickers", "Portfolio")
    if not all(s in wb.sheetnames for s in needed):
        log("  fix_watchlist_orphans: required sheets missing, skipping")
        return (0, 0, 0)

    ws_w = wb["Watchlist"]
    ws_t = wb["Tickers"]
    ws_p = wb["Portfolio"]

    tickers_lookup = _build_tickers_lookup(ws_t)
    csv_lookup = _build_csv_lookup(Path(csv_path))
    held = _portfolio_tickers(ws_p)

    orphans = find_orphans(ws_w)
    if not orphans:
        return (0, 0, 0)

    # Bottom-up so deletions don't shift later rows
    orphans.sort(key=lambda x: -x[0])

    deleted = 0
    renamed = 0
    unresolved = 0
    any_deletion = False

    for row_num, raw_id, numeric_id in orphans:
        ticker = company = sector = yahoo = ""
        market = asset_type = ftse_ticker = ""
        in_tickers_already = numeric_id in tickers_lookup

        if in_tickers_already:
            t = tickers_lookup[numeric_id]
            ticker = t["ticker"]
            company = t["company"]
            sector = t["sector"]
            yahoo = t["yahoo"]
            market = t["market"]
            asset_type = t["asset_type"]
            ftse_ticker = t["ftse_ticker"]
        elif numeric_id in csv_lookup:
            c = csv_lookup[numeric_id]
            ticker = c["ticker"]
            yahoo = ticker  # NASDAQ/NYSE: same symbol
            market = c["market"]
            asset_type = c["asset_type"]
            # Try yfinance for company/sector if available
            try:
                import yfinance as yf
                info = yf.Ticker(ticker).info
                company = info.get("longName") or info.get("shortName") or ticker
                sector = info.get("sector") or ""
            except Exception:
                company = ticker  # fallback - just use the ticker as the name
            log(f"  Watchlist row {row_num}: {raw_id} resolved via CSV -> {ticker} ({company or '(no name)'})")
        else:
            log(f"  Watchlist row {row_num}: {raw_id} not resolvable (not in Tickers sheet or CSV) - leaving alone")
            unresolved += 1
            continue

        ticker_upper = ticker.upper()

        # Backfill the Tickers sheet if the row only came from CSV
        if not in_tickers_already:
            in_pf = "Yes" if ticker_upper in held else "No"
            in_wl = "No" if ticker_upper in held else "Yes"
            _append_to_tickers(
                ws_t,
                ticker=ticker, company=company, sector=sector, asset_id=numeric_id,
                yahoo=yahoo, market=market, asset_type=asset_type,
                ftse_ticker=ftse_ticker, in_portfolio=in_pf, in_watchlist=in_wl,
                log=log,
            )

        if ticker_upper in held:
            ws_w.delete_rows(row_num, 1)
            any_deletion = True
            deleted += 1
            log(f"  Watchlist row {row_num}: deleted (resolves to {ticker}, already in Portfolio)")
        else:
            row = list(ws_w.iter_rows(min_row=row_num, max_row=row_num))[0]
            row[3].value = ticker
            row[4].value = yahoo
            if not row[1].value and company:
                row[1].value = company
            if not row[2].value and sector:
                row[2].value = sector
            renamed += 1
            log(f"  Watchlist row {row_num}: renamed to {ticker} ({company or '(no name)'})")

    if any_deletion:
        _repair_watchlist_formulas(ws_w)

    return (deleted, renamed, unresolved)


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--apply", action="store_true", help="Write changes back. Without this, runs as a dry-run.")
    p.add_argument("--backup", action="store_true", help="When applying, also write a timestamped backup.")
    p.add_argument("--master", default=str(MASTER), help="Path to eToro_Master.xlsx")
    p.add_argument("--csv", default=str(MATCH_CSV), help="Path to etoro_portfolio_tickermatch.csv")
    args = p.parse_args()

    master = Path(args.master)
    if not master.exists():
        print(f"ERROR: {master} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {master} ...")
    wb = openpyxl.load_workbook(master)

    deleted, renamed, unresolved = fix_watchlist_orphans(wb, csv_path=Path(args.csv))

    actions = deleted + renamed
    if args.apply and actions:
        if args.backup:
            stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            backup = master.with_suffix(f".bak-{stamp}.xlsx")
            shutil.copy2(master, backup)
            print(f"Wrote backup: {backup}")
        wb.save(master)
        print(f"Saved. Deleted {deleted}, renamed {renamed}, unresolved {unresolved}.")
    elif actions:
        print(f"\nDRY-RUN: would delete {deleted}, rename {renamed}, leave {unresolved} unresolved. Re-run with --apply to commit.")
    elif unresolved:
        print(f"\nFound {unresolved} orphan(s) but none could be resolved. Add them to data/etoro_portfolio_tickermatch.csv and re-run.")
    else:
        print("No orphans found.")


if __name__ == "__main__":
    main()
